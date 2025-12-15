"""
Test completo del flujo TC: 455 -> plantilla5924
Ejecutar desde carpeta src/nuevo
"""
import sys
import os
from pathlib import Path

# Agregar parent path para imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook
from src.modelo.transformador import TransformadorDatos
from src.config.polizas import CONFIGURACION_POLIZAS

print("="*80)
print("TEST COMPLETO: Flujo TC (455 -> plantilla5924)")
print("="*80)

# Archivos en esta carpeta
archivo_455 = Path("455 Report_AseguradoraSaldos_Unificado noviembre 2025 5924.xlsx")
archivo_plantilla = Path("../plantillas/plantilla5924.xlsx")

print(f"\n✓ Archivo 455: {archivo_455}")
print(f"✓ Plantilla: {archivo_plantilla}")

if not archivo_455.exists():
    print(f"❌ ERROR: No existe {archivo_455}")
    sys.exit(1)

if not archivo_plantilla.exists():
    print(f"❌ ERROR: No existe {archivo_plantilla}")
    sys.exit(1)

# Obtener config TC
poliza_tc = CONFIGURACION_POLIZAS.get('TC')
if not poliza_tc:
    print("❌ ERROR: Configuración TC no encontrada")
    sys.exit(1)

print(f"\n✓ Configuración TC cargada")
print(f"  - Hoja origen: {poliza_tc.get('hoja_origen_requerida')}")
print(f"  - Número póliza: {poliza_tc.get('numero_poliza_fijo')}")
print(f"  - Nombre producto: {poliza_tc.get('nombre_producto_fijo')}")
print(f"  - País residencia: {poliza_tc.get('pais_residencia_fijo')}")

# Intentar transformación
print("\n" + "="*80)
print("EJECUTANDO TRANSFORMACIÓN")
print("="*80)

try:
    transformador = TransformadorDatos(callback_mensaje=lambda m: print(f"  {m}"))
    
    wb_resultado, nombre_descarga = transformador.transformar(
        archivo_origen=str(archivo_455),
        archivo_plantilla=str(archivo_plantilla),
        poliza_info=poliza_tc
    )
    
    print(f"\n✓ Transformación completada")
    print(f"  - Nombre: {nombre_descarga}")
    print(f"  - Workbook: {wb_resultado}")
    
    # Guardar temporal
    temp_file = f"resultado_tc_test.xlsx"
    wb_resultado.save(temp_file)
    print(f"\n✓ Guardado en: {temp_file}")
    
    # Verificar resultado
    print("\n" + "="*80)
    print("VERIFICACIÓN DE RESULTADO")
    print("="*80)
    
    wb_check = load_workbook(temp_file, data_only=True)
    ws_check = wb_check.active
    
    print(f"\n✓ Hojas: {wb_check.sheetnames}")
    print(f"✓ Hoja activa: {ws_check.title}")
    
    # Buscar headers y primeras filas
    print(f"\n📊 Primeros datos guardados:\n")
    
    # Encontrar fila de headers
    for fila_idx in range(1, min(15, ws_check.max_row + 1)):
        row_text = ' '.join([str(c.value or '').upper() for c in ws_check[fila_idx]])
        if "PRIMER APELLIDO" in row_text:
            print(f"  Headers en fila {fila_idx}")
            headers = [c.value for c in ws_check[fila_idx]]
            
            # Mostrar datos primera fila
            if fila_idx + 1 <= ws_check.max_row:
                fila_datos = ws_check[fila_idx + 1]
                print(f"\n  Primera fila de datos (fila {fila_idx + 1}):\n")
                
                for i, (h, cell) in enumerate(zip(headers[:20], fila_datos[:20])):
                    if h:
                        print(f"    {h:30s} = {cell.value}")
            break
    
    # Verificar campos fijos
    print(f"\n🔍 Verificación campos fijos TC:\n")
    
    campos_buscar = [
        ('NUMERO POLIZA', '5924'),
        ('NOMBRE PRODUCTO', 'SALDO DE DEUDA T + C'),
        ('PAIS DE RESIDENCIA', '239')
    ]
    
    for nombre_campo, valor_esperado in campos_buscar:
        encontrado = False
        for col_idx, h in enumerate(headers):
            if h and nombre_campo.upper() in str(h).upper():
                valor_real = fila_datos[col_idx].value
                status = "✓" if valor_real == valor_esperado else "⚠"
                print(f"  {status} {nombre_campo:30s}")
                print(f"     Esperado: {valor_esperado}")
                print(f"     Obtenido: {valor_real}\n")
                encontrado = True
                break
        
        if not encontrado:
            print(f"  ❌ {nombre_campo}: NO ENCONTRADO\n")
    
    print("="*80)
    print("✓ TEST COMPLETADO EXITOSAMENTE")
    print("="*80)
    
except Exception as e:
    import traceback
    print(f"\n❌ ERROR: {e}")
    print(traceback.format_exc())
    sys.exit(1)
