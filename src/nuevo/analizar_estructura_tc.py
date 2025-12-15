"""
Análisis de estructura TC: ver las primeras filas de datos para entender el mapeo
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

plantilla_tc = Path("plantillaTC.xlsx")

print("="*80)
print("ESTRUCTURA PLANTILLA TC - Primeras filas de datos")
print("="*80)

wb = load_workbook(plantilla_tc, data_only=True)
ws = wb['T+C (5924)']

print(f"\n📋 Hoja: T+C (5924)")
print(f"   Max filas: {ws.max_row}")
print(f"   Max columnas: {ws.max_column}")

# Encontrar encabezados
print(f"\n🔍 Buscando fila de encabezados...")
fila_headers = None
for fila_idx in range(1, min(15, ws.max_row + 1)):
    row_text = ' '.join([str(c.value or '').upper() for c in ws[fila_idx]])
    if "PRIMER APELLIDO" in row_text and "TIPO IDENTIFICACION" in row_text:
        fila_headers = fila_idx
        print(f"   ✓ Headers en fila {fila_idx}\n")
        break

if not fila_headers:
    print("   ❌ No se encontraron headers")
    sys.exit(1)

# Obtener headers
headers = [c.value for c in ws[fila_headers]]

# Mostrar estructura: headers + primeras 2 filas de datos
print("="*80)
print("FILA DE ENCABEZADOS (fila {})".format(fila_headers))
print("="*80)

print(f"\n{'Col':<4} {'Header':<35} {'Fila {}'.format(fila_headers+1):<25} {'Fila {}'.format(fila_headers+2):<25}")
print("-" * 90)

for col_idx, h in enumerate(headers):
    if not h:
        continue
    
    # Datos de las dos primeras filas
    data1 = ws.cell(fila_headers + 1, col_idx + 1).value
    data2 = ws.cell(fila_headers + 2, col_idx + 1).value
    
    h_str = str(h)[:33]
    d1_str = str(data1)[:23] if data1 else ""
    d2_str = str(data2)[:23] if data2 else ""
    
    print(f"{col_idx+1:<4} {h_str:<35} {d1_str:<25} {d2_str:<25}")

print("\n" + "="*80)
print("ANÁLISIS DE VALORES ESPECIALES")
print("="*80)

# Buscar campos especiales
campos_especiales = ['NUMERO POLIZA', 'NOMBRE PRODUCTO', 'PAIS DE RESIDENCIA', 'SALDO']

for campo in campos_especiales:
    for col_idx, h in enumerate(headers):
        if h and campo.upper() in str(h).upper():
            val1 = ws.cell(fila_headers + 1, col_idx + 1).value
            val2 = ws.cell(fila_headers + 2, col_idx + 1).value
            print(f"\n{campo} (Col {col_idx + 1}):")
            print(f"   Fila {fila_headers + 1}: {val1}")
            print(f"   Fila {fila_headers + 2}: {val2}")
            break

# Mostrar que columnas tienen fórmulas
print(f"\n\n" + "="*80)
print("COLUMNAS CON FÓRMULAS (fila {})".format(fila_headers+1))
print("="*80)

print(f"\n")
for col_idx, h in enumerate(headers):
    if not h:
        continue
    
    cell = ws.cell(fila_headers + 1, col_idx + 1)
    if cell.data_type == 'f' and cell.value:
        print(f"   Col {col_idx+1:<2d} ({h:<30s}): {cell.value}")

print("\n" + "="*80)
