import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Archivos a analizar
ORIGEN_455 = Path('src/nuevo/455 Report_AseguradoraSaldos_Unificado noviembre 2025 5924.xlsx')
PLANTILLA_TC = Path('src/plantillas/plantilla5924.xlsx')
PLANTILLA_DV = Path('src/plantillas/plantilla5852.xlsx')

print("=" * 80)
print("ANÁLISIS DETALLADO: 455 → plantillaTC vs flujo actual DV")
print("=" * 80)

# 1. Analizar origen 455
print("\n1. ARCHIVO ORIGEN 455:")
df_455 = pd.read_excel(ORIGEN_455, sheet_name='Report_AseguradoraSaldos_COVID', header=None)
headers_455 = df_455.iloc[5].tolist()
print(f"   Hoja: Report_AseguradoraSaldos_COVID")
print(f"   Fila headers: 5 (datos desde fila 6)")
print(f"   Total columnas con datos: {len([h for h in headers_455 if pd.notna(h)])}")

# 2. Analizar plantilla TC
print("\n2. PLANTILLA TC (destino):")
wb_tc = load_workbook(PLANTILLA_TC)
ws_tc = wb_tc['T+C (5924)']
headers_tc = [cell.value for cell in ws_tc[7]]  # Fila 7 (1-based)
print(f"   Hoja: T+C (5924)")
print(f"   Fila headers: 6 (0-based) = 7 (1-based)")
print(f"   Datos desde fila: 7 (0-based) = 8 (1-based)")
print(f"   Total columnas: {len([h for h in headers_tc if h])}")

# 3. Analizar plantilla DV para comparación
print("\n3. PLANTILLA DV (referencia actual):")
wb_dv = load_workbook(PLANTILLA_DV)
ws_dv = wb_dv.worksheets[0]
headers_dv = [cell.value for cell in ws_dv[5]]
print(f"   Fila headers: 4 (0-based) = 5 (1-based)")
print(f"   Datos desde fila: 5 (0-based) = 6 (1-based)")
print(f"   Total columnas: {len([h for h in headers_dv if h])}")

# 4. Comparar columnas TC vs DV
print("\n4. COLUMNAS EXCLUSIVAS EN TC (no en DV):")
set_tc = set([str(h).upper().strip() for h in headers_tc if h])
set_dv = set([str(h).upper().strip() for h in headers_dv if h])
exclusivas_tc = set_tc - set_dv
for col in sorted(exclusivas_tc):
    idx = [i for i, h in enumerate(headers_tc) if h and str(h).upper().strip() == col]
    if idx:
        print(f"   Col {idx[0]+1}: {col}")

# 5. Analizar fórmulas en plantilla TC
print("\n5. FÓRMULAS EN PLANTILLA TC (fila 8, primera fila de datos):")
fila_formula_tc = 8  # 1-based
formulas_tc = {}
for col_idx, cell in enumerate(ws_tc[fila_formula_tc], start=1):
    if cell.data_type == 'f' and cell.value:
        header = headers_tc[col_idx-1] if col_idx-1 < len(headers_tc) else f"Col{col_idx}"
        formulas_tc[col_idx] = (header, cell.value)
        print(f"   Col {col_idx:2d} ({header}): {cell.value}")

# 6. Analizar fórmulas en plantilla DV para comparar
print("\n6. FÓRMULAS EN PLANTILLA DV (fila 6, primera fila de datos):")
fila_formula_dv = 6
formulas_dv = {}
for col_idx, cell in enumerate(ws_dv[fila_formula_dv], start=1):
    if cell.data_type == 'f' and cell.value:
        header = headers_dv[col_idx-1] if col_idx-1 < len(headers_dv) else f"Col{col_idx}"
        formulas_dv[col_idx] = (header, cell.value)

print(f"   Total fórmulas DV: {len(formulas_dv)}")
print(f"   Total fórmulas TC: {len(formulas_tc)}")

# 7. Mapeo sugerido 455 → TC
print("\n7. MAPEO SUGERIDO 455 → TC:")
mapeo_sugerido = []
for idx_orig, h_orig in enumerate(headers_455):
    if pd.notna(h_orig):
        h_orig_str = str(h_orig).upper().strip()
        # Buscar coincidencia en TC
        for idx_dest, h_dest in enumerate(headers_tc):
            if h_dest:
                h_dest_str = str(h_dest).upper().strip()
                if h_orig_str == h_dest_str or h_orig_str in h_dest_str or h_dest_str in h_orig_str:
                    mapeo_sugerido.append((idx_orig+1, h_orig, idx_dest+1, h_dest))
                    break

print(f"   Total mapeos automáticos: {len(mapeo_sugerido)}")
for orig_col, orig_name, dest_col, dest_name in mapeo_sugerido[:10]:  # Primeros 10
    print(f"   455 Col {orig_col:2d} ({orig_name}) → TC Col {dest_col:2d} ({dest_name})")

# 8. Campos sin mapeo directo
print("\n8. CAMPOS TC SIN EQUIVALENTE DIRECTO EN 455:")
headers_455_upper = set([str(h).upper().strip() for h in headers_455 if pd.notna(h)])
campos_sin_mapeo = []
for idx, h in enumerate(headers_tc):
    if h:
        h_upper = str(h).upper().strip()
        tiene_mapeo = False
        for h_orig in headers_455_upper:
            if h_upper in h_orig or h_orig in h_upper or h_upper == h_orig:
                tiene_mapeo = True
                break
        if not tiene_mapeo:
            campos_sin_mapeo.append((idx+1, h))

for col_idx, col_name in campos_sin_mapeo:
    # Ver si tiene fórmula
    formula_info = formulas_tc.get(col_idx, None)
    if formula_info:
        print(f"   Col {col_idx:2d} ({col_name}): FÓRMULA - {formula_info[1]}")
    else:
        print(f"   Col {col_idx:2d} ({col_name}): SIN MAPEO (requiere valor fijo o lógica)")

print("\n" + "=" * 80)
print("Análisis completado. Revisar diferencias clave arriba.")
print("=" * 80)
