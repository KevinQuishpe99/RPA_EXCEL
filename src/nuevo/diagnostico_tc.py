"""
Diagnóstico para verificar mapeo de datos 455 -> TC
Ejecutar desde carpeta src/nuevo donde están los archivos
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Archivos esperados en esta carpeta
archivo_455 = Path("455 Report_AseguradoraSaldos_Unificado noviembre 2025 5924.xlsx")
plantilla_tc = Path("plantillaTC.xlsx")

print("="*80)
print("DIAGNÓSTICO: Mapeo 455 -> Plantilla TC")
print("="*80)

if not archivo_455.exists():
    print(f"❌ No se encontró: {archivo_455}")
    sys.exit(1)
print(f"✓ Archivo 455: {archivo_455}")

if not plantilla_tc.exists():
    print(f"❌ No se encontró: {plantilla_tc}")
    sys.exit(1)
print(f"✓ Plantilla TC: {plantilla_tc}")

print("\n" + "="*80)
print("ANÁLISIS ARCHIVO 455")
print("="*80)

# Leer archivo 455
xls_455 = pd.ExcelFile(archivo_455)
hojas_455 = xls_455.sheet_names
print(f"\n📋 Hojas: {hojas_455}")

if "Report_AseguradoraSaldos_COVID" not in hojas_455:
    print(f"❌ No existe 'Report_AseguradoraSaldos_COVID'")
    sys.exit(1)

df_455 = pd.read_excel(archivo_455, sheet_name="Report_AseguradoraSaldos_COVID", header=None)
print(f"📊 Datos: {len(df_455)} filas × {len(df_455.columns)} columnas")

# Encontrar encabezados (fila 5, 0-based)
print(f"\n🔍 Encabezados en fila 5 (0-based):")
headers_455 = [str(h).strip() if pd.notna(h) else "" for h in df_455.iloc[5].tolist()]
for i, h in enumerate(headers_455[:15]):
    if h:
        print(f"   Col {i+1:2d}: {h}")

print("\n" + "="*80)
print("ANÁLISIS PLANTILLA TC")
print("="*80)

wb_tc = load_workbook(plantilla_tc, data_only=False)
print(f"\n📋 Hojas: {wb_tc.sheetnames}")

# Encontrar hoja TC
hoja_tc = None
for sheet in wb_tc.sheetnames:
    if "5924" in sheet or "T+C" in sheet.upper():
        hoja_tc = sheet
        break

if not hoja_tc:
    hoja_tc = wb_tc.sheetnames[0]

print(f"✓ Usando hoja: {hoja_tc}")

ws_tc = wb_tc[hoja_tc]
print(f"📊 Estructura: {ws_tc.max_row} filas × {ws_tc.max_column} columnas")

# Encontrar encabezados (buscar "PRIMER APELLIDO")
print(f"\n🔍 Buscando encabezados...")
fila_headers = None
for fila_idx in range(1, min(15, ws_tc.max_row + 1)):
    row_text = ' '.join([str(c.value or '').upper() for c in ws_tc[fila_idx]])
    if "PRIMER APELLIDO" in row_text and "TIPO IDENTIFICACION" in row_text:
        fila_headers = fila_idx
        print(f"   ✓ Headers encontrados en fila {fila_idx}")
        headers_tc = [c.value for c in ws_tc[fila_idx]]
        break

if not fila_headers:
    print("   ❌ No se encontraron headers")
    sys.exit(1)

print(f"\n   Primeros 15 headers TC:")
for i, h in enumerate(headers_tc[:15]):
    if h:
        print(f"   Col {i+1:2d}: {h}")

print("\n" + "="*80)
print("MAPEO COLUMNAS 455 -> TC")
print("="*80)

mapeos = {}
for idx_455, h455 in enumerate(headers_455):
    if not h455:
        continue
    
    h455_upper = h455.upper()
    
    for idx_tc, htc in enumerate(headers_tc):
        if not htc:
            continue
        
        htc_upper = str(htc).upper()
        
        # Búsqueda de coincidencia
        if h455_upper == htc_upper or h455_upper in htc_upper or htc_upper in h455_upper:
            mapeos[idx_455] = (idx_tc + 1, h455, htc)
            print(f"\n455 Col {idx_455+1:2d} ({h455:35s})")
            print(f"  → TC Col {idx_tc+1:2d} ({htc})")
            break

print(f"\n✓ Total columnas mapeadas: {len(mapeos)}")

print("\n" + "="*80)
print("CAMPOS FIJOS REQUERIDOS")
print("="*80)

campos_fijos = {
    'NUMERO POLIZA': '5924',
    'NOMBRE PRODUCTO': 'SALDO DE DEUDA T + C',
    'PAIS DE RESIDENCIA': '239'
}

print(f"\nBuscando campos fijos en TC:\n")
for campo, valor in campos_fijos.items():
    encontrado = False
    for idx, h in enumerate(headers_tc):
        if h and campo.upper() in str(h).upper():
            print(f"✓ {campo:25s} → Col {idx+1:2d} ({h})")
            print(f"  Valor a establecer: {valor}")
            encontrado = True
            break
    if not encontrado:
        print(f"❌ {campo:25s} → NO ENCONTRADO")

print("\n" + "="*80)
print("PRIMEROS DATOS 455 (fila 6, primeras columnas)")
print("="*80)

datos_muestra = df_455.iloc[6, :15].tolist()
print(f"\n")
for i, (h, d) in enumerate(zip(headers_455[:15], datos_muestra)):
    if h:
        print(f"  {h:35s} = {str(d)[:50]}")

print("\n" + "="*80)
print("✓ Diagnóstico completado")
print("="*80)
