"""
Mapeo detallado: 455 -> TC
Comparación exacta de cómo deben pasar los datos
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

archivo_455 = Path("455 Report_AseguradoraSaldos_Unificado noviembre 2025 5924.xlsx")
plantilla_tc = Path("plantillaTC.xlsx")

print("="*80)
print("MAPEO DETALLADO: 455 -> TC")
print("="*80)

# Leer 455
df_455 = pd.read_excel(archivo_455, sheet_name="Report_AseguradoraSaldos_COVID", header=None)
headers_455 = [str(h).strip() if pd.notna(h) else "" for h in df_455.iloc[5].tolist()]
fila1_455 = df_455.iloc[6, :].tolist()  # Primera fila de datos
fila2_455 = df_455.iloc[7, :].tolist()  # Segunda fila de datos

# Leer TC
wb_tc = load_workbook(plantilla_tc, data_only=True)
ws_tc = wb_tc['T+C (5924)']
headers_tc = [c.value for c in ws_tc[7]]
fila1_tc = [ws_tc.cell(8, col_idx + 1).value for col_idx in range(len(headers_tc))]
fila2_tc = [ws_tc.cell(9, col_idx + 1).value for col_idx in range(len(headers_tc))]

print(f"\n📊 COMPARACIÓN FILA 1 (primeros datos):")
print("="*80)
print(f"\n{'455 Col':<6} {'Header 455':<30} {'Valor 455':<20} | {'TC Col':<6} {'Header TC':<30} {'Valor TC':<20}")
print("-"*160)

# Mapeo manual basado en la coincidencia de nombres
mapeo_manual = {
    1: (0, "PRIMER APELLIDO"),
    2: (1, "SEGUNDO APELLIDO"),
    3: (2, "PRIMER NOMBRE"),
    4: (3, "SEGUNDO NOMBRE"),
    5: (4, "OFICINA"),
    6: (5, "TIPO IDENTIFICACION"),
    7: (6, "NUMERO DE IDENTIFICACION"),
    8: (7, "FECHA DE NACIMIENTO"),
    9: (11, "SEXO/GENERO"),  # Col 12 en TC (10 es EDAD calculado)
    10: (12, "ESTADO CIVIL"),
    11: (13, "NACIONALIDAD"),
    12: (14, "PAIS DE RESIDENCIA"),  # Importante: de PAIS DE ORIGEN del 455
    13: (15, "PROVINCIA"),
    14: (16, "CIUDAD"),
    15: (24, "DIRECCION"),
    16: (25, "TELEFONO CASA"),
    17: (26, "TELEFONO TRABAJO"),
    18: (27, "CELULAR"),
    19: (29, "EMAIL"),
    20: (30, "OCUPACION"),
    21: (31, "ACTIVIDAD ECONOMICA"),
    22: (32, "INGRESOS"),
    23: (33, "PATRIMONIO"),
    24: (34, "SALDO A LA FECHA"),  # SALDO ACTUAL (col 29) del 455
    25: (36, "FECHA DE INICIO DE CREDITO"),
    26: (37, "FECHA DE TERMINACION DE CREDITO"),
    27: (38, "PLAZO DE CREDITO"),
    28: (39, "PRIMA NETA"),
    29: (40, "IMP"),
    30: (41, "PRIMA TOTAL"),
}

# Mostrar mapeos
for col_455, (idx_tc, header_tc) in mapeo_manual.items():
    if col_455 <= len(headers_455):
        h455 = headers_455[col_455 - 1] if col_455 <= len(headers_455) else ""
        v455 = str(fila1_455[col_455 - 1])[:18] if col_455 <= len(fila1_455) else ""
        v_tc = str(fila1_tc[idx_tc])[:18] if idx_tc < len(fila1_tc) else ""
        
        print(f"{col_455:<6} {h455[:29]:<30} {v455:<20} | {idx_tc+1:<6} {header_tc[:29]:<30} {v_tc:<20}")

print("\n" + "="*80)
print("📌 PUNTOS CLAVE DEL MAPEO:")
print("="*80)

print(f"""
1. COLUMNAS QUE SE SALTAN en 455:
   - Col 1: Vacía
   - Col 10: SEXO/GENERO (fila 10 en TC tiene EDAD calculado)
   - Algunos campos sin nombre en 455

2. CAMPOS FIJOS que deben escribirse SIEMPRE:
   - TC Col 15 (PAIS DE RESIDENCIA) = "239" 
   - TC Col 57 (NUMERO POLIZA) = "5924"
   - TC Col 58 (NOMBRE PRODUCTO) = "SALDO DE DEUDA T + C"

3. COLUMNAS CALCULADAS en TC que NO se deben sobrescribir:
   - TC Col 9 (EDAD en texto): Fórmula que calcula edad
   - TC Col 10 (EDAD en decimal): Fórmula que calcula edad
   - TC Col 11 (%): Porcentaje
   - TC Col 18 (CONCATENADO): Concatenación de pais/provincia/ciudad
   - TC Col 36 (SUMA ASEGURAR SDP): Copia de SALDO A LA FECHA

4. MAPEO ESPECIAL:
   - 455 Col 12 (SEXO/GENERO, posición) → TC Col 12 (SEXO/GENERO)
   - 455 Col 15 (PAIS DE ORIGEN) → TC Col 15 (PAIS DE RESIDENCIA) = "239" (override)
   - 455 Col 29 (SALDO ACTUAL) → TC Col 35 (SALDO A LA FECHA)

5. COLUMNAS QUE NO se mapean de 455:
   - TC Col 19 (PAIS nombre): Se llena desde tabla de códigos
   - TC Col 20 (PROVINCIA nombre): Se llena desde tabla de códigos
   - TC Col 21 (CIUDAD nombre): Se llena desde tabla de códigos
   - TC Col 22, 23, 24: Conversiones de códigos
   - TC Col 28 (DIRECCION TRABAJO): Requiere validación o está vacío
""")

print("="*80)
