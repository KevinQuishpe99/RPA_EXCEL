#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Script para debuguear los campos fijos TC"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook

# Add parent path to imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.polizas import CONFIGURACION_POLIZAS

print("=" * 80)
print("DEBUG: Campos fijos TC")
print("=" * 80)

# Rutas
ruta_plantilla = r"E:\s1\RPA_EXCEL\src\nuevo\plantillaTC.xlsx"

# Cargar plantilla y ver headers
print("\n[PLANTILLA] Cargando headers de plantilla TC...")
wb = load_workbook(ruta_plantilla, data_only=False)
ws = wb["T+C (5924)"]

print(f"  Hoja: {ws.title}")
print(f"  Headers en fila 7 (1-based):")
for col_idx in range(1, 60):  # 59 columnas
    cell = ws.cell(7, col_idx)
    header_text = cell.value or "EMPTY"
    print(f"    Col {col_idx:2d} (idx {col_idx-1:2d}): {repr(header_text)}")

# Verificar poliza_info
print("\n[CONFIG] Verificando poliza_info para TC...")
poliza_config = CONFIGURACION_POLIZAS['TC'].copy()
poliza_info = {
    'prefijo': poliza_config['prefijo'],
    'numero': poliza_config.get('numero_poliza_fijo'),
    'nombre': poliza_config.get('nombre_producto_fijo'),
    'hoja_origen_requerida': poliza_config.get('hoja_origen_requerida'),
    'pais_residencia': poliza_config.get('pais_residencia_fijo'),
    'numero_poliza_fijo': poliza_config.get('numero_poliza_fijo'),
    'nombre_producto_fijo': poliza_config.get('nombre_producto_fijo'),
    'pais_residencia_fijo': poliza_config.get('pais_residencia_fijo'),
}

print(f"  prefijo: {repr(poliza_info.get('prefijo'))}")
print(f"  numero_poliza_fijo: {repr(poliza_info.get('numero_poliza_fijo'))}")
print(f"  nombre_producto_fijo: {repr(poliza_info.get('nombre_producto_fijo'))}")
print(f"  pais_residencia_fijo: {repr(poliza_info.get('pais_residencia_fijo'))}")

# Buscar headers manualmente
print("\n[BUSQUEDA] Buscando headers en la plantilla...")
from openpyxl.cell.cell import MergedCell

headers_destino = list(ws.iter_rows(min_row=7, max_row=7, values_only=False))[0]
print(f"  Total headers: {len(headers_destino)}")

# Buscar NUMERO POLIZA
print("\n  Buscando 'NUMERO POLIZA'...")
for idx, cell in enumerate(headers_destino):
    if cell.value:
        header_upper = str(cell.value).upper()
        if 'NUMERO' in header_upper and 'POLIZA' in header_upper:
            print(f"    ENCONTRADO en columna {idx+1} (idx {idx}): {repr(cell.value)}")
            print(f"    Cell position: {cell.coordinate}")
            # Simular escritura
            print(f"    Escribiria en fila 8, columna {idx+1}")

# Buscar NOMBRE PRODUCTO
print("\n  Buscando 'NOMBRE PRODUCTO'...")
for idx, cell in enumerate(headers_destino):
    if cell.value:
        header_upper = str(cell.value).upper()
        if 'NOMBRE' in header_upper and 'PRODUCTO' in header_upper:
            print(f"    ENCONTRADO en columna {idx+1} (idx {idx}): {repr(cell.value)}")
            print(f"    Cell position: {cell.coordinate}")
            print(f"    Escribiria en fila 8, columna {idx+1}")

# Buscar PAIS DE RESIDENCIA
print("\n  Buscando 'PAIS DE RESIDENCIA'...")
for idx, cell in enumerate(headers_destino):
    if cell.value:
        header_upper = str(cell.value).upper()
        if 'PAIS' in header_upper and 'RESIDENCIA' in header_upper:
            print(f"    ENCONTRADO en columna {idx+1} (idx {idx}): {repr(cell.value)}")
            print(f"    Cell position: {cell.coordinate}")
            print(f"    Escribiria en fila 8, columna {idx+1}")

print("\n" + "=" * 80)
