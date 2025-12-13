# núcleo_transformacion.py
# Módulo principal que contiene la lógica de transformación de datos

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from datetime import datetime, date
import copy
import re


class TransformadorDatos:
    """
    Clase responsable de transformar datos de archivo origen a archivo destino.
    """
    
    def __init__(self, config_poliza, config_sistema):
        """
        Inicializa el transformador.
        
        Args:
            config_poliza: Información de la póliza a procesar
            config_sistema: Configuración general del sistema
        """
        self.config_poliza = config_poliza
        self.config_sistema = config_sistema
        
        # Cache para optimización
        self._cache_mapeo_columnas = None
        self._cache_headers_destino = None
        self._cache_indices_columnas = {}
        self._cache_estilos = {}
        
        # Crear estilos pre-compilados
        self._preparar_estilos()
    
    def _preparar_estilos(self):
        """Pre-crea estilos para reutilización"""
        self._cache_estilos['fuente_calibri'] = Font(name='Calibri')
        self._cache_estilos['alineacion_centrada'] = Alignment(
            horizontal='center', 
            vertical='center'
        )
        self._cache_estilos['borde_delgado'] = Side(style='thin')
        self._cache_estilos['borde_celda'] = Border(
            left=self._cache_estilos['borde_delgado'],
            right=self._cache_estilos['borde_delgado'],
            top=self._cache_estilos['borde_delgado'],
            bottom=self._cache_estilos['borde_delgado']
        )
    
    def limpiar_datos_destino(self, ws_destino):
        """
        Limpia datos existentes en la hoja destino, manteniendo encabezados.
        
        Args:
            ws_destino: Hoja de destino
        """
        # Limpiar datos desde fila 6 en adelante
        for row in ws_destino.iter_rows(min_row=6, max_row=ws_destino.max_row):
            for cell in row:
                if not isinstance(cell, MergedCell) and cell.data_type != 'f':
                    cell.value = None
    
    def validar_fila(self, fila_datos):
        """
        Valida si una fila debe ser procesada.
        
        Una fila es válida si su primera columna tiene valor y no es una palabra clave.
        
        Args:
            fila_datos: Array con datos de fila
            
        Returns:
            bool: True si fila es válida
        """
        if len(fila_datos) == 0:
            return False
        
        primera_col = fila_datos[0]
        
        # Validar si tiene valor
        if pd.isna(primera_col):
            return False
        
        if isinstance(primera_col, str):
            valor = primera_col.strip().upper()
            
            # Detectar palabras clave de totales
            if valor in ['NAN', 'NONE', 'NULL', 'TOTAL', 'CUADRE', 'PRECANCELACION', '']:
                return False
        
        return True
    
    def aplicar_transformaciones(self, valor, columna_destino, valor_tipo_ident=None):
        """
        Aplica transformaciones especiales a los valores.
        
        Args:
            valor: Valor original
            columna_destino: Nombre de columna destino
            valor_tipo_ident: Valor de TIPO IDENTIFICACION (si aplica)
            
        Returns:
            Valor transformado
        """
        from ..config import TRANSFORMACIONES
        
        if pd.isna(valor):
            return valor
        
        columna_upper = str(columna_destino).upper()
        
        # Quitar ceros iniciales de PROVINCIA y CIUDAD
        if 'PROVINCIA' in columna_upper or 'CIUDAD' in columna_upper:
            if isinstance(valor, str):
                # Convertir a número si es posible
                valor_limpio = valor.lstrip('0')
                return valor_limpio if valor_limpio else '0'
        
        # PAIS DE RESIDENCIA: siempre 239
        if 'PAIS' in columna_upper and 'RESIDENCIA' in columna_upper:
            return TRANSFORMACIONES['PAIS_RESIDENCIA_FIJO']
        
        # NACIONALIDAD: 239 si TIPO IDENTIFICACION es 00
        if 'NACIONALIDAD' in columna_upper:
            if valor_tipo_ident and str(valor_tipo_ident).strip() == '00':
                return TRANSFORMACIONES['NACIONALIDAD_SI_TIPO_00']
        
        return valor
    
    def obtener_mapeo_columnas(self, headers_origen, headers_destino):
        """
        Obtiene mapeo de índices de columnas origen → destino.
        
        Args:
            headers_origen: Headers de archivo origen
            headers_destino: Headers de archivo destino
            
        Returns:
            dict: {idx_origen: idx_destino, ...}
        """
        # Usar cache si disponible
        if self._cache_mapeo_columnas is not None:
            return self._cache_mapeo_columnas.copy()
        
        mapeo = {}
        
        # Crear diccionario de headers destino
        headers_destino_dict = {}
        for idx, cell in enumerate(headers_destino):
            if cell.value:
                nombre = str(cell.value).strip().upper()
                headers_destino_dict[nombre] = idx + 1
        
        # Mapear cada columna origen
        for idx_origen, header_origen in enumerate(headers_origen):
            if pd.isna(header_origen):
                continue
            
            nombre_origen = str(header_origen).strip().upper()
            
            # Búsqueda exacta
            if nombre_origen in headers_destino_dict:
                mapeo[idx_origen] = headers_destino_dict[nombre_origen]
                continue
            
            # Búsqueda parcial (flexible)
            for nombre_destino, idx_destino in headers_destino_dict.items():
                # Ignorar palabras pequeñas
                palabras_origen = [p for p in nombre_origen.split() if len(p) > 2]
                palabras_destino = [p for p in nombre_destino.split() if len(p) > 2]
                
                # Contar coincidencias
                coincidencias = sum(1 for p in palabras_origen if p in palabras_destino)
                
                if coincidencias >= 2:  # Al menos 2 palabras coinciden
                    mapeo[idx_origen] = idx_destino
                    break
        
        # Cachear resultado
        self._cache_mapeo_columnas = mapeo.copy()
        
        return mapeo
