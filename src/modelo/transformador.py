# src/modelo/transformador.py
"""
Modelo de Transformador - Orquestación principal
Refactorizado en arquitectura modular
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime

# Importar módulos especializados
from .estilos import EstilosExcel
from .mapeo_columnas import obtener_mapeo_columnas
from .mapeo_tc import MAPEO_TC_MANUAL, NO_SOBRESCRIBIR_TC, CAMPOS_FIJOS_TC
from .transferencia_datos import TransferenciaDatos
from .totales_pie import agregar_totales_columnas, agregar_pie_pagina, limpiar_bordes_todas_filas_excepto_pie
from .tabla_dinamica import crear_hoja2_tabla_dinamica


class TransformadorDatos:
    """Orquestador principal de transformación de datos"""
    
    def __init__(self, callback_mensaje=None):
        self.callback_mensaje = callback_mensaje
        self._cache_headers_destino = None
        self._cache_mapeo_columnas = None
        self._cache_indices_columnas = {}
        self._formulas_cache = {}
        self._formulas_pattern = re.compile(r'(\$?[A-Z]+\$?)(\d+)')
        self.estilos = EstilosExcel()
        self.transferencia = TransferenciaDatos(
            self.estilos,
            self._cache_indices_columnas,
            self._formulas_cache,
            self._formulas_pattern
        )
    
    def enviar_mensaje(self, mensaje):
        """Envía mensaje al callback si existe"""
        if self.callback_mensaje:
            self.callback_mensaje(mensaje)
    
    def transformar(self, archivo_origen, archivo_plantilla, poliza_info):
        """
        Transforma datos del archivo origen a la plantilla
        Retorna: (wb_resultado, nombre_archivo_descarga)
        """
        try:
            self.enviar_mensaje("=" * 80)
            self.enviar_mensaje("INICIANDO TRANSFORMACIÓN")
            self.enviar_mensaje("=" * 80)
            self.enviar_mensaje("Leyendo archivo origen...")
            
            # Leer archivo origen según póliza
            hoja_origen = None
            if poliza_info and isinstance(poliza_info, dict):
                hoja_origen = poliza_info.get('hoja_origen_requerida')
            if not hoja_origen:
                # Fallback compatible con DV (413)
                hoja_origen = "Report_AseguradoraMensual"

            df_origen = pd.read_excel(
                archivo_origen,
                sheet_name=hoja_origen,
                header=None,
                engine='openpyxl'
            )
            
            if len(df_origen) < 10:
                raise Exception(f"Archivo origen vacío o muy pequeño ({len(df_origen)} filas)")
            
            self.enviar_mensaje(f"✓ Archivo origen leído: {len(df_origen)} filas")
            
            # Buscar encabezados
            fila_encabezados_origen, headers_origen = self.buscar_encabezados(df_origen)
            
            if headers_origen is None:
                raise Exception("No se encontraron encabezados válidos")
            
            self.enviar_mensaje(f"✓ Encabezados encontrados en fila {fila_encabezados_origen + 1}")
            
            # Copiar plantilla
            wb = load_workbook(archivo_plantilla, data_only=False)
            
            # Detectar hoja destino
            hoja_destino = self.detectar_hoja_destino(wb, poliza_info)
            
            if not hoja_destino:
                raise Exception("No se pudo detectar hoja destino")
            
            ws = wb[hoja_destino]
            self.enviar_mensaje(f"✓ Usando hoja: {hoja_destino}")
            
            # Encontrar fila de encabezados destino (dinámicamente)
            fila_encabezados_destino = self.encontrar_fila_encabezados_destino(ws)
            fila_datos_destino = fila_encabezados_destino + 1
            
            # Obtener headers destino
            headers_destino = list(ws[fila_encabezados_destino])
            
            # Mapear columnas
            # Para TC, usar mapeo manual más preciso
            if poliza_info and poliza_info.get('prefijo') == 'TC':
                # Usar mapeo manual específico para TC
                mapeo = MAPEO_TC_MANUAL.copy()
                self.enviar_mensaje(f"✓ Usando mapeo manual TC")
            else:
                # Para DV, usar mapeo inteligente automático
                mapeo = obtener_mapeo_columnas(
                    headers_origen,
                    headers_destino,
                    self._cache_mapeo_columnas,
                    self._cache_headers_destino
                )
            
            # Actualizar cache
            self._cache_mapeo_columnas = mapeo.copy()
            self._cache_headers_destino = headers_destino
            
            self.enviar_mensaje(f"✓ {len(mapeo)} columnas mapeadas")
            
            # Limpiar datos existentes
            self.limpiar_datos_destino(ws, fila_inicio=fila_datos_destino)
            
            # Completar poliza_info con valores fijos para TC
            if poliza_info and poliza_info.get('prefijo') == 'TC':
                # Asegurar que tiene los campos esperados por _aplicar_campos_fijos_tc
                poliza_info = dict(poliza_info)  # Hacer copia para no modificar original
                if 'numero_poliza_fijo' not in poliza_info:
                    poliza_info['numero_poliza_fijo'] = poliza_info.get('numero')
                if 'nombre_producto_fijo' not in poliza_info:
                    poliza_info['nombre_producto_fijo'] = poliza_info.get('nombre')
                if 'pais_residencia_fijo' not in poliza_info:
                    poliza_info['pais_residencia_fijo'] = poliza_info.get('pais_residencia')
            
            # Transferir datos con parámetros dinámicos
            filas_procesadas = self.transferencia.transferir_datos(
                ws, df_origen, fila_encabezados_origen,
                headers_origen, headers_destino, mapeo, self.enviar_mensaje,
                fila_destino_inicio=fila_datos_destino,
                fila_plantilla=fila_datos_destino,
                poliza_info=poliza_info
            )
            
            self.enviar_mensaje(f"✓ {filas_procesadas} filas procesadas")
            
            # Agregar totales a columnas
            self.enviar_mensaje("Agregando totales a columnas...")
            ultima_fila_datos_nueva = filas_procesadas + 5
            fila_total = ultima_fila_datos_nueva + 1
            agregar_totales_columnas(ws, ultima_fila_datos_nueva, headers_destino, self.estilos, self.enviar_mensaje)
            
            # Agregar pie de página (deja una fila vacía después de totales)
            self.enviar_mensaje("Agregando pie de página...")
            agregar_pie_pagina(ws, fila_total, headers_destino, self.estilos, self.enviar_mensaje)
                        # Limpiar bordes de todas las filas después del pie
            self.enviar_mensaje("Limpiando bordes...")
            fila_final_pie = fila_total + 10  # Aproximadamente donde termina el pie
            limpiar_bordes_todas_filas_excepto_pie(ws, fila_final_pie, callback=self.enviar_mensaje)
                        # Crear Hoja2 con tabla dinámica
            self.enviar_mensaje("Creando Hoja2 con tabla dinámica...")
            crear_hoja2_tabla_dinamica(wb, ws, ultima_fila_datos_nueva, headers_destino, self.estilos, self.enviar_mensaje)
            
            # Generar nombre archivo
            fecha_mes = self.extraer_fecha_mes(df_origen, headers_origen)
            nombre_descarga = self.generar_nombre_archivo(poliza_info, fecha_mes)
            
            self.enviar_mensaje("✓ Transformación completada")
            
            return wb, nombre_descarga
            
        except Exception as e:
            raise Exception(f"Error en transformación: {str(e)}")
    
    def buscar_encabezados(self, df_origen):
        """Busca la fila de encabezados en el archivo origen"""
        for idx_fila in range(min(10, len(df_origen))):
            fila_actual = df_origen.iloc[idx_fila].tolist()
            headers_validos = [h for h in fila_actual if pd.notna(h) and str(h).strip() != '']
            
            if len(headers_validos) >= 5:
                headers_texto = [h for h in headers_validos if isinstance(h, str)]
                if len(headers_texto) >= 3:
                    return idx_fila, fila_actual
        
        return None, None
    
    def detectar_hoja_destino(self, wb, poliza_info):
        """Detecta la hoja destino basándose en la póliza"""
        for sheet_name in wb.sheetnames:
            if 'CODIGO' not in sheet_name.upper() and 'HOJA1' not in sheet_name.upper():
                if poliza_info:
                    prefijo = poliza_info.get('prefijo', '')
                    if prefijo in sheet_name.upper():
                        return sheet_name
                else:
                    return sheet_name
        
        return wb.sheetnames[0]
    
    def limpiar_datos_destino(self, ws, fila_inicio=6):
        """Limpia datos existentes en hoja destino"""
        for row in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row):
            for cell in row:
                if not isinstance(cell, MergedCell) and cell.data_type != 'f':
                    cell.value = None
    
    def extraer_fecha_mes(self, df_origen, headers_origen):
        """Extrae fecha del mes desde columna FECHA DE INICIO DE CREDITO"""
        col_fecha = None
        
        for idx, header in enumerate(headers_origen):
            if pd.notna(header) and 'FECHA DE INICIO DE CREDITO' in str(header).upper():
                col_fecha = idx
                break
        
        if col_fecha is None:
            return datetime.now()
        
        for idx in range(6, min(len(df_origen), 100)):
            try:
                fecha_valor = df_origen.iloc[idx, col_fecha]
                if pd.notna(fecha_valor):
                    if isinstance(fecha_valor, datetime):
                        return fecha_valor
                    elif isinstance(fecha_valor, pd.Timestamp):
                        return fecha_valor.to_pydatetime()
                    else:
                        return pd.to_datetime(fecha_valor).to_pydatetime()
            except:
                continue
        
        return datetime.now()
    
    def encontrar_fila_encabezados_destino(self, ws):
        """Encuentra la fila de encabezados en la plantilla destino"""
        for fila_idx in range(1, min(20, ws.max_row + 1)):
            row = ws[fila_idx]
            texto_fila = ' '.join([str(c.value or '').upper() for c in row])
            if 'PRIMER APELLIDO' in texto_fila and 'TIPO IDENTIFICACION' in texto_fila:
                return fila_idx
        return 5  # Por defecto fila 5 si no encuentra
    
    def generar_nombre_archivo(self, poliza_info, fecha_mes):
        """Genera el nombre del archivo resultado"""
        meses_espanol = {
            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
            5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
            9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
        }
        
        mes_nombre = meses_espanol.get(fecha_mes.month, 'mes').capitalize()
        año = fecha_mes.year
        
        prefijo = poliza_info.get('nombre_archivo', 'Facturación') if poliza_info else 'Facturación'
        
        return f"{prefijo} {mes_nombre} {año}.xlsx"
