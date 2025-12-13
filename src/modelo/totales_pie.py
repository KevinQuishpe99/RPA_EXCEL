# src/modelo/totales_pie.py
"""
Lógica para agregar totales y pie de página
"""

from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def agregar_totales_columnas(ws, ultima_fila_datos, headers_destino, estilos, callback=None):
    """Agrega fórmulas SUM para columnas - PLAZO DE CREDITO usa COUNTA (clientes) - SIN DUPLICAR"""
    columnas_totales = {
        'MONTO CREDITO': ['MONTO CREDITO', 'MONTO CRÉDITO'],
        'PLAZO DE CREDITO': ['PLAZO DE CREDITO', 'PLAZO DE CRÉDITO'],
        'PRIMA NETA': ['PRIMA NETA'],
        'HGR': ['HGR']
    }
    
    # Buscar columnas
    columnas_encontradas = {}
    for idx, cell in enumerate(headers_destino):
        if cell.value:
            header_str = str(cell.value).strip().upper()
            header_str_limpio = ' '.join(header_str.split())
            
            for col_nombre, variantes in columnas_totales.items():
                if col_nombre not in columnas_encontradas:
                    for variante in variantes:
                        variante_upper = variante.upper()
                        if variante_upper in header_str_limpio or header_str_limpio in variante_upper:
                            col_num = idx + 1
                            col_letter = get_column_letter(col_num)
                            columnas_encontradas[col_nombre] = (col_num, col_letter)
                            if callback:
                                callback(f"  → Columna '{col_nombre}' encontrada en {col_letter}")
                            break
                    if col_nombre in columnas_encontradas:
                        break
    
    # Fila de totales
    fila_total = ultima_fila_datos + 1
    formulas_agregadas = 0
    
    # Agregar SUM o COUNTA según columna
    for col_nombre, (col_num, col_letter) in columnas_encontradas.items():
        try:
            cell_total = ws.cell(fila_total, col_num)
            if isinstance(cell_total, MergedCell):
                continue
            
            # PLAZO DE CREDITO usa COUNTA (contar clientes), otros usan SUM
            if col_nombre == 'PLAZO DE CREDITO':
                formula = f"=COUNTA({col_letter}6:{col_letter}{ultima_fila_datos})"
                label_desc = "clientes (COUNTA)"
            else:
                formula = f"=SUM({col_letter}6:{col_letter}{ultima_fila_datos})"
                label_desc = "suma (SUM)"
            
            # Sobrescribir siempre con la fórmula correcta (incluso si tiene valor previo)
            cell_total.value = formula
            cell_total.font = estilos.fuente_calibri_negrita
            cell_total.alignment = estilos.alineacion_centrada
            cell_total.border = None  # SIN BORDES en fila de totales
            cell_total.number_format = '#,##0.00'
            cell_total.fill = estilos.fill_amarillo
            
            formulas_agregadas += 1
            if callback:
                callback(f"  ✓ {label_desc} agregada para '{col_nombre}' en {col_letter}{fila_total}")
            
            # Si es PLAZO DE CREDITO, agregar "CLIENTES" en columna anterior
            if col_nombre == 'PLAZO DE CREDITO' and col_num > 1:
                col_anterior = col_num - 1
                cell_clientes = ws.cell(fila_total, col_anterior)
                if not isinstance(cell_clientes, MergedCell):
                    cell_clientes.value = "CLIENTES"
                    cell_clientes.font = estilos.fuente_calibri_negrita
                    cell_clientes.alignment = estilos.alineacion_centrada
                    cell_clientes.border = None  # SIN BORDES en fila de totales
                    cell_clientes.fill = estilos.fill_amarillo
                    if callback:
                        callback(f"  ✓ Etiqueta 'CLIENTES' agregada en {get_column_letter(col_anterior)}{fila_total}")
        except Exception:
            continue
    
    # IMP y PRIMA TOTAL también van en fila_total (misma fila que otros totales)
    # Buscar columnas AL, AM, AN
    col_al = None  # PRIMA NETA
    col_am = None  # IMP
    col_an = None  # PRIMA TOTAL
    
    for idx, cell in enumerate(headers_destino):
        if cell.value:
            header_str = str(cell.value).strip().upper()
            header_str_limpio = ' '.join(header_str.split())
            
            if 'PRIMA NETA' in header_str_limpio and col_al is None:
                col_al = idx + 1
            elif 'IMP' in header_str_limpio and col_am is None:
                col_am = idx + 1
            elif 'PRIMA TOTAL' in header_str_limpio and col_an is None:
                col_an = idx + 1
    
    # AM (IMP): =AL{fila_total}*4% - EN LA MISMA FILA_TOTAL
    if col_am is not None and col_al is not None:
        try:
            col_letter_al = get_column_letter(col_al)
            col_letter_am = get_column_letter(col_am)
            cell_am = ws.cell(fila_total, col_am)
            
            if not isinstance(cell_am, MergedCell):
                formula_am = f"={col_letter_al}{fila_total}*4%"
                cell_am.value = formula_am
                cell_am.font = estilos.fuente_calibri_negrita
                cell_am.alignment = estilos.alineacion_centrada
                cell_am.border = None  # SIN BORDES en fila de totales
                cell_am.number_format = '#,##0.00'
                cell_am.fill = estilos.fill_amarillo
                if callback:
                    callback(f"  ✓ Fórmula IMP agregada en {col_letter_am}{fila_total}")
        except Exception:
            pass
    
    # AN (PRIMA TOTAL): =+AL{fila_total}+AM{fila_total} - EN LA MISMA FILA_TOTAL
    if col_an is not None and col_al is not None and col_am is not None:
        try:
            col_letter_al = get_column_letter(col_al)
            col_letter_am = get_column_letter(col_am)
            col_letter_an = get_column_letter(col_an)
            cell_an = ws.cell(fila_total, col_an)
            
            if not isinstance(cell_an, MergedCell):
                formula_an = f"=+{col_letter_al}{fila_total}+{col_letter_am}{fila_total}"
                cell_an.value = formula_an
                cell_an.font = estilos.fuente_calibri_negrita
                cell_an.alignment = estilos.alineacion_centrada
                cell_an.border = None  # SIN BORDES en fila de totales
                cell_an.number_format = '#,##0.00'
                cell_an.fill = estilos.fill_amarillo
                if callback:
                    callback(f"  ✓ Fórmula PRIMA TOTAL agregada en {col_letter_an}{fila_total}")
        except Exception:
            pass
    
    if formulas_agregadas > 0 and callback:
        callback(f"  ✓ {formulas_agregadas} total(es) agregado(s) en fila {fila_total}")


def limpiar_bordes_todas_filas_excepto_pie(ws, fila_pie_inicio, callback=None):
    """Quita TODOS los bordes de todas las filas después del pie de página"""
    try:
        from openpyxl.styles import Border
        no_border = Border()
        
        # Comenzar después del pie (al final del pie)
        fila_inicio = fila_pie_inicio + 1
        max_row = ws.max_row
        
        if fila_inicio <= max_row:
            for fila in range(fila_inicio, max_row + 1):
                for col in range(1, ws.max_column + 1):
                    try:
                        cell = ws.cell(fila, col)
                        cell.border = no_border
                    except Exception:
                        pass
            
            if callback:
                filas_limpiadas = max_row - fila_inicio + 1
                callback(f"  ✓ {filas_limpiadas} filas sin bordes después del pie de página")
    except Exception as e:
        if callback:
            callback(f"  ⚠️ Error al limpiar bordes: {str(e)}")


def limpiar_bordes_entre_filas(ws, fila_inicio, fila_fin, callback=None):
    """Quita TODOS los bordes entre fila_inicio y fila_fin inclusive.
    Útil para garantizar que después de los totales todas las filas queden sin bordes
    antes de construir el pie de página.
    """
    try:
        from openpyxl.styles import Border
        no_border = Border()

        max_row = ws.max_row
        max_col = ws.max_column

        fila_inicio = max(1, int(fila_inicio))
        fila_fin = min(max_row, int(fila_fin))

        if fila_inicio <= fila_fin:
            for fila in range(fila_inicio, fila_fin + 1):
                for col in range(1, max_col + 1):
                    try:
                        ws.cell(fila, col).border = no_border
                    except Exception:
                        pass
            if callback:
                callback(f"  ✓ Filas {fila_inicio}-{fila_fin} sin bordes")
    except Exception as e:
        if callback:
            callback(f"  ⚠️ Error limpiando bordes entre filas: {str(e)}")


def aplicar_formato_celda(cell, valor, estilos, aplicar_borde=True, numero_formato=None, fill=None, forzar_borde=False):
    """Aplica formato a una celda.
    - Bordes: se aplican si hay contenido real y aplicar_borde=True, o si forzar_borde=True.
    - Las celdas vacías quedan sin borde salvo que se fuerce (útil para celdas operandos vacías).
    """
    cell.value = valor
    cell.font = estilos.fuente_calibri
    cell.alignment = estilos.alineacion_centrada
    
    tiene_contenido = False
    if valor is not None:
        valor_str = str(valor).strip()
        if valor_str and valor_str != '' and valor_str != '-':
            tiene_contenido = True
    
    if (aplicar_borde and tiene_contenido) or forzar_borde:
        cell.border = estilos.borde_celda
    else:
        cell.border = None
    
    if numero_formato:
        cell.number_format = numero_formato
    if fill:
        cell.fill = fill


def agregar_pie_pagina(ws, fila_total, headers_destino, estilos, callback=None):
    """Agrega pie de página con PRE CANCELACION, BASE 0%, BASE 12%, tabla de pólizas
    SOLO aplica bordes a celdas con datos
    fila_total es la fila de totales, el pie comenzará en fila_total + 2 dejando una vacía"""
    try:
        # Columnas fijas para el pie (AN, AO, AP, AQ, AR)
        col_an = 40  # AN
        col_ao = 41  # AO
        col_ap = 42  # AP
        col_aq = 43  # AQ
        col_ar = 44  # AR
        
        col_letter_an = get_column_letter(col_an)
        col_letter_ao = get_column_letter(col_ao)
        col_letter_ap = get_column_letter(col_ap)
        col_letter_aq = get_column_letter(col_aq)
        col_letter_ar = get_column_letter(col_ar)
        
        # Buscar PRIMA NETA (AL) y HGR para las fórmulas
        col_al = None
        col_ao_hgr = None
        for idx, cell in enumerate(headers_destino):
            if cell.value:
                header_str = str(cell.value).strip().upper()
                if 'PRIMA NETA' in header_str and col_al is None:
                    col_al = idx + 1
                if 'HGR' in header_str and col_ao_hgr is None:
                    col_ao_hgr = idx + 1
        
        if col_al is None:
            col_al = 38  # Valor por defecto para AL
        
        if col_ao_hgr is None:
            col_ao_hgr = 41  # Valor por defecto para AO (HGR)
        
        col_letter_al = get_column_letter(col_al)
        col_letter_ao_hgr = get_column_letter(col_ao_hgr)
        
        # Limpiar bordes DESPUÉS de la fila de totales, antes de construir el pie
        # Se limpia desde la fila siguiente a totales hasta el final actual del libro
        limpiar_bordes_entre_filas(ws, fila_total + 1, ws.max_row, callback=callback)

        # Filas del pie - dejar una fila vacía después de totales
        fila_pre_cancelacion = fila_total + 2
        fila_formula_porcentaje = fila_pre_cancelacion - 1  # Una fila ARRIBA de PRE CANCELACION
        fila_base_12 = fila_pre_cancelacion + 3
        fila_tabla_encabezado = fila_base_12 + 3
        fila_poliza_1 = fila_tabla_encabezado + 1
        fila_poliza_2 = fila_poliza_1 + 1
        fila_totales_tabla = fila_poliza_2 + 1
        
        # Fila de fórmula porcentaje (AO/AL) - ARRIBA de PRE CANCELACION
        aplicar_formato_celda(
            ws.cell(fila_formula_porcentaje, col_ao), 
            f"={col_letter_ao}{fila_total}/{col_letter_al}{fila_total}", 
            estilos,
            aplicar_borde=True,  # SÍ aplicar borde
            numero_formato='0.00%'
        )
        
        # Celda AN - "PRE CANCELACION" (texto con borde por tener información)
        aplicar_formato_celda(
            ws.cell(fila_pre_cancelacion, col_an), 
            "PRE CANCELACION", 
            estilos,
            aplicar_borde=True
        )
        
        # Celda AO en fila PRE CANCELACION: operando (vacía, pero con borde porque interviene en fórmula)
        aplicar_formato_celda(
            ws.cell(fila_pre_cancelacion, col_ao), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )

        # En la fila INFERIOR, colocar la celda AO amarilla con la fórmula
        # Igual a Figura 1: =AO{fila_total}-AO{fila_pre_cancelacion}
        fila_ao_inferior = fila_pre_cancelacion + 1
        formula_ao_inferior = f"={col_letter_ao}{fila_total}-{col_letter_ao}{fila_pre_cancelacion}"
        aplicar_formato_celda(
            ws.cell(fila_ao_inferior, col_ao), 
            formula_ao_inferior, 
            estilos,
            numero_formato='0.00',
            fill=estilos.fill_amarillo
        )
        
        # Celda vacía con borde en AO (2 filas después de la fórmula amarilla, antes de BASE 12%)
        fila_ao_vacia = fila_ao_inferior + 2
        aplicar_formato_celda(
            ws.cell(fila_ao_vacia, col_ao), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        
        # Celda AP - Operando vacío (se usará en fórmulas), borde forzado
        aplicar_formato_celda(
            ws.cell(fila_pre_cancelacion, col_ap), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        
        # Celda AQ - "BASE 0%" (texto con borde y fondo amarillo)
        aplicar_formato_celda(
            ws.cell(fila_pre_cancelacion, col_aq), 
            "BASE 0%", 
            estilos,
            aplicar_borde=True,
            fill=estilos.fill_amarillo
        )

        # Celda AR - "TOTAL PRE CANCELACIONES" (texto con borde)
        aplicar_formato_celda(
            ws.cell(fila_pre_cancelacion, col_ar), 
            "TOTAL PRE CANCELACIONES", 
            estilos,
            aplicar_borde=True
        )
        
        # Ajustar ancho de columna AR para "TOTAL PRE CANCELACIONES"
        ws.column_dimensions[col_letter_ar].width = 25
        
        # Celda AR debajo de TOTAL PRE CANCELACIONES: vacía con borde
        aplicar_formato_celda(
            ws.cell(fila_pre_cancelacion + 1, col_ar), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        
        # Fila de BASE 12% (texto con borde y fondo amarillo) colocada en columna AQ
        aplicar_formato_celda(
            ws.cell(fila_base_12, col_aq), 
            "BASE 12%", 
            estilos,
            aplicar_borde=True,
            fill=estilos.fill_amarillo
        )
        
        # Celda AP en fila BASE 12%: vacía con borde (al lado izquierdo de BASE 12%)
        aplicar_formato_celda(
            ws.cell(fila_base_12, col_ap), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        
        # Celda AR en fila BASE 12%: vacía con borde
        aplicar_formato_celda(
            ws.cell(fila_base_12, col_ar), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        
        # Fila DEBAJO de BASE 12%: AO copia valor de arriba
        fila_ao_copia = fila_base_12 + 1
        aplicar_formato_celda(
            ws.cell(fila_ao_copia, col_ao), 
            f"={col_letter_ao}{fila_base_12}", 
            estilos,
            numero_formato='0.00'
        )
        
        # Encabezados tabla de pólizas (texto con borde)
        aplicar_formato_celda(
            ws.cell(fila_tabla_encabezado, col_ao), 
            "#", 
            estilos,
            aplicar_borde=True
        )
        
        aplicar_formato_celda(
            ws.cell(fila_tabla_encabezado, col_ap), 
            "TOTAL", 
            estilos,
            aplicar_borde=True
        )
        
        # Pólizas (texto en columna AN, con borde; celdas AO/AP vacías con borde)
        aplicar_formato_celda(
            ws.cell(fila_poliza_1, col_an), 
            "5852", 
            estilos,
            aplicar_borde=True
        )
        aplicar_formato_celda(
            ws.cell(fila_poliza_1, col_ao), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        aplicar_formato_celda(
            ws.cell(fila_poliza_1, col_ap), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        
        aplicar_formato_celda(
            ws.cell(fila_poliza_2, col_an), 
            "7650", 
            estilos,
            aplicar_borde=True
        )
        aplicar_formato_celda(
            ws.cell(fila_poliza_2, col_ao), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        aplicar_formato_celda(
            ws.cell(fila_poliza_2, col_ap), 
            "", 
            estilos,
            aplicar_borde=True,
            forzar_borde=True
        )
        
        # Totales tabla - Columna AO (con fondo amarillo)
        aplicar_formato_celda(
            ws.cell(fila_totales_tabla, col_ao), 
            f"=+{col_letter_ao}{fila_poliza_1}+{col_letter_ao}{fila_poliza_2}", 
            estilos,
            numero_formato='0.00',
            fill=estilos.fill_amarillo
        )
        
        # Totales tabla - Columna AP (con fondo amarillo)
        aplicar_formato_celda(
            ws.cell(fila_totales_tabla, col_ap), 
            f"=+{col_letter_ap}{fila_poliza_1}+{col_letter_ap}{fila_poliza_2}", 
            estilos,
            numero_formato='0.00',
            fill=estilos.fill_amarillo
        )
        
        if callback:
            callback("  ✓ Pie de página agregado (PRE CANCELACION, BASE 0%, BASE 12%, tabla de pólizas)")
        
    except Exception as e:
        if callback:
            callback(f"  ⚠️ Error al agregar pie de página: {str(e)}")
