# src/modelo/archivo.py
"""
Modelo de Archivo
Maneja la lógica de lectura/escritura de archivos
"""

import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


class ArchivoOrigen:
    """Representa el archivo 413 de origen"""
    
    def __init__(self, ruta):
        self.ruta = ruta
        self.nombre = os.path.basename(ruta)
        self.existe = os.path.exists(ruta)
        self.df = None
        self.hoja_datos = None
    
    def es_valido(self):
        """Valida que el archivo sea válido"""
        return self.existe and self.ruta.endswith(('.xlsx', '.xls'))
    
    def cargar(self, nombre_hoja='Report_AseguradoraMensual'):
        """Carga el archivo"""
        try:
            self.df = pd.read_excel(
                self.ruta,
                sheet_name=nombre_hoja,
                header=None,
                engine='openpyxl'
            )
            self.hoja_datos = nombre_hoja
            return True
        except Exception as e:
            print(f"Error al cargar: {e}")
            return False
    
    def obtener_datos(self):
        """Retorna los datos como DataFrame"""
        return self.df
    
    def obtener_filas(self):
        """Retorna cantidad de filas"""
        return len(self.df) if self.df is not None else 0


class ArchivoPlantilla:
    """Representa el archivo plantilla5852.xlsx"""
    
    def __init__(self, ruta):
        self.ruta = ruta
        self.nombre = os.path.basename(ruta)
        self.existe = os.path.exists(ruta)
        self.hojas = []
    
    def es_valido(self):
        """Valida que el archivo sea válido"""
        return self.existe and self.ruta.endswith(('.xlsx', '.xls'))
    
    def cargar_hojas(self):
        """Carga lista de hojas"""
        try:
            wb = load_workbook(self.ruta, read_only=True, data_only=True)
            self.hojas = wb.sheetnames
            wb.close()
            return True
        except:
            return False
    
    def obtener_hojas(self):
        """Retorna lista de hojas válidas (excluye auxiliares)"""
        return [
            h for h in self.hojas
            if 'CODIGO' not in h.upper() and 'HOJA1' not in h.upper()
        ]
    
    def obtener_todas_hojas(self):
        """Retorna todas las hojas"""
        return self.hojas


class ArchivoResultado:
    """Representa el archivo resultado transformado"""
    
    def __init__(self, nombre_base, poliza):
        self.poliza = poliza
        self.nombre_base = nombre_base
        self.nombre_archivo = self._generar_nombre()
        self.ruta_temporal = None
    
    def _generar_nombre(self):
        """Genera el nombre del archivo resultado"""
        from datetime import datetime
        
        # Obtener mes y año actual
        ahora = datetime.now()
        mes = ahora.strftime('%B').capitalize()
        año = ahora.year
        
        # Generar nombre con póliza
        prefijo = self.poliza.obtener_nombre_archivo()
        return f"{prefijo} {mes} {año}.xlsx"
    
    def establecer_ruta_temporal(self, ruta):
        """Establece la ruta temporal del archivo"""
        self.ruta_temporal = ruta
    
    def obtener_nombre_descarga(self):
        """Retorna el nombre para descargar"""
        return self.nombre_archivo
    
    def obtener_ruta_temporal(self):
        """Retorna la ruta temporal"""
        return self.ruta_temporal
