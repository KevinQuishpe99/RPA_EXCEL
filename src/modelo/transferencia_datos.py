# src/modelo/transferencia_datos.py
"""
Lógica de transferencia de datos y validación de filas
"""

import pandas as pd
from openpyxl.cell.cell import MergedCell
from datetime import datetime, date


class TransferenciaDatos:
    """Maneja transferencia de datos de origen a destino"""
    
    def __init__(self, estilos, cache_indices, formulas_cache, formulas_pattern):
        self.estilos = estilos
        self._cache_indices_columnas = cache_indices
        self._formulas_cache = formulas_cache
        self._formulas_pattern = formulas_pattern
        self._headers_cache = {}
    
    def transferir_datos(self, ws, df_origen, fila_encabezados, headers_origen, headers_destino, mapeo, callback=None, fila_destino_inicio=6, fila_plantilla=6, poliza_info=None):
        """Transfiere datos de origen a destino replicando la lógica original"""
        fila_destino = fila_destino_inicio
        filas_procesadas = 0
        fila_origen_inicio = fila_encabezados + 1
        limite_origen = len(df_origen)
        
        # Para TC, la primera columna podría estar vacía - buscar primera columna mapeada
        col_validacion = 0
        if poliza_info and poliza_info.get('prefijo') == 'TC':
            # Encontrar la primera columna mapeada (generalmente será 1 o mayor)
            if mapeo:
                col_indices = sorted([k for k in mapeo.keys() if isinstance(k, int)])
                if col_indices and col_indices[0] > 0:
                    col_validacion = col_indices[0]
        
        primera_columna_array = df_origen.iloc[fila_origen_inicio:, col_validacion].values

        # Índices cacheados para columnas especiales
        idx_pais_residencia_dest = self._cache_indices_columnas.get('idx_pais_residencia_dest')
        if idx_pais_residencia_dest is None:
            for idx, cell in enumerate(headers_destino):
                if cell.value and 'PAIS DE RESIDENCIA' in str(cell.value).upper():
                    idx_pais_residencia_dest = idx + 1
                    self._cache_indices_columnas['idx_pais_residencia_dest'] = idx_pais_residencia_dest
                    break

        idx_pais_origen = self._cache_indices_columnas.get('idx_pais_origen')
        idx_provincia_orig = self._cache_indices_columnas.get('idx_provincia_orig')
        idx_ciudad_orig = self._cache_indices_columnas.get('idx_ciudad_orig')

        col_provincia_dest = 15
        col_ciudad_dest = 16

        # Forzar mapeo de PROVINCIA y CIUDAD
        if idx_provincia_orig is None:
            for idx, header in enumerate(headers_origen):
                if pd.notna(header):
                    header_str = str(header).strip().upper()
                    if 'PROVINCIA' in header_str and 'PAIS' not in header_str:
                        mapeo[idx] = col_provincia_dest
                        idx_provincia_orig = idx
                        self._cache_indices_columnas['idx_provincia_orig'] = idx
                        break
        else:
            mapeo[idx_provincia_orig] = col_provincia_dest

        if idx_ciudad_orig is None:
            for idx, header in enumerate(headers_origen):
                if pd.notna(header):
                    header_str = str(header).strip().upper()
                    if 'CIUDAD' in header_str:
                        mapeo[idx] = col_ciudad_dest
                        idx_ciudad_orig = idx
                        self._cache_indices_columnas['idx_ciudad_orig'] = idx
                        break
        else:
            mapeo[idx_ciudad_orig] = col_ciudad_dest

        if idx_pais_origen is None:
            for idx, header in enumerate(headers_origen):
                if pd.notna(header) and 'PAIS DE ORIGEN' in str(header).upper():
                    idx_pais_origen = idx
                    self._cache_indices_columnas['idx_pais_origen'] = idx
                    break

        for idx_local, idx_origen in enumerate(range(fila_origen_inicio, limite_origen)):
            try:
                primera_col = primera_columna_array[idx_local]
                fila_valida = False

                if pd.notna(primera_col):
                    if isinstance(primera_col, str):
                        valor_limpio = primera_col.strip()
                        if valor_limpio and valor_limpio.upper() not in ['NAN', 'NONE', 'NULL', 'TOTAL', 'CUADRE', 'PRECANCELACION', '']:
                            fila_valida = True
                    else:
                        fila_valida = True

                if not fila_valida:
                    continue

                if isinstance(primera_col, str):
                    primera_col_str = primera_col.strip().upper()
                    if 'TOTAL' in primera_col_str or 'CUADRE' in primera_col_str or 'PRECANCELACION' in primera_col_str:
                        break
            except Exception:
                continue

            try:
                self.transferir_fila_optimizada(
                    df_origen,
                    idx_origen,
                    ws,
                    fila_destino,
                    mapeo,
                    headers_origen,
                    headers_destino,
                    idx_pais_origen,
                    idx_pais_residencia_dest,
                    fila_plantilla=fila_plantilla,
                    poliza_info=poliza_info,
                )
                filas_procesadas += 1
                fila_destino += 1
            except Exception:
                continue

        return filas_procesadas

    def transferir_fila_optimizada(self, df_origen, idx_origen, ws_destino, fila_destino,
                                   mapeo_columnas, headers_origen, headers_destino,
                                   idx_pais_origen=None, idx_pais_residencia_dest=None, fila_plantilla=6, poliza_info=None):
        """Copia fórmulas y datos aplicando las mismas transformaciones del monolito"""
        row_origen = df_origen.iloc[idx_origen]

        # Detectar tipo de identificación
        tipo_identificacion = None
        idx_tipo_id = self._cache_indices_columnas.get('idx_tipo_identificacion')
        if idx_tipo_id is None:
            for idx, header in enumerate(headers_origen):
                if pd.notna(header) and 'TIPO IDENTIFICACION' in str(header).upper():
                    idx_tipo_id = idx
                    self._cache_indices_columnas['idx_tipo_identificacion'] = idx
                    break
        if idx_tipo_id is not None and idx_tipo_id < len(row_origen):
            tipo_valor = row_origen.iloc[idx_tipo_id]
            if pd.notna(tipo_valor):
                tipo_identificacion = str(tipo_valor).strip()

        # Paso 1: copiar fórmulas de la fila 6
        if fila_destino != fila_plantilla:
            diferencia_filas = fila_destino - fila_plantilla

            if fila_plantilla not in self._formulas_cache:
                formulas_plantilla = {}
                max_cols = min(ws_destino.max_column, 200)
                for idx, cell_plantilla in enumerate(ws_destino[fila_plantilla], start=1):
                    if idx > max_cols:
                        break
                    if not isinstance(cell_plantilla, MergedCell) and cell_plantilla.data_type == 'f':
                        formulas_plantilla[idx] = str(cell_plantilla.value)
                self._formulas_cache[fila_plantilla] = formulas_plantilla

            formulas_plantilla = self._formulas_cache[fila_plantilla]

            def reemplazar_ref(match):
                col_ref = match.group(1)
                fila_ref = int(match.group(2))
                match_completo = match.group(0)
                if '$' in match_completo:
                    partes = match_completo.split(str(fila_ref))
                    if len(partes) > 0 and '$' in partes[0]:
                        return match_completo
                return f"{col_ref}{fila_ref + diferencia_filas}"

            headers_edad_cache = {}
            for col in formulas_plantilla.keys():
                if col-1 < len(headers_destino) and headers_destino[col-1].value:
                    header_cell = str(headers_destino[col-1].value).upper()
                    if 'EDAD' in header_cell:
                        headers_edad_cache[col] = True

            for col, formula_original in formulas_plantilla.items():
                try:
                    cell_destino = ws_destino.cell(fila_destino, col)
                    if isinstance(cell_destino, MergedCell):
                        continue
                    formula_ajustada = self._formulas_pattern.sub(reemplazar_ref, formula_original) if diferencia_filas != 0 else formula_original
                    if col in headers_edad_cache and 'ROUND' not in formula_ajustada.upper():
                        formula_sin_igual = formula_ajustada[1:] if formula_ajustada.startswith('=') else formula_ajustada
                        formula_ajustada = f"=ROUND({formula_sin_igual},2)"
                    cell_destino.value = formula_ajustada
                except Exception:
                    continue

        # Paso 2: copiar datos mapeados con transformaciones
        self._aplicar_transformaciones(
            df_origen, idx_origen, ws_destino, fila_destino, mapeo_columnas,
            headers_origen, headers_destino, tipo_identificacion, idx_pais_origen, poliza_info
        )

        # Paso 3: aplicar campos fijos según póliza (DESPUÉS de copiar datos para sobrescribir)
        if poliza_info and poliza_info.get('prefijo') == 'TC':
            # Para TC, aplicar campos fijos específicos
            self._aplicar_campos_fijos_tc(ws_destino, fila_destino, headers_destino, poliza_info)
        else:
            # Para DV, aplicar campos fijos tradicionales
            if idx_pais_residencia_dest is not None:
                try:
                    cell_destino = ws_destino.cell(fila_destino, idx_pais_residencia_dest)
                    if not isinstance(cell_destino, MergedCell) and cell_destino.data_type != 'f':
                        cell_destino.value = '239'
                except Exception:
                    pass
            self._escribir_numero_poliza(ws_destino, fila_destino, headers_destino)
            self._escribir_nombre_producto(ws_destino, fila_destino, headers_destino)

    def _aplicar_transformaciones(self, df_origen, idx_origen, ws_destino, fila_destino,
                                  mapeo_columnas, headers_origen, headers_destino,
                                  tipo_identificacion, idx_pais_origen, poliza_info=None):
        """Aplica transformaciones de datos según tipo de columna"""
        # Importar constantes TC si es necesario
        no_sobrescribir_tc = set()
        if poliza_info and poliza_info.get('prefijo') == 'TC':
            from .mapeo_tc import NO_SOBRESCRIBIR_TC
            no_sobrescribir_tc = NO_SOBRESCRIBIR_TC
        
        row_origen = df_origen.iloc[idx_origen]
        valor_pais_origen = None
        
        if idx_pais_origen is not None and idx_pais_origen < len(row_origen):
            valor_pais = row_origen.iloc[idx_pais_origen]
            if pd.notna(valor_pais) and str(valor_pais).strip() != '':
                valor_pais_origen = valor_pais

        es_provincia_ciudad_cols = {15, 16}
        es_columna_ap_bc = set(range(42, 56))
        row_values = row_origen.values if hasattr(row_origen, 'values') else [row_origen.iloc[i] for i in range(len(row_origen))]

        for idx_origen_col, col_destino in mapeo_columnas.items():
            try:
                # Para TC, no sobrescribir ciertas columnas (tienen fórmulas)
                # col_destino es 1-based, pero NO_SOBRESCRIBIR_TC usa índices 0-based
                if (col_destino - 1) in no_sobrescribir_tc:
                    continue
                
                if idx_origen_col >= len(row_values):
                    continue

                valor = row_values[idx_origen_col]

                if pd.isna(valor):
                    continue

                valor_str = valor.strip() if isinstance(valor, str) else str(valor).strip()
                if not valor_str or valor_str.lower() == 'nan':
                    continue

                cell_destino = ws_destino.cell(fila_destino, col_destino)
                if isinstance(cell_destino, MergedCell):
                    continue
                if col_destino not in es_provincia_ciudad_cols and col_destino not in es_columna_ap_bc and cell_destino.data_type == 'f':
                    continue

                header_orig = None
                if idx_origen_col < len(headers_origen):
                    if idx_origen_col not in self._headers_cache:
                        self._headers_cache[idx_origen_col] = str(headers_origen[idx_origen_col]).strip().upper()
                    header_orig = self._headers_cache[idx_origen_col]

                    # Aplicar transformación según tipo de columna
                    valor = self._transformar_valor(valor, valor_str, header_orig, col_destino, tipo_identificacion, poliza_info)

                cell_destino.value = valor
                
                # Aplicar formato
                if 'FECHA' in (header_orig or '') and isinstance(valor, (date, datetime)):
                    cell_destino.number_format = 'mm/dd/yyyy'
                elif isinstance(valor, (int, float)) and not isinstance(valor, bool):
                    cell_destino.number_format = '0.00'
            except Exception:
                continue

    def _transformar_valor(self, valor, valor_str, header_orig, col_destino, tipo_identificacion, poliza_info=None):
        """Aplica transformación específica según el tipo de columna"""
        header_upper = header_orig.upper() if header_orig else ""
        
        # Columnas numéricas que deben ser enteros sin ceros adelante
        numeras_sin_ceros_patterns = [
            'PROVINCIA' in header_upper and 'PAIS' not in header_upper,
            'CIUDAD' in header_upper,
            'NACIONALIDAD' in header_upper,
            'PAIS DE RESIDENCIA' in header_upper,
            'PAIS DE ORIGEN' in header_upper,
        ]
        
        # Listas de columnas decimales
        decimales_patterns = [
            'MONTO CREDITO' in header_upper,
            'MONTO CRÉDITO' in header_upper,
            'PLAZO DE CREDITO' in header_upper or 'PLAZO DE CRÉDITO' in header_upper,
            'PRIMA NETA' in header_upper,
            'INGRESOS' in header_upper,
            'PATRIMONIO' in header_upper,
            'SALDO' in header_upper and 'FECHA' not in header_upper,
        ]
        
        # Aplicar conversión de números sin ceros adelante
        if any(numeras_sin_ceros_patterns):
            try:
                if isinstance(valor, (int, float)):
                    return int(valor)
                elif isinstance(valor_str, str) and valor_str:
                    # Quitar ceros iniciales
                    if len(valor_str) > 1 and valor_str[0] == '0' and valor_str[1:].isdigit():
                        return int(valor_str[1:])
                    elif valor_str.isdigit():
                        return int(valor_str)
                    try:
                        if valor_str.replace('.', '', 1).replace('-', '', 1).isdigit():
                            return int(float(valor_str))
                    except Exception:
                        pass
            except Exception:
                pass
            return valor_str
        
        # Aplicar decimales limitados a 2
        if any(decimales_patterns):
            try:
                return round(float(valor), 2)
            except Exception:
                return valor_str
        
        # Manejo de fechas
        if 'FECHA' in header_upper:
            try:
                fecha_obj = None
                if isinstance(valor, (datetime, pd.Timestamp)):
                    fecha_obj = valor.date() if hasattr(valor, 'date') else valor
                elif isinstance(valor_str, str):
                    if ' ' in valor_str:
                        valor_str = valor_str.split(' ')[0]
                    if '-' in valor_str or '/' in valor_str:
                        fecha_parsed = pd.to_datetime(valor_str)
                        fecha_obj = fecha_parsed.date() if hasattr(fecha_parsed, 'date') else fecha_parsed
                if fecha_obj:
                    return fecha_obj.date() if hasattr(fecha_obj, 'date') else fecha_obj
            except Exception:
                pass
        
        return valor_str

    def _aplicar_campos_fijos_tc(self, ws_destino, fila_destino, headers_destino, poliza_info):
        """Aplica valores fijos específicos de la póliza TC - sobrescribe datos mapeados"""
        if not poliza_info:
            return
        
        # NUMERO DE POLIZA fijo (5924)
        numero_poliza = poliza_info.get('numero_poliza_fijo')
        if numero_poliza:
            for idx, cell in enumerate(headers_destino):
                if cell.value:
                    header_upper = str(cell.value).upper()
                    if 'NUMERO' in header_upper and 'POLIZA' in header_upper:
                        try:
                            cell_destino = ws_destino.cell(fila_destino, idx + 1)
                            if not isinstance(cell_destino, MergedCell) and cell_destino.data_type != 'f':
                                cell_destino.value = numero_poliza
                                break
                        except Exception:
                            pass
        
        # NOMBRE PRODUCTO fijo
        nombre_producto = poliza_info.get('nombre_producto_fijo')
        if nombre_producto:
            for idx, cell in enumerate(headers_destino):
                if cell.value:
                    header_upper = str(cell.value).upper()
                    if 'NOMBRE' in header_upper and 'PRODUCTO' in header_upper:
                        try:
                            cell_destino = ws_destino.cell(fila_destino, idx + 1)
                            if not isinstance(cell_destino, MergedCell) and cell_destino.data_type != 'f':
                                cell_destino.value = nombre_producto
                                break
                        except Exception:
                            pass
        
        # PAIS DE RESIDENCIA fijo (239)
        pais_residencia = poliza_info.get('pais_residencia_fijo')
        if pais_residencia:
            for idx, cell in enumerate(headers_destino):
                if cell.value:
                    header_upper = str(cell.value).upper()
                    if 'PAIS' in header_upper and 'RESIDENCIA' in header_upper:
                        try:
                            cell_destino = ws_destino.cell(fila_destino, idx + 1)
                            if not isinstance(cell_destino, MergedCell) and cell_destino.data_type != 'f':
                                cell_destino.value = pais_residencia
                                break
                        except Exception:
                            pass

    def _escribir_numero_poliza(self, ws_destino, fila_destino, headers_destino):
        """Escribe número de póliza fijo para DV (5852)"""
        if not hasattr(self, '_idx_numero_poliza'):
            self._idx_numero_poliza = None
            for idx, cell in enumerate(headers_destino):
                if cell.value and 'NUMERO' in str(cell.value).upper() and 'POLIZA' in str(cell.value).upper():
                    self._idx_numero_poliza = idx + 1
                    break
        if self._idx_numero_poliza is not None:
            try:
                cell_poliza = ws_destino.cell(fila_destino, self._idx_numero_poliza)
                if not isinstance(cell_poliza, MergedCell) and cell_poliza.data_type != 'f':
                    cell_poliza.value = '5852'
            except Exception:
                pass

    def _escribir_nombre_producto(self, ws_destino, fila_destino, headers_destino):
        """Escribe nombre producto fijo para DV (MONTO DEL CREDITO)"""
        if not hasattr(self, '_idx_nombre_producto'):
            self._idx_nombre_producto = None
            for idx, cell in enumerate(headers_destino):
                if cell.value and 'NOMBRE' in str(cell.value).upper() and 'PRODUCTO' in str(cell.value).upper():
                    self._idx_nombre_producto = idx + 1
                    break
        if self._idx_nombre_producto is not None:
            try:
                cell_producto = ws_destino.cell(fila_destino, self._idx_nombre_producto)
                if not isinstance(cell_producto, MergedCell) and cell_producto.data_type != 'f':
                    cell_producto.value = 'MONTO DEL CREDITO'
            except Exception:
                pass
