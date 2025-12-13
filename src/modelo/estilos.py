# src/modelo/estilos.py
"""
Estilos reutilizables para celdas Excel
"""

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class EstilosExcel:
    """Contenedor de estilos predefinidos para Excel"""
    
    def __init__(self):
        self._preparar_estilos()
    
    def _preparar_estilos(self):
        """Pre-crea estilos para mejor rendimiento"""
        # Fuentes
        self.fuente_calibri = Font(name='Calibri')
        self.fuente_calibri_negrita = Font(name='Calibri', bold=True)
        
        # Alineaciones
        self.alineacion_centrada = Alignment(horizontal='center', vertical='center')
        
        # Bordes
        self.borde_delgado = Side(style='thin')
        self.borde_celda = Border(
            left=self.borde_delgado,
            right=self.borde_delgado,
            top=self.borde_delgado,
            bottom=self.borde_delgado
        )
        
        # Rellenos
        self.fill_amarillo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.fill_gris = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        self.fill_azul = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
