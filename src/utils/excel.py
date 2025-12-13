# src/utils/excel.py
"""
Utilidades para manejo de Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


def cargar_datos_excel(ruta, nombre_hoja=None):
    """
    Carga datos de un archivo Excel.
    
    Args:
        ruta (str): Ruta del archivo
        nombre_hoja (str): Nombre de la hoja (si None, usa primera)
    
    Returns:
        pd.DataFrame: DataFrame con los datos
    """
    try:
        df = pd.read_excel(ruta, sheet_name=nombre_hoja, engine='openpyxl')
        return df
    except Exception as e:
        print(f"Error al cargar Excel: {e}")
        return None


def obtener_hojas_excel(ruta):
    """
    Obtiene lista de hojas de un archivo Excel.
    
    Args:
        ruta (str): Ruta del archivo
    
    Returns:
        list: Lista de nombres de hojas
    """
    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        hojas = wb.sheetnames
        wb.close()
        return hojas
    except:
        return []


def filtrar_hojas_validas(hojas, excluir_palabras_clave=None):
    """
    Filtra hojas válidas excluyendo según palabras clave.
    
    Args:
        hojas (list): Lista de nombres de hojas
        excluir_palabras_clave (list): Palabras clave para excluir
    
    Returns:
        list: Hojas filtradas
    """
    if excluir_palabras_clave is None:
        excluir_palabras_clave = ['CODIGO', 'AUXILIAR', 'HOJA1', 'TEMP']
    
    resultado = []
    for hoja in hojas:
        valida = True
        for palabra in excluir_palabras_clave:
            if palabra.upper() in hoja.upper():
                valida = False
                break
        if valida:
            resultado.append(hoja)
    
    return resultado


def aplicar_formato_celda(celda, estilo=None):
    """
    Aplica formato a una celda.
    
    Args:
        celda: Celda de openpyxl
        estilo (dict): Diccionario con estilos a aplicar
    """
    if estilo is None:
        estilo = {}
    
    # Fuente
    if 'fuente' in estilo or 'tamaño' in estilo:
        celda.font = Font(
            name=estilo.get('fuente', 'Calibri'),
            size=estilo.get('tamaño', 11)
        )
    
    # Alineación
    if 'alineacion' in estilo:
        celda.alignment = Alignment(
            horizontal=estilo['alineacion'].get('horizontal', 'left'),
            vertical=estilo['alineacion'].get('vertical', 'center')
        )
    
    # Bordes
    if estilo.get('borde', False):
        lado_delgado = Side(style='thin')
        celda.border = Border(
            left=lado_delgado,
            right=lado_delgado,
            top=lado_delgado,
            bottom=lado_delgado
        )
