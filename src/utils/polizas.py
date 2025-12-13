# Utilidades para detección de pólizas
import re
from openpyxl import load_workbook


def detectar_poliza_desde_plantilla(archivo_plantilla, configuracion_polizas):
    """
    Detecta automáticamente qué póliza usar basándose en las hojas de la plantilla.
    
    Args:
        archivo_plantilla: Ruta al archivo plantilla5852.xlsx
        configuracion_polizas: Dict con configuración de pólizas
        
    Returns:
        dict: Información de póliza detectada o None
        {
            'nombre_hoja': 'DV(5852)',
            'tipo_poliza': 'DV',
            'numero_poliza': '5852',
            'config': {...}
        }
    """
    try:
        # Leer hojas de la plantilla
        wb = load_workbook(archivo_plantilla, read_only=True, data_only=True)
        hojas_disponibles = wb.sheetnames
        wb.close()
        
        # Buscar hojas que coincidan con pólizas configuradas
        polizas_encontradas = []
        
        for nombre_hoja in hojas_disponibles:
            # Ignorar hojas auxiliares
            if 'CODIGO' in nombre_hoja.upper() or 'HOJA1' in nombre_hoja.upper():
                continue
            
            # Buscar coincidencias con pólizas configuradas
            for tipo_poliza, config in configuracion_polizas.items():
                for patron in config['patrones_hoja']:
                    if re.search(patron, nombre_hoja, re.IGNORECASE):
                        # Extraer número de póliza
                        numero_poliza = _extraer_numero(nombre_hoja)
                        
                        polizas_encontradas.append({
                            'nombre_hoja': nombre_hoja,
                            'tipo_poliza': tipo_poliza,
                            'numero_poliza': numero_poliza,
                            'config': config
                        })
                        break
        
        # Si se encontró solo una hoja, usarla
        if len(polizas_encontradas) == 1:
            return polizas_encontradas[0]
        # Si hay múltiples, retornar la primera
        elif len(polizas_encontradas) > 1:
            return polizas_encontradas[0]
        
        return None
        
    except Exception as e:
        print(f"Error al detectar póliza: {e}")
        return None


def _extraer_numero(nombre_hoja):
    """Extrae número de póliza del nombre de hoja"""
    # Buscar entre paréntesis
    match = re.search(r'\((\d+)\)', nombre_hoja)
    if match:
        return match.group(1)
    
    # Buscar después de espacio
    match = re.search(r'\s+(\d+)', nombre_hoja)
    if match:
        return match.group(1)
    
    return None


def generar_nombre_archivo(poliza_info, fecha_mes=None):
    """
    Genera nombre de archivo automáticamente.
    
    Args:
        poliza_info: Dict con información de póliza
        fecha_mes: datetime object con la fecha (opcional)
        
    Returns:
        str: Nombre del archivo generado
    """
    from datetime import datetime
    from ..config import MESES_ESPANOL
    
    if not fecha_mes:
        fecha_mes = datetime.now()
    
    # Obtener información de póliza
    config = poliza_info.get('config', {})
    prefijo = config.get('nombre_archivo', 'Facturación')
    
    # Obtener mes y año
    mes_num = fecha_mes.month
    nombre_mes = MESES_ESPANOL.get(mes_num, 'mes').capitalize()
    año = fecha_mes.year
    
    # Generar nombre
    nombre_archivo = f"{prefijo} {nombre_mes} {año}.xlsx"
    
    return nombre_archivo


def obtener_hojas_procesables(archivo_plantilla):
    """
    Obtiene lista de hojas que pueden ser procesadas (pólizas válidas).
    
    Args:
        archivo_plantilla: Ruta al archivo
        
    Returns:
        list: Nombres de hojas válidas
    """
    try:
        wb = load_workbook(archivo_plantilla, read_only=True, data_only=True)
        hojas = [
            h for h in wb.sheetnames 
            if 'CODIGO' not in h.upper() and 'HOJA1' not in h.upper()
        ]
        wb.close()
        return hojas
    except:
        return []
