# Utilidades para búsqueda y detección
import re
import os
from openpyxl import load_workbook

def buscar_plantilla(ubicaciones_posibles=None):
    """
    Busca el archivo plantilla5852.xlsx en múltiples ubicaciones.
    
    Args:
        ubicaciones_posibles: Lista de rutas a buscar. Si es None, usa ubicaciones estándar.
        
    Returns:
        str: Ruta al archivo plantilla5852.xlsx encontrado
        None: Si no se encuentra
    """
    if ubicaciones_posibles is None:
        import sys
        
        ubicaciones_posibles = []
        
        # Si es EXE compilado (PyInstaller)
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            ubicaciones_posibles.append(os.path.join(base_path, "src", "plantillas", "plantilla5852.xlsx"))
            ubicaciones_posibles.append(os.path.join(base_path, "plantillas", "plantilla5852.xlsx"))
            ubicaciones_posibles.append(os.path.join(base_path, "plantilla5852.xlsx"))
        
        # Directorio del script
        directorio_script = os.path.dirname(os.path.abspath(__file__))
        ubicaciones_posibles.append(os.path.join(directorio_script, "..", "plantillas", "plantilla5852.xlsx"))
        ubicaciones_posibles.append(os.path.join(directorio_script, "plantilla5852.xlsx"))
        
        # Directorio actual
        directorio_actual = os.getcwd()
        ubicaciones_posibles.append(os.path.join(directorio_actual, "src", "plantillas", "plantilla5852.xlsx"))
        ubicaciones_posibles.append(os.path.join(directorio_actual, "plantillas", "plantilla5852.xlsx"))
        ubicaciones_posibles.append(os.path.join(directorio_actual, "plantilla5852.xlsx"))
        
        # Carpeta 'programa'
        ubicaciones_posibles.append(os.path.join(directorio_actual, "programa", "plantilla5852.xlsx"))
        ubicaciones_posibles.append(os.path.join(directorio_script, "programa", "plantilla5852.xlsx"))
        ubicaciones_posibles.append(os.path.join(directorio_actual, "programa", "plantillas", "plantilla5852.xlsx"))
    
    # Buscar en cada ubicación
    for ruta in ubicaciones_posibles:
        if os.path.exists(ruta):
            return ruta
    
    return None


def extraer_numero_poliza(nombre_hoja):
    """
    Extrae el número de póliza del nombre de la hoja.
    
    Ejemplos:
        'DV(5852)' → '5852'
        'RC 6789' → '6789'
        'AP(1234)' → '1234'
        'DV' → None
        
    Args:
        nombre_hoja: Nombre de la hoja
        
    Returns:
        str: Número de póliza o None
    """
    # Buscar número entre paréntesis: DV(5852)
    match = re.search(r'\((\d+)\)', nombre_hoja)
    if match:
        return match.group(1)
    
    # Buscar número después de espacio: DV 5852
    match = re.search(r'\s+(\d+)', nombre_hoja)
    if match:
        return match.group(1)
    
    return None


def obtener_hojas_validas(archivo_plantilla):
    """
    Obtiene lista de hojas válidas en la plantilla (excluyendo auxiliares).
    
    Args:
        archivo_plantilla: Ruta al archivo plantilla5852.xlsx
        
    Returns:
        list: Nombres de hojas válidas
    """
    try:
        wb = load_workbook(archivo_plantilla, read_only=True, data_only=True)
        hojas = wb.sheetnames
        wb.close()
        
        # Filtrar hojas auxiliares
        hojas_validas = [
            h for h in hojas 
            if 'CODIGO' not in h.upper() and 'HOJA1' not in h.upper()
        ]
        
        return hojas_validas
    except Exception as e:
        print(f"Error al leer hojas: {e}")
        return []


def validar_patron_regex(patron, texto):
    """
    Valida si un texto coincide con un patrón regex.
    
    Args:
        patron: Patrón regex (str)
        texto: Texto a validar
        
    Returns:
        bool: True si coincide
    """
    try:
        return bool(re.search(patron, texto, re.IGNORECASE))
    except:
        return False


def limpiar_nombre_archivo(nombre):
    """
    Limpia nombre de archivo eliminando caracteres inválidos.
    
    Args:
        nombre: Nombre del archivo
        
    Returns:
        str: Nombre limpio
    """
    # Reemplazar caracteres inválidos
    caracteres_invalidos = r'[<>:"/\\|?*]'
    nombre_limpio = re.sub(caracteres_invalidos, '_', nombre)
    
    # Eliminar espacios múltiples
    nombre_limpio = re.sub(r'\s+', ' ', nombre_limpio)
    
    # Limitar longitud
    if len(nombre_limpio) > 200:
        nombre_limpio = nombre_limpio[:200]
    
    return nombre_limpio.strip()
