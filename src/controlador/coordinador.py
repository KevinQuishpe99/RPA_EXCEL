# src/controlador/coordinador.py
"""
Controlador Coordinador
Orquesta la interacción entre Modelo y Vista
"""

import threading
from pathlib import Path
from src.modelo import (
    Poliza, ArchivoOrigen, ArchivoPlantilla, 
    ArchivoResultado, TransformadorDatos
)
from src.config.polizas import CONFIGURACION_POLIZAS, TRANSFORMACIONES

# Qt helper for thread-safe UI updates (no-op if Qt is not available)
try:
    from PySide6.QtCore import QTimer
except Exception:
    QTimer = None


class CoordinadorPrincipal:
    """Coordinador principal - orquesta Model y View"""
    
    def __init__(self, vista):
        self.vista = vista
        self.archivo_actual = None
        self.poliza_actual = None
        self.polizas_disponibles = {}
        self.transformador = None
        self.archivo_plantilla = None
        
        self._inicializar()
    
    def _inicializar(self):
        """Inicializa el coordinador"""
        # Crear instancias de pólizas
        self._crear_polizas()
        
        # Crear transformador con callback de mensajes
        self.transformador = TransformadorDatos(
            callback_mensaje=self._mostrar_mensaje
        )
        
        # Conectar callbacks / señales según tipo de vista
        if hasattr(self.vista, 'callback_seleccionar_archivo'):
            self.vista.callback_seleccionar_archivo = self.archivo_seleccionado
        if hasattr(self.vista, 'callback_transformar'):
            self.vista.callback_transformar = self.iniciar_transformacion
        if hasattr(self.vista, 'callback_descargar'):
            self.vista.callback_descargar = self.descargar_archivo
        # Tkinter: botón
        if hasattr(self.vista, 'btn_transformar') and hasattr(self.vista.btn_transformar, 'config'):
            try:
                self.vista.btn_transformar.config(command=self.iniciar_transformacion)
            except Exception:
                pass
        # PySide6: señales
        try:
            if hasattr(self.vista, 'solicitar_transformacion'):
                self.vista.solicitar_transformacion.connect(self.iniciar_transformacion)
            if hasattr(self.vista, 'archivo_seleccionado'):
                self.vista.archivo_seleccionado.connect(self.archivo_seleccionado)
            if hasattr(self.vista, 'descargar_resultado'):
                self.vista.descargar_resultado.connect(self.descargar_archivo)
        except Exception:
            pass
        
        # Actualizar lista de pólizas
        self._actualizar_polizas_vista()
        
        # Buscar plantilla
        self._buscar_plantilla()

    # ===== Helpers =====
    def _en_ui(self, func, *args, **kwargs):
        """Ejecuta un callable en el hilo de UI si QTimer está disponible."""
        if QTimer:
            QTimer.singleShot(0, lambda: func(*args, **kwargs))
        else:
            func(*args, **kwargs)
    
    def _crear_polizas(self):
        """Crea instancias de pólizas desde configuración"""
        for tipo, config in CONFIGURACION_POLIZAS.items():
            self.polizas_disponibles[tipo] = Poliza(tipo, config)
    
    def _actualizar_polizas_vista(self):
        """Actualiza la lista de pólizas en vista"""
        nombres_polizas = list(self.polizas_disponibles.keys())
        try:
            if hasattr(self.vista, 'establecer_polizas'):
                self.vista.establecer_polizas(nombres_polizas)
            elif hasattr(self.vista, 'set_polizas'):
                self.vista.set_polizas(nombres_polizas)
        except Exception:
            pass
    
    def _buscar_plantilla(self):
        """Busca el archivo plantilla"""
        rutas_busqueda = [
            Path.cwd() / 'src' / 'plantillas' / 'plantilla5852.xlsx',
            Path.cwd() / 'plantillas' / 'plantilla5852.xlsx',
            Path.cwd() / 'plantilla5852.xlsx',
            Path.cwd() / 'src' / 'plantilla5852.xlsx',
            Path.cwd() / 'resources' / 'plantilla5852.xlsx',
        ]
        
        for ruta in rutas_busqueda:
            if ruta.exists():
                self.archivo_plantilla = ArchivoPlantilla(str(ruta))
                if self.archivo_plantilla.cargar_hojas():
                    self._add_msg(f"✓ Plantilla encontrada: {ruta.name}\n")
                    return
        
        self._add_msg("⚠ Plantilla no encontrada\n")
    
    def archivo_seleccionado(self, ruta):
        """Manejador cuando se selecciona un archivo"""
        self.archivo_actual = ArchivoOrigen(ruta)
        
        if not self.archivo_actual.es_valido():
            self.vista.mostrar_error("Error", "Archivo inválido")
            return
        
        self._add_msg(f"✓ Archivo seleccionado: {self.archivo_actual.nombre}\n")
        # Progreso inicial tras selección de archivo
        try:
            self.vista.establecer_progreso(10)
        except Exception:
            pass
    
    def iniciar_transformacion(self):
        """Inicia la transformación en un thread separado"""
        if not self.archivo_actual:
            self.vista.mostrar_error("Error", "Seleccione un archivo")
            return
        
        poliza_nombre = self.vista.obtener_poliza_seleccionada()
        if not poliza_nombre:
            self.vista.mostrar_error("Error", "Seleccione una póliza")
            return
        
        self.poliza_actual = self.polizas_disponibles[poliza_nombre]
        
        # Deshabilitar controles
        try:
            self.vista.habilitar_controles(False)
        except Exception:
            pass
        # Avance de progreso al iniciar
        try:
            self._set_progress(20)
        except Exception:
            pass
        
        # Ejecutar en thread
        thread = threading.Thread(target=self._ejecutar_transformacion)
        thread.daemon = True
        thread.start()
    
    def _ejecutar_transformacion(self):
        """Ejecuta la transformación de datos - LÓGICA REAL"""
        import os
        import tempfile
        
        try:
            # Crear transformador con callback de mensajes
            contador_mensajes = {'count': 0}
            
            def callback_mensaje(msg):
                self._add_msg(msg + "\n")
                # Avance progresivo suave basado en mensajes
                try:
                    contador_mensajes['count'] += 1
                    # Incremento pequeño y constante: +1% por cada mensaje hasta 75%
                    valor_actual = getattr(self.vista.barra_progreso, 'value', 20)
                    siguiente = min(75, int(valor_actual) + 1)
                    self._set_progress(siguiente)
                except Exception:
                    pass
            
            transformador = TransformadorDatos(callback_mensaje=callback_mensaje)
            
            # Obtener configuración de póliza
            poliza_config = self.poliza_actual.config
            # Avance antes de transformar
            try:
                self._set_progress(25)
            except Exception:
                pass
            
            # Determinar plantilla según póliza (DV -> 5852, TC -> 5924)
            import os
            plantilla_nombre = 'plantilla5852.xlsx'
            try:
                prefijo = str(poliza_config.get('prefijo', '')).upper()
            except Exception:
                prefijo = ''
            if prefijo == 'TC':
                plantilla_nombre = 'plantilla5924.xlsx'

            posibles_rutas = [
                os.path.join(os.getcwd(), 'src', 'plantillas', plantilla_nombre),
                os.path.join(os.getcwd(), 'plantillas', plantilla_nombre),
                os.path.join(os.getcwd(), plantilla_nombre),
            ]

            ruta_plantilla_elegida = None
            for rp in posibles_rutas:
                if os.path.exists(rp):
                    ruta_plantilla_elegida = rp
                    break

            if not ruta_plantilla_elegida:
                raise Exception(f"No se encontró la plantilla requerida: {plantilla_nombre}")

            # Validar que el archivo origen contiene la hoja requerida según póliza
            hoja_requerida = poliza_config.get('hoja_origen_requerida') if isinstance(poliza_config, dict) else None
            if hoja_requerida:
                try:
                    from openpyxl import load_workbook
                    wb_origen = load_workbook(self.archivo_actual.ruta, read_only=True, data_only=True)
                    nombres_hojas = [str(n) for n in wb_origen.sheetnames]
                    if hoja_requerida not in nombres_hojas:
                        msg = (
                            f"El archivo seleccionado no contiene la hoja requerida para la póliza seleccionada.\n\n"
                            f"Póliza: {poliza_config.get('prefijo','')}\n"
                            f"Hoja requerida: {hoja_requerida}\n"
                            f"Hojas encontradas: {', '.join(nombres_hojas)}\n\n"
                            f"Para DV (5852) se espera un archivo 413. Para TC (5924) se espera un archivo 455."
                        )
                        self._en_ui(self.vista.mostrar_error, "Hoja requerida no encontrada", msg)
                        # Re-habilitar controles y abortar
                        try:
                            if hasattr(self.vista, 'habilitar_controles'):
                                self._en_ui(self.vista.habilitar_controles, True)
                        except Exception:
                            pass
                        return
                except Exception:
                    # Si no se puede leer, continuar y dejar que el transformador reporte el error
                    pass

            # Ejecutar transformación
            wb_resultado, nombre_descarga = transformador.transformar(
                archivo_origen=self.archivo_actual.ruta,
                archivo_plantilla=ruta_plantilla_elegida,
                poliza_info=poliza_config
            )
            # Avance después de transformar exitosamente
            try:
                self._set_progress(80)
            except Exception:
                pass
            
            # Guardar en archivo temporal
            temp_dir = tempfile.gettempdir()
            ruta_temp = os.path.join(temp_dir, nombre_descarga)
            
            wb_resultado.save(ruta_temp)
            try:
                self._set_progress(90)
            except Exception:
                pass
            
            # Establecer archivo para descargar con nombre sugerido
            try:
                if hasattr(self.vista, 'establecer_archivo_resultado'):
                    self._en_ui(self.vista.establecer_archivo_resultado, ruta_temp, nombre_descarga)
                elif hasattr(self.vista, 'set_archivo_resultado_temp'):
                    self._en_ui(self.vista.set_archivo_resultado_temp, ruta_temp, nombre_descarga)
            except Exception:
                pass
            # Resaltar descarga y bloquear transformar
            try:
                try:
                    self._en_ui(self.vista.resaltar_descargar)
                except Exception:
                    try:
                        self._en_ui(self.vista.highlight_descargar)
                    except Exception:
                        pass
            except Exception:
                pass
            
            self._add_msg(f"✓ Archivo preparado: {nombre_descarga}\n")
            self._add_msg("\nHaz clic en 'Descargar Resultado' para elegir dónde guardarlo\n")
            self._add_msg("\n🎉 ¡Transformación completada exitosamente!\n")
            try:
                self._set_progress(100)
            except Exception:
                pass
            
        except Exception as e:
            import traceback
            error_detalle = traceback.format_exc()
            self._add_msg(f"\n✗ Error: {str(e)}\n")
            self._add_msg(f"\nDetalle:\n{error_detalle}\n")
            self._en_ui(self.vista.mostrar_error, "Error", str(e))
            try:
                self._set_progress(0)
            except Exception:
                pass
            # Habilitar controles en caso de error
            try:
                self._en_ui(self.vista.habilitar_controles, True)
            except Exception:
                pass
    
    def _mostrar_mensaje(self, mensaje):
        """Muestra mensaje en la vista"""
        self._add_msg(mensaje + "\n")
    
    def descargar_archivo(self, ruta_origen, ruta_destino):
        """Descarga el archivo a la ubicación seleccionada y abre en la carpeta correcta"""
        import shutil
        import subprocess
        import os
        import time
        
        if not ruta_origen or not os.path.exists(ruta_origen):
            self.vista.mostrar_error("Error", "Archivo temporal no encontrado")
            return
        
        try:
            # Copiar archivo a ubicación seleccionada
            shutil.copy2(ruta_origen, ruta_destino)
            self._add_msg(f"✓ Archivo guardado en:\n{ruta_destino}\n")
            
            # Esperar a que el archivo esté completamente escrito
            time.sleep(0.5)

            # Mostrar mensaje de éxito primero; abrir Explorer SOLO después de aceptar
            ruta_destino = os.path.abspath(ruta_destino)
            carpeta = os.path.dirname(ruta_destino)
            self.vista.mostrar_exito("Éxito", f"Archivo guardado en:\n{ruta_destino}")

            # Ahora abrir la carpeta con el archivo seleccionado
            if os.name == 'nt':  # Windows
                try:
                    subprocess.Popen(f'explorer /select,"{ruta_destino}"', shell=True)
                except Exception:
                    subprocess.Popen(f'explorer "{carpeta}"', shell=True)
            elif os.name == 'posix':  # Mac/Linux
                subprocess.Popen(['open', '-R', ruta_destino])

            self._add_msg(f"\n✓ Abriendo carpeta: {carpeta}\n")
            # Resaltar "Analizar Otro Archivo" y deshabilitar Descargar
            try:
                self.vista.resaltar_analizar_otro()
            except Exception:
                try:
                    self.vista.highlight_analizar()
                except Exception:
                    pass
            
        except Exception as e:
            self.vista.mostrar_error("Error", f"No se pudo guardar el archivo:\n{str(e)}")

    # ===== Helpers for dual UI =====
    def _add_msg(self, msg):
        try:
            if hasattr(self.vista, 'agregar_mensaje'):
                self._en_ui(self.vista.agregar_mensaje, msg)
            elif hasattr(self.vista, 'add_message'):
                self._en_ui(self.vista.add_message, msg)
        except Exception:
            pass

    def _set_progress(self, value):
        try:
            if hasattr(self.vista, 'establecer_progreso'):
                self._en_ui(self.vista.establecer_progreso, value)
            elif hasattr(self.vista, 'set_progress'):
                self._en_ui(self.vista.set_progress, value)
        except Exception:
            pass
            


