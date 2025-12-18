#!/usr/bin/env python3
"""Test sin Qt para ver dónde se queda"""

import sys
import os
sys.path.insert(0, os.getcwd())

from src.modelo import TransformadorDatos
from src.config.polizas import CONFIGURACION_POLIZAS

print("[TEST] Iniciando test...")

# Encontrar archivo 413
archivo_413 = None
for root, dirs, files in os.walk('.'):
    for file in files:
        if '413' in file and file.endswith('.xlsx'):
            archivo_413 = os.path.join(root, file)
            break

if not archivo_413:
    print("[ERROR] No se encontró archivo 413")
    sys.exit(1)

print(f"[TEST] Archivo origen: {archivo_413}")

# Obtener config DV
poliza_config = CONFIGURACION_POLIZAS.get('DV')
print(f"[TEST] Config DV: {poliza_config}")

# Plantilla
plantilla = 'src/plantillas/plantilla5852.xlsx'
if not os.path.exists(plantilla):
    print(f"[ERROR] Plantilla no existe: {plantilla}")
    sys.exit(1)

print(f"[TEST] Plantilla: {plantilla}")

# Crear transformador
def callback(msg):
    print(f"[CB] {msg}")

transformador = TransformadorDatos(callback_mensaje=callback)
print("[TEST] Transformador creado")

# Validar archivo origen
print("[DEBUG] Validando hoja requerida...")
hoja_requerida = poliza_config.get('hoja_origen_requerida')
print(f"[DEBUG] hoja_requerida: {hoja_requerida}")

if hoja_requerida:
    print("[DEBUG] Leyendo archivo origen...")
    from openpyxl import load_workbook
    print("[DEBUG] Abriendo workbook...")
    wb_origen = load_workbook(archivo_413, read_only=True, data_only=True)
    print(f"[DEBUG] Hojas: {wb_origen.sheetnames}")
    wb_origen.close()
    print("[DEBUG] Workbook cerrado")

# Transformar
print("[DEBUG] Antes de llamar transformador.transformar()")
try:
    print("[DEBUG] Llamando transformador.transformar()...")
    wb_resultado, nombre_descarga = transformador.transformar(
        archivo_origen=archivo_413,
        archivo_plantilla=plantilla,
        poliza_info=poliza_config
    )
    print(f"[DEBUG] ✓ Transformación completada: {nombre_descarga}")
except Exception as e:
    print(f"[ERROR] {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("[TEST] ✓ Éxito")
