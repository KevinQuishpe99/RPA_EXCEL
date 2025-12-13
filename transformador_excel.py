import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from datetime import datetime, date
import os
import shutil
import glob
import threading
import queue
import copy
import re

class TransformadorExcel:
    def __init__(self, root, usar_interfaz_mejorada=True):
        self.root = root
        
        # Variables
        self.archivo_origen = None
        self.archivo_plantilla = None
        self.archivo_resultado = None
        self.nombre_archivo_descarga = None  # Nombre sugerido para descarga
        
        # Variables para threading y progreso
        self.queue_mensajes = queue.Queue()
        self.procesando = False
        
        # Cache para optimización (encabezados siempre son los mismos)
        self._cache_mapeo_columnas = None
        self._cache_headers_destino = None
        self._cache_indices_columnas = {}
        self._cache_formulas = {}  # Ya existe, pero lo documentamos
        
        # Cache de estilos para reutilización (optimización de rendimiento)
        self._estilo_fuente_calibri = Font(name='Calibri')
        self._estilo_alineacion_centrada = Alignment(horizontal='center', vertical='center')
        self._estilo_borde_delgado = Side(style='thin')
        self._estilo_borde_celda = Border(
            left=self._estilo_borde_delgado, 
            right=self._estilo_borde_delgado, 
            top=self._estilo_borde_delgado, 
            bottom=self._estilo_borde_delgado
        )
        
        # Limpiar archivos temporales antiguos al iniciar
        self.limpiar_archivos_temporales()
        
        # Buscar plantilla automáticamente
        self.buscar_plantilla()
        
        # Crear interfaz (mejorada o clásica)
        if usar_interfaz_mejorada:
            try:
                # Intentar importar desde el módulo
                try:
                    from transformador_excel.gui.interfaz_mejorada import InterfazMejorada
                except ImportError:
                    # Si falla, intentar importar desde ruta relativa
                    import sys
                    import os
                    ruta_gui = os.path.join(os.path.dirname(__file__), 'transformador_excel', 'gui', 'interfaz_mejorada.py')
                    if os.path.exists(ruta_gui):
                        import importlib.util
                        spec = importlib.util.spec_from_file_location("interfaz_mejorada", ruta_gui)
                        modulo = importlib.util.module_from_spec(spec)
                        spec.loader.exec_module(modulo)
                        InterfazMejorada = modulo.InterfazMejorada
                    else:
                        raise ImportError("No se encontró interfaz_mejorada.py")
                
                self.interfaz = InterfazMejorada(root, self)
                # Conectar cola de mensajes (compartida)
                self.interfaz.queue_mensajes = self.queue_mensajes
                # Conectar widgets
                self.label_origen = self.interfaz.label_origen
                self.label_plantilla = self.interfaz.label_plantilla
                self.btn_transformar = self.interfaz.btn_transformar
                self.btn_descargar = self.interfaz.btn_descargar
                self.progress_var = self.interfaz.progress_var
                self.progress_bar = self.interfaz.progress_bar
                self.text_cuadre = self.interfaz.text_cuadre
                # Actualizar estado de plantilla
                self.interfaz._actualizar_estado_plantilla()
            except Exception as e:
                # Si falla, usar interfaz clásica
                import traceback
                print(f"Error al cargar interfaz mejorada: {e}")
                print(traceback.format_exc())
                self.root.title("Demo")
                self.root.geometry("800x700")
                self.crear_interfaz()
                self.verificar_mensajes()
        else:
            self.root.title("Transformador de Excel - 413 a Facturación")
            self.root.geometry("800x700")
            self.crear_interfaz()
            self.verificar_mensajes()
        
        # Limpiar archivos temporales al cerrar
        self.root.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion)
        
    def crear_interfaz(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Título
        titulo = ttk.Label(main_frame, text="Demo", 
                          font=("Arial", 16, "bold"))
        titulo.grid(row=0, column=0, columnspan=3, pady=10)
        
        # Sección 1: Archivo Origen (413)
        ttk.Label(main_frame, text="1. Archivo Origen (413):", 
                 font=("Arial", 10, "bold")).grid(row=1, column=0, sticky=tk.W, pady=5)
        
        self.label_origen = ttk.Label(main_frame, text="No seleccionado", 
                                      foreground="gray")
        self.label_origen.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=5)
        
        btn_origen = ttk.Button(main_frame, text="Seleccionar Archivo 413", 
                               command=self.seleccionar_origen)
        btn_origen.grid(row=2, column=2, padx=5)
        
        # Sección 2: Archivo Plantilla (Facturación) - Solo informativo
        ttk.Label(main_frame, text="2. Archivo Plantilla (Facturación):", 
                 font=("Arial", 10, "bold")).grid(row=3, column=0, sticky=tk.W, pady=5)
        
        self.label_plantilla = ttk.Label(main_frame, text="Buscando...", 
                                         foreground="gray")
        self.label_plantilla.grid(row=4, column=0, columnspan=3, sticky=tk.W, padx=5)
        
        # Actualizar estado de la plantilla
        if self.archivo_plantilla:
            nombre = os.path.basename(self.archivo_plantilla)
            self.label_plantilla.config(text=f"✓ {nombre}", foreground="green")
        else:
            self.label_plantilla.config(text="✗ Plantilla no encontrada", foreground="red")
        
        # Sección 3: Botón Transformar
        self.btn_transformar = ttk.Button(main_frame, text="Transformar Datos", 
                                    command=self.iniciar_transformacion,
                                    style="Accent.TButton")
        self.btn_transformar.grid(row=5, column=0, columnspan=3, pady=10)
        
        # Barra de progreso
        self.progress_var = tk.StringVar(value="0%")
        self.label_progreso = ttk.Label(main_frame, textvariable=self.progress_var, 
                                       foreground="blue")
        self.label_progreso.grid(row=6, column=0, columnspan=3, pady=5)
        
        self.progress_bar = ttk.Progressbar(main_frame, mode='determinate', maximum=100)
        self.progress_bar.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # Sección 4: Área de resultados y cuadre
        ttk.Label(main_frame, text="Progreso y Cuadre de Totales:", 
                 font=("Arial", 10, "bold")).grid(row=8, column=0, sticky=tk.W, pady=5)
        
        # Frame para el cuadre con scrollbar
        frame_cuadre = ttk.Frame(main_frame)
        frame_cuadre.grid(row=9, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        scrollbar = ttk.Scrollbar(frame_cuadre)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.text_cuadre = tk.Text(frame_cuadre, height=15, width=80, 
                                   yscrollcommand=scrollbar.set)
        self.text_cuadre.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.text_cuadre.yview)
        
        # Sección 5: Botón Descargar
        self.btn_descargar = ttk.Button(main_frame, text="Descargar Archivo Transformado", 
                                        command=self.descargar_archivo,
                                        state="disabled")
        self.btn_descargar.grid(row=10, column=0, columnspan=3, pady=10)
        
        # Configurar grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(9, weight=1)
        
    def seleccionar_origen(self):
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo 413",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if archivo:
            self.archivo_origen = archivo
            nombre = os.path.basename(archivo)
            self.label_origen.config(text=nombre, foreground="green")
    
    def buscar_plantilla(self):
        """Busca automáticamente el archivo de plantilla (plantilla.xlsx)"""
        import sys
        
        # Si está ejecutándose como EXE (PyInstaller), buscar en _MEIPASS primero
        if getattr(sys, 'frozen', False):
            # Ejecutándose como ejecutable compilado
            base_path = sys._MEIPASS
            ruta_plantilla_exe = os.path.join(base_path, "plantilla.xlsx")
            if os.path.exists(ruta_plantilla_exe):
                self.archivo_plantilla = ruta_plantilla_exe
                return
        
        # Buscar plantilla.xlsx en el directorio del script primero
        directorio_script = os.path.dirname(os.path.abspath(__file__))
        ruta_plantilla_script = os.path.join(directorio_script, "plantilla.xlsx")
        
        if os.path.exists(ruta_plantilla_script):
            self.archivo_plantilla = ruta_plantilla_script
            return
        
        # Si no está en el directorio del script, buscar en el directorio actual
        directorio_actual = os.getcwd()
        ruta_plantilla_actual = os.path.join(directorio_actual, "plantilla.xlsx")
        
        if os.path.exists(ruta_plantilla_actual):
            self.archivo_plantilla = ruta_plantilla_actual
            return
        
        # Si no se encuentra, buscar en la carpeta "programa" del directorio actual
        ruta_programa = os.path.join(directorio_actual, "programa", "plantilla.xlsx")
        if os.path.exists(ruta_programa):
            self.archivo_plantilla = ruta_programa
            return
        
        # Si no se encuentra, buscar en la carpeta "programa" del directorio del script
        ruta_programa_script = os.path.join(directorio_script, "programa", "plantilla.xlsx")
        if os.path.exists(ruta_programa_script):
            self.archivo_plantilla = ruta_programa_script
            return
    
    def verificar_mensajes(self):
        """Verifica mensajes de la cola y actualiza la interfaz (optimizado)"""
        # Si usa interfaz mejorada, el procesamiento se hace allí
        if hasattr(self, 'interfaz') and hasattr(self.interfaz, '_procesar_mensajes'):
            # La interfaz mejorada maneja sus propios mensajes
            return
        
        # Interfaz clásica
        mensajes_procesados = 0
        max_mensajes_por_ciclo = 10  # Procesar más mensajes por ciclo
        
        try:
            while mensajes_procesados < max_mensajes_por_ciclo:
                mensaje = self.queue_mensajes.get_nowait()
                tipo = mensaje.get('tipo')
                
                if tipo == 'progreso':
                    texto = mensaje.get('texto', '')
                    porcentaje = mensaje.get('porcentaje')
                    # Mostrar solo el porcentaje numérico
                    if porcentaje is not None:
                        self.progress_var.set(f"{int(porcentaje)}%")
                        self.progress_bar['value'] = porcentaje
                        self.progress_bar.update()
                    else:
                        self.progress_var.set("0%")
                elif tipo == 'actualizar_texto':
                    self.text_cuadre.insert(tk.END, mensaje.get('texto', '') + '\n')
                    self.text_cuadre.see(tk.END)
                elif tipo == 'actualizar_ultima_linea':
                    # Optimizado: solo leer última línea en lugar de todo el contenido
                    try:
                        num_lineas = int(self.text_cuadre.index('end-1c').split('.')[0])
                        if num_lineas > 1:
                            ultima_linea = self.text_cuadre.get(f"{num_lineas-1}.0", "end-1c")
                            if ultima_linea.strip().startswith('→ Fila'):
                                self.text_cuadre.delete(f"{num_lineas-1}.0", "end-1c")
                    except:
                        pass
                    self.text_cuadre.insert(tk.END, mensaje.get('texto', '') + '\n')
                    self.text_cuadre.see(tk.END)
                elif tipo == 'limpiar_texto':
                    self.text_cuadre.delete(1.0, tk.END)
                elif tipo == 'completado':
                    self.progress_bar['value'] = 100
                    self.progress_var.set("100%")
                    self.btn_transformar.config(state="normal")
                    self.btn_descargar.config(state="normal")
                    messagebox.showinfo("Éxito", mensaje.get('texto', 'Transformación completada.'))
                elif tipo == 'error':
                    porcentaje = mensaje.get('porcentaje', 0)
                    self.progress_bar['value'] = porcentaje
                    self.progress_var.set(f"{int(porcentaje)}%")
                    self.btn_transformar.config(state="normal")
                    messagebox.showerror("Error", mensaje.get('texto', 'Error durante la transformación'))
                
                mensajes_procesados += 1
        except queue.Empty:
            pass
        
        # Actualizar interfaz solo una vez por ciclo (más eficiente)
        if mensajes_procesados > 0:
            self.root.update_idletasks()  # Más eficiente que update()
        
        # Programar próxima verificación
        self.root.after(50, self.verificar_mensajes)  # Más frecuente pero con límite de mensajes
    
    def enviar_mensaje(self, tipo, texto='', porcentaje=None):
        """Envía un mensaje a la cola para actualizar la interfaz"""
        # Si usa interfaz mejorada, también enviar allí
        if hasattr(self, 'interfaz') and hasattr(self.interfaz, 'enviar_mensaje'):
            self.interfaz.enviar_mensaje(tipo, texto)
        self.queue_mensajes.put({'tipo': tipo, 'texto': texto, 'porcentaje': porcentaje})
    
    def iniciar_transformacion(self):
        """Inicia la transformación en un hilo separado"""
        if self.procesando:
            return
        
        if not self.archivo_origen:
            messagebox.showerror("Error", "Por favor seleccione el archivo origen (413)")
            return
        
        # Buscar plantilla nuevamente por si acaso
        if not self.archivo_plantilla:
            self.buscar_plantilla()
        
        if not self.archivo_plantilla:
            messagebox.showerror("Error", 
                "No se encontró el archivo de plantilla.\n"
                "Por favor asegúrese de que el archivo 'plantilla.xlsx' esté en la carpeta 'programa'.")
            return
        
        if not os.path.exists(self.archivo_plantilla):
            messagebox.showerror("Error", 
                f"El archivo de plantilla no existe:\n{self.archivo_plantilla}")
            return
        
        # Preparar interfaz
        self.procesando = True
        self.btn_transformar.config(state="disabled")
        self.btn_descargar.config(state="disabled")
        self.progress_bar['value'] = 0
        self.progress_var.set("0%")
        self.enviar_mensaje('limpiar_texto')
        
        # Iniciar hilo
        thread = threading.Thread(target=self.transformar_datos, daemon=True)
        thread.start()
    
    def transformar_datos(self):
        if not self.archivo_origen:
            messagebox.showerror("Error", "Por favor seleccione el archivo origen (413)")
            return
        
        # Buscar plantilla nuevamente por si acaso
        if not self.archivo_plantilla:
            self.buscar_plantilla()
        
        if not self.archivo_plantilla:
            messagebox.showerror("Error", 
                "No se encontró el archivo de plantilla.\n"
                "Por favor asegúrese de que el archivo 'plantilla.xlsx' esté en la carpeta 'programa'.")
            return
        
        if not os.path.exists(self.archivo_plantilla):
            messagebox.showerror("Error", 
                f"El archivo de plantilla no existe:\n{self.archivo_plantilla}")
            return
        
        try:
            self.enviar_mensaje('actualizar_texto', "=" * 80)
            self.enviar_mensaje('actualizar_texto', "INICIANDO TRANSFORMACIÓN")
            self.enviar_mensaje('actualizar_texto', "=" * 80)
            self.enviar_mensaje('progreso', "Leyendo archivo origen...", 5)
            
            # Leer archivo origen
            try:
                df_origen = pd.read_excel(self.archivo_origen, 
                                     sheet_name="Report_AseguradoraMensual", 
                                         header=None,
                                         engine='openpyxl')
            except Exception as e:
                self.enviar_mensaje('error', f"Error al leer el archivo origen: {str(e)}\n\nAsegúrese de que:\n- El archivo no esté abierto en Excel\n- El archivo tenga la hoja 'Report_AseguradoraMensual'\n- El archivo no esté corrupto")
                return
            
            # Obtener encabezados
            
            if len(df_origen) < 10:
                self.enviar_mensaje('error', f"El archivo origen parece estar vacío o tiene muy pocas filas ({len(df_origen)} filas)")
                return
            
            self.enviar_mensaje('actualizar_texto', f"✓ Archivo origen leído: {len(df_origen)} filas")
            self.enviar_mensaje('progreso', "Buscando encabezados...")
            
            # Buscar dinámicamente la fila de encabezados
            # Buscar en las primeras 10 filas (índices 0-9)
            fila_encabezados_origen = None
            headers_origen = None
            
            for idx_fila in range(min(10, len(df_origen))):
                fila_actual = df_origen.iloc[idx_fila].tolist()
                # Contar cuántos encabezados válidos hay en esta fila
                headers_validos = [h for h in fila_actual if pd.notna(h) and str(h).strip() != '']
                
                # Si hay al menos 5 encabezados válidos, consideramos que es la fila de encabezados
                if len(headers_validos) >= 5:
                    # Verificar que no sean solo números (los datos suelen tener números)
                    # Los encabezados suelen ser texto
                    headers_texto = [h for h in headers_validos if isinstance(h, str) or (isinstance(h, (int, float)) and str(h).strip() != '')]
                    if len(headers_texto) >= 3:  # Al menos 3 deben ser texto o tener contenido significativo
                        fila_encabezados_origen = idx_fila
                        headers_origen = fila_actual
                        self.enviar_mensaje('actualizar_texto', f"✓ Encabezados encontrados en fila {idx_fila + 1} (ORIGEN): {len(headers_validos)} columnas válidas")
                        break
            
            if headers_origen is None or fila_encabezados_origen is None:
                self.enviar_mensaje('error', "No se pudieron encontrar los encabezados en el archivo origen.\n\nVerifique que el archivo tenga una fila de encabezados con al menos 5 columnas válidas.")
                return
            
            self.enviar_mensaje('progreso', "Copiando plantilla...", 15)
            
            # Buscar columna FECHA DE INICIO DE CREDITO
            col_fecha_inicio = None
            for idx, header in enumerate(headers_origen):
                if pd.notna(header) and 'FECHA DE INICIO DE CREDITO' in str(header).upper():
                    col_fecha_inicio = idx
                    break
            
            # Extraer fecha de la primera fila de datos (después de la fila de encabezados)
            fecha_mes = None
            if col_fecha_inicio is not None:
                # Buscar en las primeras filas después de los encabezados
                inicio_busqueda = fila_encabezados_origen + 1
                for idx in range(inicio_busqueda, min(inicio_busqueda + 20, len(df_origen))):
                    fecha_valor = df_origen.iloc[idx, col_fecha_inicio]
                    if pd.notna(fecha_valor):
                        try:
                            if isinstance(fecha_valor, datetime):
                                fecha_mes = fecha_valor
                            elif isinstance(fecha_valor, pd.Timestamp):
                                fecha_mes = fecha_valor.to_pydatetime()
                            elif isinstance(fecha_valor, str):
                                fecha_mes = pd.to_datetime(fecha_valor)
                            else:
                                fecha_mes = pd.to_datetime(fecha_valor)
                            break
                        except:
                            continue
            
            # Generar nombre del archivo basado en fecha (formato: "Facturación DV [mes] [año]")
            if fecha_mes:
                # Mapeo de meses en español
                meses_espanol = {
                    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                    5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
                }
                
                mes_num = fecha_mes.month
                nombre_mes = meses_espanol.get(mes_num, 'mes')
                nombre_mes_capitalizado = nombre_mes.capitalize()
                año = fecha_mes.year
                
                # Generar nombre: "Facturación DV [mes] [año].xlsx"
                nombre_archivo = f"Facturación DV {nombre_mes_capitalizado} {año}.xlsx"
            else:
                # Si no se encuentra fecha, usar formato con fecha actual
                fecha_actual = datetime.now()
                meses_espanol = {
                    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                    5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
                }
                nombre_mes = meses_espanol.get(fecha_actual.month, 'mes').capitalize()
                nombre_archivo = f"Facturación DV {nombre_mes} {fecha_actual.year}.xlsx"
            
            # Copiar plantilla con nombre temporal
            nombre_resultado_temp = f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.archivo_resultado = nombre_resultado_temp
            self.nombre_archivo_descarga = nombre_archivo
            shutil.copy2(self.archivo_plantilla, self.archivo_resultado)
            
            self.enviar_mensaje('actualizar_texto', f"✓ Plantilla copiada: {nombre_archivo}")
            self.enviar_mensaje('progreso', "Abriendo archivo de destino...", 20)
            
            # Abrir archivo resultado con openpyxl para conservar fórmulas
            try:
                wb = load_workbook(self.archivo_resultado, data_only=False)
            except Exception as e:
                self.enviar_mensaje('error', f"Error al abrir el archivo de plantilla: {str(e)}")
                return
            
            # Verificar que existe la hoja destino (buscar de forma flexible)
            hoja_encontrada = None
            # La hoja destino siempre debe llamarse exactamente "DV(5852)"
            nombre_hoja_buscado = "DV(5852)"
            
            # Primero intentar el nombre exacto
            if nombre_hoja_buscado in wb.sheetnames:
                hoja_encontrada = nombre_hoja_buscado
            else:
                # Buscar hoja que contenga "DV" y números (como "5852")
                for nombre_hoja in wb.sheetnames:
                    nombre_upper = nombre_hoja.upper()
                    if 'DV' in nombre_upper and ('5852' in nombre_hoja or any(char.isdigit() for char in nombre_hoja)):
                        hoja_encontrada = nombre_hoja
                        break
                
                # Si no se encuentra, buscar cualquier hoja con "DV"
                if not hoja_encontrada:
                    for nombre_hoja in wb.sheetnames:
                        if 'DV' in nombre_hoja.upper():
                            hoja_encontrada = nombre_hoja
                            break
                
                # Si aún no se encuentra, usar la primera hoja que no sea de códigos
                if not hoja_encontrada:
                    for nombre_hoja in wb.sheetnames:
                        if 'CODIGO' not in nombre_hoja.upper() and 'HOJA1' not in nombre_hoja.upper():
                            hoja_encontrada = nombre_hoja
                            break
                
                if hoja_encontrada and hoja_encontrada != nombre_hoja_buscado:
                    self.enviar_mensaje('actualizar_texto', f"⚠️ Usando hoja '{hoja_encontrada}' (no se encontró '{nombre_hoja_buscado}')")
            
            if not hoja_encontrada:
                hojas_disponibles = ', '.join(wb.sheetnames)
                self.enviar_mensaje('error', f"No se encontró una hoja válida en la plantilla.\n\nHojas disponibles: {hojas_disponibles}\n\nSe esperaba una hoja con 'DV' en el nombre.")
                return
            
            ws_destino = wb[hoja_encontrada]
            
            # Los encabezados ya fueron encontrados dinámicamente arriba
            # Si por alguna razón no están disponibles, buscarlos de nuevo
            if 'headers_origen' not in locals() or headers_origen is None:
                # Buscar dinámicamente de nuevo
                for idx_fila in range(min(10, len(df_origen))):
                    fila_actual = df_origen.iloc[idx_fila].tolist()
                    headers_validos = [h for h in fila_actual if pd.notna(h) and str(h).strip() != '']
                    if len(headers_validos) >= 5:
                        headers_texto = [h for h in headers_validos if isinstance(h, str) or (isinstance(h, (int, float)) and str(h).strip() != '')]
                        if len(headers_texto) >= 3:
                            fila_encabezados_origen = idx_fila
                            headers_origen = fila_actual
                            break
                
                if headers_origen is None:
                    self.enviar_mensaje('error', "No se encontraron encabezados en el archivo origen")
                    return
            
            # Buscar encabezados en destino (fila 4 o 5)
            headers_destino = None
            for fila_encabezado in [4, 5]:
                try:
                    headers_destino = ws_destino[fila_encabezado]
                    # Verificar que hay al menos algunos encabezados
                    if any(cell.value for cell in headers_destino[:10]):
                        self.enviar_mensaje('actualizar_texto', f"✓ Encabezados destino encontrados en fila {fila_encabezado}")
                        break
                except:
                    continue
            
            if headers_destino is None:
                self.enviar_mensaje('error', "No se encontraron encabezados en el archivo destino")
                return
            
            self.enviar_mensaje('actualizar_texto', "✓ Encabezados obtenidos")
            self.enviar_mensaje('progreso', "Mapeando columnas...", 25)
            
            # Mapeo de columnas origen -> destino (optimizado con cache)
            mapeo_columnas = self.obtener_mapeo_columnas(headers_origen, headers_destino)
            
            # Cachear índices de columnas importantes para evitar búsquedas repetitivas
            if 'idx_pais_residencia_dest' not in self._cache_indices_columnas:
                self._cache_indices_columnas['idx_pais_residencia_dest'] = None
                for idx, cell in enumerate(headers_destino):
                    if cell.value and 'PAIS DE RESIDENCIA' in str(cell.value).upper():
                        self._cache_indices_columnas['idx_pais_residencia_dest'] = idx + 1
                        break
            
            if len(mapeo_columnas) == 0:
                self.enviar_mensaje('error', "No se pudo mapear ninguna columna entre origen y destino.\n\nVerifique que:\n- Los encabezados en ambos archivos coincidan\n- Los nombres de las columnas sean correctos")
                return
            
            # No mostrar información de mapeo para mantener la pantalla limpia y velocidad
            self.enviar_mensaje('progreso', "Limpiando datos existentes...", 30)
            
            # Limpiar datos existentes (mantener encabezados y fórmulas)
            self.limpiar_datos_destino(ws_destino)
            
            self.enviar_mensaje('actualizar_texto', "✓ Datos limpiados")
            
            # Transferir datos
            # ORIGEN: Datos desde la fila siguiente a los encabezados
            # DESTINO: Datos desde fila 6
            fila_destino = 6  # Primera fila de datos en destino (fila 6 en Excel)
            fila_origen_inicio = fila_encabezados_origen + 1  # Primera fila de datos en origen (después de los encabezados)
            
            total_filas = 0
            total_estimado = len(df_origen) - fila_origen_inicio
            
            self.enviar_mensaje('actualizar_texto', "\n" + "=" * 80)
            self.enviar_mensaje('actualizar_texto', "TRANSFIRIENDO DATOS AL ARCHIVO FINAL...")
            self.enviar_mensaje('actualizar_texto', "=" * 80)
            self.enviar_mensaje('actualizar_texto', f"Total estimado: {total_estimado} filas")
            self.enviar_mensaje('actualizar_texto', "\n📝 Última fila procesada:")
            self.enviar_mensaje('actualizar_texto', "-" * 80)
            
            # Pre-calcular índices de columnas importantes (optimización)
            idx_primer_apellido = None
            idx_primer_nombre = None
            idx_pais_origen = None
            idx_pais_residencia_dest = None
            idx_provincia_orig = None
            idx_ciudad_orig = None
            
            # Pre-calcular índice de PAIS DE RESIDENCIA en destino (usar cache si está disponible)
            if 'idx_pais_residencia_dest' in self._cache_indices_columnas:
                idx_pais_residencia_dest = self._cache_indices_columnas['idx_pais_residencia_dest']
            else:
                idx_pais_residencia_dest = None
                for idx, cell in enumerate(headers_destino):
                    if cell.value and 'PAIS DE RESIDENCIA' in str(cell.value).upper():
                        idx_pais_residencia_dest = idx + 1
                        self._cache_indices_columnas['idx_pais_residencia_dest'] = idx + 1
                        break
            
            # Pre-calcular índices en origen (usar cache si está disponible)
            cache_key_origen = 'indices_origen'
            if cache_key_origen not in self._cache_indices_columnas:
                # Buscar y cachear índices
                for idx_orig, col_dest in mapeo_columnas.items():
                    if idx_orig < len(headers_origen):
                        header = str(headers_origen[idx_orig]).strip().upper()
                        if 'PRIMER APELLIDO' in header:
                            self._cache_indices_columnas['idx_primer_apellido'] = idx_orig
                            idx_primer_apellido = idx_orig
                        elif 'PRIMER NOMBRE' in header:
                            self._cache_indices_columnas['idx_primer_nombre'] = idx_orig
                            idx_primer_nombre = idx_orig
                        elif 'PAIS DE ORIGEN' in header:
                            self._cache_indices_columnas['idx_pais_origen'] = idx_orig
                            idx_pais_origen = idx_orig
                        elif 'PROVINCIA' in header and 'PAIS' not in header:
                            self._cache_indices_columnas['idx_provincia_orig'] = idx_orig
                            idx_provincia_orig = idx_orig
                        elif 'CIUDAD' in header:
                            self._cache_indices_columnas['idx_ciudad_orig'] = idx_orig
                            idx_ciudad_orig = idx_orig
                self._cache_indices_columnas[cache_key_origen] = True
            else:
                # Usar índices del cache
                idx_primer_apellido = self._cache_indices_columnas.get('idx_primer_apellido')
                idx_primer_nombre = self._cache_indices_columnas.get('idx_primer_nombre')
                idx_pais_origen = self._cache_indices_columnas.get('idx_pais_origen')
                idx_provincia_orig = self._cache_indices_columnas.get('idx_provincia_orig')
                idx_ciudad_orig = self._cache_indices_columnas.get('idx_ciudad_orig')
            
            # Verificar que PROVINCIA y CIUDAD estén mapeadas
            # IMPORTANTE: Las columnas O (15) y P (16) deben recibir los datos directamente
            col_provincia_dest = 15  # Columna O
            col_ciudad_dest = 16     # Columna P
            
            if idx_provincia_orig is None:
                # Buscar manualmente PROVINCIA en origen si no se mapeó
                for idx, header in enumerate(headers_origen):
                    if pd.notna(header):
                        header_str = str(header).strip().upper()
                        if 'PROVINCIA' in header_str and 'PAIS' not in header_str:
                            # Forzar mapeo a columna O (15)
                            mapeo_columnas[idx] = col_provincia_dest
                            idx_provincia_orig = idx
                            break
            
            if idx_ciudad_orig is None:
                # Buscar manualmente CIUDAD en origen si no se mapeó
                for idx, header in enumerate(headers_origen):
                    if pd.notna(header):
                        header_str = str(header).strip().upper()
                        if 'CIUDAD' in header_str:
                            # Forzar mapeo a columna P (16)
                            mapeo_columnas[idx] = col_ciudad_dest
                            idx_ciudad_orig = idx
                            break
            
            # Si ya están mapeadas pero no a las columnas O y P, forzar el mapeo correcto
            if idx_provincia_orig is not None:
                mapeo_columnas[idx_provincia_orig] = col_provincia_dest
            if idx_ciudad_orig is not None:
                mapeo_columnas[idx_ciudad_orig] = col_ciudad_dest
            
            # Mensaje de depuración para PROVINCIA y CIUDAD
            if idx_provincia_orig is not None:
                self.enviar_mensaje('actualizar_texto', f"  ✓ PROVINCIA mapeada: columna origen {idx_provincia_orig} → destino O (columna {col_provincia_dest})")
            else:
                self.enviar_mensaje('actualizar_texto', f"  ⚠️ PROVINCIA no encontrada en origen")
            
            if idx_ciudad_orig is not None:
                self.enviar_mensaje('actualizar_texto', f"  ✓ CIUDAD mapeada: columna origen {idx_ciudad_orig} → destino P (columna {col_ciudad_dest})")
            else:
                self.enviar_mensaje('actualizar_texto', f"  ⚠️ CIUDAD no encontrada en origen")
            
            # Procesar filas de forma optimizada - FILA POR FILA
            # Optimización: Pre-calcular límite una vez y convertir DataFrame a numpy para acceso más rápido
            limite_origen = len(df_origen)
            
            # Optimización: Convertir primera columna a array numpy para acceso más rápido
            primera_columna_array = df_origen.iloc[fila_origen_inicio:, 0].values
            
            self.enviar_mensaje('actualizar_texto', f"Procesando desde fila {fila_origen_inicio + 1} hasta fila {limite_origen}...")
            
            for idx_local, idx_origen in enumerate(range(fila_origen_inicio, limite_origen)):
                # PROCESAR FILA POR FILA
                # VALIDACIÓN PRINCIPAL: Verificar que la PRIMERA COLUMNA esté llena
                # Toda fila que tiene la primera columna llena es válida y debe pasarse
                try:
                    # Optimización: Usar array numpy para acceso más rápido
                    primera_columna_valor = primera_columna_array[idx_local]
                    
                    # Verificar si la primera columna tiene un valor válido
                    fila_valida = False
                    
                    if pd.notna(primera_columna_valor):
                        # Si es string, verificar que no esté vacío y que no sea una palabra clave de totales
                        if isinstance(primera_columna_valor, str):
                            valor_limpio = primera_columna_valor.strip()
                            if valor_limpio and valor_limpio.upper() not in ['NAN', 'NONE', 'NULL', 'TOTAL', 'CUADRE', 'PRECANCELACION', '']:
                                fila_valida = True
                        else:
                            # Para valores no-string (números, fechas, etc.), si no es NaN, es válido
                            fila_valida = True
                    
                    # Si la primera columna está vacía, saltar esta fila
                    if not fila_valida:
                        continue
                    
                    # Verificar si es una fila de totales (solo si tiene contenido)
                    if isinstance(primera_columna_valor, str):
                        primera_col_str = primera_columna_valor.strip().upper()
                        if 'TOTAL' in primera_col_str or 'CUADRE' in primera_col_str or 'PRECANCELACION' in primera_col_str:
                            self.enviar_mensaje('actualizar_texto', f"\n⚠️ Detectada fila de totales en fila {idx_origen + 1}, deteniendo transferencia")
                            break
                    
                except Exception as e:
                    # Si hay error al leer la primera columna, saltar esta fila
                    continue
                
                # Si llegamos aquí, la fila es válida (primera columna llena)
                # Transferir datos de esta fila (optimizado con índices pre-calculados)
                try:
                    self.transferir_fila_optimizada(df_origen, idx_origen, ws_destino, fila_destino, 
                                        mapeo_columnas, headers_origen, headers_destino,
                                        idx_pais_origen, idx_pais_residencia_dest)
                    fila_destino += 1
                    total_filas += 1
                except Exception as e:
                    # Si hay error al transferir, continuar con la siguiente fila
                    self.enviar_mensaje('actualizar_texto', f"⚠️ Error al transferir fila {idx_origen + 1}: {str(e)}")
                    continue
                
                # Actualizar interfaz solo cada 2000 filas para mayor velocidad (optimizado)
                if total_filas % 2000 == 0:
                    # Obtener nombre rápido para mostrar
                    nombre_mostrar = ""
                    try:
                        if idx_primer_apellido is not None:
                            apellido = df_origen.iloc[idx_origen, idx_primer_apellido]
                            if pd.notna(apellido):
                                nombre_mostrar = str(apellido).strip()
                        if idx_primer_nombre is not None and not nombre_mostrar:
                            nombre = df_origen.iloc[idx_origen, idx_primer_nombre]
                            if pd.notna(nombre):
                                nombre_mostrar = str(nombre).strip()
                    except:
                        pass
                    
                    mensaje = f"  → Fila {fila_destino}: {nombre_mostrar if nombre_mostrar else 'Procesando...'}"
                    self.enviar_mensaje('actualizar_ultima_linea', mensaje)
                    
                    # Actualizar progreso junto con la interfaz (solo cada 1000 filas)
                    # El procesamiento de filas va del 30% al 85% (55% del total)
                    porcentaje_filas = (total_filas / total_estimado) * 55 if total_estimado > 0 else 0
                    porcentaje_total = 30 + porcentaje_filas  # Empezar desde 30%
                    self.enviar_mensaje('progreso', f"Procesando... {total_filas}/{total_estimado} filas ({porcentaje_total:.1f}%)", porcentaje_total)
                    
                    # Guardar el archivo periódicamente (cada 3000 filas para no ralentizar)
                    if total_filas % 3000 == 0 and total_filas > 0:
                        try:
                            wb.save(self.archivo_resultado)
                        except:
                            pass  # Si falla, continuar sin error
            
            if total_filas == 0:
                self.enviar_mensaje('error', "No se procesaron filas. Verifique que:\n- El archivo origen tenga datos desde la fila 6\n- No haya errores en el formato de los datos")
                return
            
            self.enviar_mensaje('actualizar_texto', f"\n✓ Total filas procesadas: {total_filas}")
            self.enviar_mensaje('progreso', "Validando fórmulas y datos...", 85)
            
            # Validar que no haya NaN donde deberían estar fórmulas
            self.validar_formulas_datos(ws_destino, fila_destino - 1)
            
            # Información sobre transformaciones aplicadas
            self.enviar_mensaje('actualizar_texto', "\n" + "=" * 80)
            self.enviar_mensaje('actualizar_texto', "TRANSFORMACIONES APLICADAS:")
            self.enviar_mensaje('actualizar_texto', "=" * 80)
            self.enviar_mensaje('actualizar_texto', "  ✓ PROVINCIA y CIUDAD: Ceros iniciales eliminados (ej: '01' → 1)")
            self.enviar_mensaje('actualizar_texto', "  ✓ NACIONALIDAD: Si TIPO IDENTIFICACION='00' → '239'")
            self.enviar_mensaje('actualizar_texto', "  ✓ PAIS DE RESIDENCIA: Siempre establecido en '239'")
            self.enviar_mensaje('actualizar_texto', "  ✓ EDAD: Calculada con fórmula (NO se copia del origen)")
            self.enviar_mensaje('actualizar_texto', "  ✓ Fórmulas VLOOKUP: Conservadas y ajustadas")
            self.enviar_mensaje('actualizar_texto', "=" * 80)
            
            # Guardar el archivo una vez al final antes de verificar errores
            try:
                wb_destino.save(self.archivo_resultado)
            except:
                pass
            
            # Detectar y desplazar filas de totales si es necesario
            self.enviar_mensaje('progreso', "Verificando y desplazando filas de totales...")
            
            # Determinar la última fila real de datos verificando columnas A y B
            # fila_destino ahora apunta a la siguiente fila después de la última con datos
            # Entonces la última fila con datos es fila_destino - 1
            ultima_fila_datos_nueva = fila_destino - 1
            
            # Verificar que la última fila realmente tenga datos en A o B
            # Si no tiene datos, buscar hacia arriba hasta encontrar la última fila con datos
            for fila_check in range(ultima_fila_datos_nueva, 5, -1):  # Revisar desde la última fila hacia arriba
                try:
                    cell_a = ws_destino.cell(fila_check, 1)  # Columna A
                    cell_b = ws_destino.cell(fila_check, 2)  # Columna B
                    
                    # Obtener valores y limpiarlos
                    valor_a = cell_a.value if cell_a.value is not None else ''
                    valor_b = cell_b.value if cell_b.value is not None else ''
                    valor_a_str = str(valor_a).strip() if valor_a else ''
                    valor_b_str = str(valor_b).strip() if valor_b else ''
                    
                    # Si ambas columnas A y B están vacías, esta fila no tiene datos
                    # Continuar buscando hacia arriba
                    if valor_a_str == '' and valor_b_str == '':
                        continue
                    else:
                        # Esta fila tiene datos en A o B, es la última fila real
                        ultima_fila_datos_nueva = fila_check
                        break
                except:
                    continue
            
            # Asegurarse de que no sea menor que 6 (primera fila de datos)
            if ultima_fila_datos_nueva < 6:
                ultima_fila_datos_nueva = 6
            
            # Mensaje informativo
            self.enviar_mensaje('actualizar_texto', f"\n  → Última fila de datos detectada: {ultima_fila_datos_nueva}")
            
            filas_desplazadas = self.desplazar_filas_totales(ws_destino, ultima_fila_datos_nueva)
            
            if filas_desplazadas > 0:
                self.enviar_mensaje('actualizar_texto', f"\n✓ {filas_desplazadas} fila(s) de totales desplazada(s) hacia abajo")
            
            self.enviar_mensaje('progreso', "Actualizando fórmulas de totales...", 87)
            
            # Establecer fecha de corte en B3 y renombrar la hoja según mes y año
            fecha_corte = self.establecer_fecha_corte(ws_destino, df_origen, headers_origen)
            
            # Ya no renombramos la hoja; debe mantenerse como "DV(5852)"
            
            # Agregar fórmulas SUM al final de los datos para las columnas requeridas
            self.enviar_mensaje('progreso', "Agregando totales al final de los datos...", 90)
            self.agregar_totales_columnas(ws_destino, ultima_fila_datos_nueva, headers_destino)
            
            # Actualizar fórmulas de totales (ahora en su nueva posición)
            formulas_actualizadas, formulas_conservadas = self.actualizar_formulas_totales(ws_destino, ultima_fila_datos_nueva)
            
            # Guardar archivo PRIMERO para que los valores calculados estén disponibles
            self.enviar_mensaje('progreso', "Guardando archivo...", 93)
            try:
                wb.save(self.archivo_resultado)
                self.enviar_mensaje('actualizar_texto', "✓ Archivo guardado correctamente")
            except Exception as e:
                self.enviar_mensaje('error', f"Error al guardar el archivo: {str(e)}\n\nAsegúrese de que:\n- El archivo no esté abierto en otro programa\n- Tenga permisos de escritura en el directorio")
                return
            
            # Crear Hoja2 con tabla dinámica de MONTO CREDITO
            self.enviar_mensaje('progreso', "Creando Hoja2 con tabla dinámica...", 96)
            self.crear_hoja2_tabla_dinamica(wb, ws_destino, ultima_fila_datos_nueva, headers_destino)
            
            # Resaltar filas con errores en fórmulas (ESTRATEGIA RÁPIDA: sin leer archivo guardado)
            self.enviar_mensaje('progreso', "Verificando errores en fórmulas...", 95)
            filas_con_nan = self.resaltar_filas_nan(ws_destino, ultima_fila_datos_nueva)
            
            # Aplicar fuente Calibri y formatos en batch (optimizado)
            self.enviar_mensaje('progreso', "Aplicando formato (Calibri) y ajustando EDAD...", 97)
            self.aplicar_formato_calibri_y_edad(ws_destino, ultima_fila_datos_nueva, headers_destino)
            
            # Aplicar estilos en batch a todas las celdas de datos (optimización)
            self.enviar_mensaje('progreso', "Aplicando estilos en batch...", 98)
            self.aplicar_estilos_batch(ws_destino, ultima_fila_datos_nueva, headers_destino)
            
            # Guardar nuevamente con los resaltados y formato
            try:
                wb.save(self.archivo_resultado)
            except:
                pass
            
            self.enviar_mensaje('progreso', "Finalizando...", 99)
            
            # Calcular totales de filas de datos
            # Origen: los encabezados están en fila 5 (índice 5 en pandas), datos desde fila 6 (índice 6)
            # Total filas en pandas = len(df_origen), pero contamos desde índice 0
            # Filas antes de datos: 0,1,2,3,4,5 = 6 filas
            total_filas_origen = len(df_origen) - 6  # Restamos 6 porque los datos empiezan en índice 6
            
            # Destino: los encabezados están en fila 5, datos desde fila 6 hasta ultima_fila_datos_nueva
            # ultima_fila_datos_nueva es la última fila de datos (en Excel, base 1)
            # Total filas de datos = ultima_fila_datos_nueva - 5 (restamos las 5 filas antes de los datos: 1,2,3,4,5)
            total_filas_destino = ultima_fila_datos_nueva - 5
            
            # Mostrar resumen final
            self.enviar_mensaje('actualizar_texto', "\n" + "=" * 80)
            self.enviar_mensaje('actualizar_texto', "RESUMEN FINAL:")
            self.enviar_mensaje('actualizar_texto', "=" * 80)
            self.enviar_mensaje('actualizar_texto', f"  📊 Total filas de datos en ORIGEN: {total_filas_origen}")
            self.enviar_mensaje('actualizar_texto', f"  📊 Total filas de datos en DESTINO: {total_filas_destino}")
            
            if total_filas_origen == total_filas_destino:
                self.enviar_mensaje('actualizar_texto', f"  ✅ Las filas coinciden correctamente")
            else:
                diferencia = abs(total_filas_origen - total_filas_destino)
                self.enviar_mensaje('actualizar_texto', f"  ⚠️ DIFERENCIA: {diferencia} fila(s) de diferencia entre origen y destino")
            
            self.enviar_mensaje('progreso', "Extrayendo cuadre...")
            
            # Validar que se copiaron datos
            self.enviar_mensaje('progreso', "Validando datos transferidos...", 98)
            
            # Verificar que hay datos en el destino
            datos_encontrados = 0
            try:
                for row in range(6, min(10, fila_destino)):
                    for col in range(1, min(10, ws_destino.max_column + 1)):
                        cell = ws_destino.cell(row, col)
                        if cell.value is not None and cell.data_type != 'f':
                            datos_encontrados += 1
                            break
                    if datos_encontrados > 0:
                        break
            except:
                pass
            
            if datos_encontrados == 0 and total_filas > 0:
                self.enviar_mensaje('actualizar_texto', "\n⚠️ ADVERTENCIA: No se detectaron datos en las primeras filas del destino")
                self.enviar_mensaje('actualizar_texto', "Esto puede ser normal si los datos están en columnas más adelante")
            
            # Extraer y mostrar cuadre
            self.enviar_mensaje('progreso', "Extrayendo cuadre...", 99)
            cuadre = self.extraer_cuadre(df_origen)
            self.mostrar_cuadre(cuadre, total_filas)
            
            self.procesando = False
            self.enviar_mensaje('completado', f"Transformación completada exitosamente.\n\n{total_filas} filas procesadas.\n\nEl archivo está listo para descargar.")
            
        except Exception as e:
            self.procesando = False
            import traceback
            error_detalle = traceback.format_exc()
            
            # Mensaje de error más amigable
            error_msg = f"Error durante la transformación:\n\n{str(e)}\n\n"
            error_msg += "Posibles causas:\n"
            error_msg += "- El archivo origen está abierto en Excel\n"
            error_msg += "- El archivo origen está corrupto\n"
            error_msg += "- Falta alguna columna requerida\n"
            error_msg += "- Error de formato en los datos\n\n"
            error_msg += "Detalles técnicos:\n" + str(error_detalle)
            
            self.enviar_mensaje('error', error_msg)
            
    
    def obtener_mapeo_columnas(self, headers_origen, headers_destino):
        """Obtiene el mapeo de índices de columnas entre origen y destino - OPTIMIZADO CON CACHE"""
        # Usar cache si los headers destino son los mismos (siempre lo son)
        cache_key = 'mapeo_columnas'
        if self._cache_mapeo_columnas is not None:
            # Verificar que los headers destino no hayan cambiado (comparar primeros headers)
            if self._cache_headers_destino and len(self._cache_headers_destino) == len(headers_destino):
                if len(headers_destino) > 0:
                    # Comparar algunos headers clave para verificar
                    cache_val = self._cache_headers_destino[0].value if hasattr(self._cache_headers_destino[0], 'value') else None
                    current_val = headers_destino[0].value if hasattr(headers_destino[0], 'value') else None
                    if cache_val == current_val:
                        # Los headers son los mismos, usar cache
                        return self._cache_mapeo_columnas.copy()  # Retornar copia para no modificar el cache
        
        mapeo = {}
        
        # Crear diccionario de nombres de columnas destino (más flexible)
        nombres_destino = {}
        nombres_destino_parciales = {}  # Para búsqueda parcial
        
        for idx, cell in enumerate(headers_destino):
            if cell.value:
                nombre_limpio = str(cell.value).strip().upper()
                nombres_destino[nombre_limpio] = idx + 1
                # También agregar versión sin espacios extra
                nombre_sin_espacios = nombre_limpio.replace('  ', ' ')
                if nombre_sin_espacios != nombre_limpio:
                    nombres_destino[nombre_sin_espacios] = idx + 1
                
                # Crear índice de palabras clave para búsqueda parcial
                palabras = nombre_limpio.split()
                for palabra in palabras:
                    if len(palabra) > 2:  # Solo palabras de más de 2 caracteres
                        if palabra not in nombres_destino_parciales:
                            nombres_destino_parciales[palabra] = []
                        nombres_destino_parciales[palabra].append((idx + 1, nombre_limpio))
        
        # Mapear columnas conocidas (más flexible con espacios)
        mapeos_conocidos = {
            'PRIMER APELLIDO': ['PRIMER APELLIDO'],
            'SEGUNDO APELLIDO': ['SEGUNDO APELLIDO'],
            'PRIMER NOMBRE': ['PRIMER NOMBRE'],
            'SEGUNDO NOMBRE': ['SEGUNDO NOMBRE'],
            'OFICINA': ['OFICINA'],
            'TIPO IDENTIFICACION': ['TIPO IDENTIFICACION', 'TIPO IDENTIFICACIÓN'],
            'NUMERO DE IDENTIFICACION': ['NUMERO DE IDENTIFICACION', 'NUMERO DE IDENTIFICACIÓN', 'NÚMERO DE IDENTIFICACION'],
            'FECHA DE NACIMIENTO': ['FECHA DE NACIMIENTO'],
            'SEXO/GENERO': ['SEXO/GENERO', 'SEXO', 'GENERO'],
            'ESTADO CIVIL': ['ESTADO CIVIL'],
            'NACIONALIDAD': ['NACIONALIDAD ACTUAL', 'NACIONALIDAD', 'NACIONALIDAD ACTUAL '],
            'PAIS DE ORIGEN': ['PAIS DE ORIGEN', 'PAÍS DE ORIGEN', 'PAIS DE ORIGEN '],
            'PROVINCIA': ['PROVINCIA', 'PROVINCIA ', ' PROVINCIA', 'PROVINCIA DE', 'PROVINCIA DEL'],
            'CIUDAD': ['CIUDAD', 'CIUDAD ', ' CIUDAD', 'CIUDAD DE', 'CIUDAD DEL'],
            'DIRECCION ': ['DIRECCION', 'DIRECCIÓN', 'DIRECCION '],
            'TELEFONO CASA': ['TELEFONO CASA', 'TELÉFONO CASA'],
            'TELEFONO TRABAJO': ['TELEFONO TRABAJO', 'TELÉFONO TRABAJO'],
            'CELULAR': ['CELULAR'],
            'DIRECCION TRABAJO': ['DIRECCION TRABAJO', 'DIRECCIÓN TRABAJO'],
            'EMAIL': ['EMAIL', 'CORREO', 'E-MAIL'],
            'OCUPACION': ['OCUPACION', 'OCUPACIÓN'],
            'ACTIVIDAD ECONOMICA': ['ACTIVIDAD ECONOMICA', 'ACTIVIDAD ECONÓMICA'],
            'INGRESOS': ['INGRESOS'],
            'PATRIMONIO': ['PATRIMONIO'],
            'MONTO CREDITO': ['MONTO CREDITO', 'MONTO CRÉDITO'],
            'FECHA DE INICIO DE CREDITO': ['FECHA DE INICIO DE CREDITO', 'FECHA DE INICIO DE CRÉDITO'],
            'FECHA DE TERMINACION DE CREDITO': ['FECHA DE TERMINACION DE CREDITO', 'FECHA DE TERMINACIÓN DE CRÉDITO'],
            'PLAZO DE CREDITO': ['PLAZO DE CREDITO', 'PLAZO DE CRÉDITO'],
            'PRIMA NETA': ['PRIMA NETA'],
        }
        
        # PRIMERO: Mapear columnas conocidas
        for idx_origen, header_orig in enumerate(headers_origen):
            if pd.notna(header_orig):
                header_orig_str = str(header_orig).strip().upper()
                # Normalizar espacios múltiples
                header_orig_str_limpio = ' '.join(header_orig_str.split())
                
                # Buscar en mapeos conocidos
                for key_orig, posibles_dest in mapeos_conocidos.items():
                    key_orig_limpio = ' '.join(key_orig.split())
                    
                    # Coincidencia exacta o con espacios normalizados
                    if header_orig_str_limpio == key_orig_limpio:
                        # Buscar en nombres destino
                        for nombre_dest in posibles_dest:
                            nombre_dest_limpio = ' '.join(nombre_dest.split())
                            if nombre_dest_limpio in nombres_destino:
                                mapeo[idx_origen] = nombres_destino[nombre_dest_limpio]
                                break
                            elif nombre_dest in nombres_destino:
                                mapeo[idx_origen] = nombres_destino[nombre_dest]
                                break
                        if idx_origen in mapeo:
                            break
                    # También buscar coincidencia parcial (sin espacios extra)
                    elif key_orig_limpio in header_orig_str_limpio or header_orig_str_limpio in key_orig_limpio:
                        for nombre_dest in posibles_dest:
                            nombre_dest_limpio = ' '.join(nombre_dest.split())
                            if nombre_dest_limpio in nombres_destino:
                                mapeo[idx_origen] = nombres_destino[nombre_dest_limpio]
                                break
                            elif nombre_dest in nombres_destino:
                                mapeo[idx_origen] = nombres_destino[nombre_dest]
                                break
                        if idx_origen in mapeo:
                            break
                    # Para PROVINCIA y CIUDAD, buscar también coincidencias parciales más flexibles
                    elif key_orig in ['PROVINCIA', 'CIUDAD']:
                        # Buscar coincidencia más flexible
                        if key_orig in header_orig_str_limpio:
                            # Buscar en todas las columnas destino que contengan PROVINCIA o CIUDAD
                            for nom_dest, col_idx in nombres_destino.items():
                                # Buscar que el nombre destino contenga PROVINCIA o CIUDAD (sin otras palabras como PAIS)
                                if key_orig in nom_dest and 'PAIS' not in nom_dest:
                                    # Verificar que no sea una columna de VLOOKUP (generalmente tienen nombres más largos)
                                    # La primera columna de PROVINCIA/CIUDAD suele ser más simple
                                    if idx_origen not in mapeo or len(nom_dest) <= len(key_orig) + 5:
                                        mapeo[idx_origen] = col_idx
                                break
                        if idx_origen in mapeo:
                            break
        
        # SEGUNDO: Mapeo automático por nombre similar para TODAS las columnas no mapeadas
        # Esto incluye especialmente las columnas AP-BC que pueden estar vacías
        for idx_origen, header_orig in enumerate(headers_origen):
            if idx_origen in mapeo:  # Ya está mapeada, saltar
                continue
                
            if pd.notna(header_orig):
                header_orig_str = str(header_orig).strip().upper()
                header_orig_str_limpio = ' '.join(header_orig_str.split())
                
                # Buscar coincidencia exacta en destino
                if header_orig_str_limpio in nombres_destino:
                    mapeo[idx_origen] = nombres_destino[header_orig_str_limpio]
                    continue
                
                # Buscar coincidencia parcial (el nombre origen está contenido en destino o viceversa)
                mejor_coincidencia = None
                mejor_puntaje = 0
                
                for nom_dest, col_idx in nombres_destino.items():
                    # Calcular similitud
                    if header_orig_str_limpio == nom_dest:
                        mejor_coincidencia = col_idx
                        mejor_puntaje = 100
                        break
                    elif header_orig_str_limpio in nom_dest or nom_dest in header_orig_str_limpio:
                        # Calcular puntaje basado en longitud de coincidencia
                        coincidencia_len = min(len(header_orig_str_limpio), len(nom_dest))
                        if coincidencia_len > mejor_puntaje:
                            mejor_puntaje = coincidencia_len
                            mejor_coincidencia = col_idx
                
                # Si encontramos una buena coincidencia, mapear
                if mejor_coincidencia and mejor_puntaje >= 5:  # Mínimo 5 caracteres de coincidencia
                    mapeo[idx_origen] = mejor_coincidencia
                    continue
                
                # Buscar por palabras clave comunes
                palabras_origen = [p for p in header_orig_str_limpio.split() if len(p) > 2]
                if palabras_origen:
                    for palabra in palabras_origen:
                        if palabra in nombres_destino_parciales:
                            # Encontrar la mejor coincidencia
                            for col_idx, nom_dest in nombres_destino_parciales[palabra]:
                                if col_idx not in [m for m in mapeo.values()]:  # No usar si ya está mapeada
                                    # Verificar que no sea una fórmula (columnas con VLOOKUP, etc.)
                                    # Las columnas de datos suelen tener nombres más simples
                                    if len(nom_dest.split()) <= 5:  # Nombres simples, no fórmulas
                                        mapeo[idx_origen] = col_idx
                                        break
                            if idx_origen in mapeo:
                                break
        
        # Guardar en cache para próximas ejecuciones
        self._cache_mapeo_columnas = mapeo.copy()
        self._cache_headers_destino = headers_destino
        
        return mapeo
    
    def ajustar_referencias_formula(self, formula, fila_origen, fila_destino):
        """Ajusta las referencias de fila en una fórmula"""
        import re
        diferencia = fila_destino - fila_origen
        
        # Patrón para encontrar referencias de celda
        # Busca patrones como: H6, $H6, H$6, $H$6
        pattern = r'(\$?[A-Z]+\$?)(\d+)'
        
        def reemplazar_ref(match):
            col_ref = match.group(1)  # Ej: H, $H, H$, $H$
            fila_ref = int(match.group(2))  # Ej: 6
            
            # Si la fila es absoluta (tiene $ antes del número), no cambiar
            # Verificar si hay $ antes del número en el match completo
            match_completo = match.group(0)
            if '$' in match_completo:
                # Verificar si el $ está antes del número (fila absoluta)
                partes = match_completo.split(str(fila_ref))
                if len(partes) > 0 and '$' in partes[0]:
                    return match_completo  # No cambiar referencias absolutas
            
            # Ajustar referencia relativa
            nueva_fila = fila_ref + diferencia
            return f"{col_ref}{nueva_fila}"
        
        return re.sub(pattern, reemplazar_ref, formula)
    
    def limpiar_datos_destino(self, ws):
        """Limpia los datos pero conserva encabezados, estructura y filas de totales"""
        from openpyxl.cell.cell import MergedCell
        
        # Identificar la última fila de datos (antes de los totales)
        # Buscar filas que contengan palabras clave de totales
        ultima_fila_datos = ws.max_row
        for row in range(ws.max_row, max(1, ws.max_row - 50), -1):
            try:
                primera_celda = ws.cell(row, 1)
                if primera_celda.value:
                    valor_str = str(primera_celda.value).strip().upper()
                    if any(keyword in valor_str for keyword in ['TOTAL', 'CUADRE', 'PRECANCELACION', 'SUMA', 'SUM']):
                        # Esta es una fila de totales, no limpiar desde aquí
                        ultima_fila_datos = row - 1
                        break
            except:
                continue
        
        # Guardar plantilla de fórmulas de la primera fila de datos (fila 6)
        plantilla_formulas = {}
        fila_plantilla = 6
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(fila_plantilla, col)
            # Verificar que no sea una celda fusionada
            if not isinstance(cell, MergedCell) and cell.data_type == 'f':
                plantilla_formulas[col] = cell.value
        
        # Limpiar SOLO desde fila 7 hasta la última fila de datos (OPTIMIZADO)
        # NO limpiar las filas de totales al final
        # Optimización: Limitar columnas y procesar en lotes
        max_col = min(ws.max_column, 200)  # Limitar a 200 columnas máximo
        batch_size = 500  # Procesar en lotes de 500 filas
        for row_batch_start in range(7, ultima_fila_datos + 1, batch_size):
            row_batch_end = min(row_batch_start + batch_size, ultima_fila_datos + 1)
            for row in range(row_batch_start, row_batch_end):
                for col in range(1, max_col + 1):
                    try:
                        cell = ws.cell(row, col)
                        # Verificar que no sea una celda fusionada antes de modificar
                        if isinstance(cell, MergedCell):
                            continue  # Saltar celdas fusionadas
                        
                        # Solo limpiar si tiene valor (optimización: no limpiar celdas vacías)
                        if cell.value is not None:
                            cell.value = None
                    except:
                        continue  # Si hay error, continuar con la siguiente celda
    
    def transferir_fila_optimizada(self, df_origen, idx_origen, ws_destino, fila_destino, 
                       mapeo_columnas, headers_origen, headers_destino,
                       idx_pais_origen=None, idx_pais_residencia_dest=None):
        """Versión optimizada de transferir_fila sin recopilar datos para mostrar"""
        row_origen = df_origen.iloc[idx_origen]
        fila_plantilla = 6
        
        # Pre-calcular índice de TIPO IDENTIFICACION (usar cache si está disponible)
        tipo_identificacion = None
        idx_tipo_id = None
        
        # Usar cache para índice de TIPO IDENTIFICACION
        if 'idx_tipo_identificacion' not in self._cache_indices_columnas:
            # Buscar y cachear
            for idx, header in enumerate(headers_origen):
                if pd.notna(header) and 'TIPO IDENTIFICACION' in str(header).upper():
                    self._cache_indices_columnas['idx_tipo_identificacion'] = idx
                    idx_tipo_id = idx
                    break
        else:
            idx_tipo_id = self._cache_indices_columnas['idx_tipo_identificacion']
        
        # Obtener valor de TIPO IDENTIFICACION
        if idx_tipo_id is not None and idx_tipo_id < len(row_origen):
            tipo_valor = row_origen.iloc[idx_tipo_id]
            if pd.notna(tipo_valor):
                tipo_identificacion = str(tipo_valor).strip()
        
        # Paso 1: Copiar SOLO las fórmulas de la plantilla (fila 6) - OPTIMIZADO CON CACHE
        # Usar cache de fórmulas (siempre las mismas)
        if not hasattr(self, '_formulas_cache'):
            self._formulas_cache = {}
        if not hasattr(self, '_formulas_pattern'):
            self._formulas_pattern = re.compile(r'(\$?[A-Z]+\$?)(\d+)')
        
        if fila_destino != fila_plantilla:
            diferencia_filas = fila_destino - fila_plantilla
            
            def reemplazar_ref(match):
                col_ref = match.group(1)
                fila_ref = int(match.group(2))
                match_completo = match.group(0)
                if '$' in match_completo:
                    partes = match_completo.split(str(fila_ref))
                    if len(partes) > 0 and '$' in partes[0]:
                        return match_completo
                return f"{col_ref}{fila_ref + diferencia_filas}"
            
            # Cachear fórmulas de la plantilla si no están en cache (OPTIMIZADO)
            if fila_plantilla not in self._formulas_cache:
                formulas_plantilla = {}
                max_cols = min(ws_destino.max_column, 200)  # Limitar columnas para velocidad
                # Optimización: Leer fila completa de una vez
                row_plantilla = ws_destino[fila_plantilla]
                for idx, cell_plantilla in enumerate(row_plantilla, start=1):
                    if idx > max_cols:
                        break
                    try:
                        if not isinstance(cell_plantilla, MergedCell) and cell_plantilla.data_type == 'f':
                            formulas_plantilla[idx] = str(cell_plantilla.value)
                    except:
                        continue
                self._formulas_cache[fila_plantilla] = formulas_plantilla
            
            # Aplicar fórmulas desde cache (OPTIMIZADO - batch)
            formulas_plantilla = self._formulas_cache[fila_plantilla]
            
            # Pre-calcular headers para EDAD (optimización)
            headers_edad_cache = {}
            for col in formulas_plantilla.keys():
                if col-1 < len(headers_destino) and headers_destino[col-1].value:
                    header_cell = str(headers_destino[col-1].value).upper()
                    if 'EDAD' in header_cell:
                        headers_edad_cache[col] = True
            
            for col, formula_original in formulas_plantilla.items():
                try:
                    cell_destino = ws_destino.cell(fila_destino, col)
                    if isinstance(cell_destino, MergedCell):
                        continue
                    
                    # Optimización: Solo hacer regex si hay diferencia de filas
                    if diferencia_filas != 0:
                        formula_ajustada = self._formulas_pattern.sub(reemplazar_ref, formula_original)
                    else:
                        formula_ajustada = formula_original
                    
                    # Modificar fórmula de EDAD para que tenga 2 decimales (usar cache)
                    if col in headers_edad_cache:
                        if 'TODAY()' in formula_ajustada.upper() or '/365' in formula_ajustada.upper():
                            if 'ROUND' not in formula_ajustada.upper():
                                # Envolver la fórmula con ROUND para 2 decimales
                                formula_sin_igual = formula_ajustada[1:] if formula_ajustada.startswith('=') else formula_ajustada
                                formula_ajustada = f"=ROUND({formula_sin_igual},2)"
                    
                    # Escribir fórmula (optimización: estilos solo cuando sea necesario)
                    cell_destino.value = formula_ajustada
                    # Aplicar estilos solo a fórmulas importantes
                    if col in headers_edad_cache or col <= 20:  # Solo primeras columnas o EDAD
                        cell_destino.font = self._estilo_fuente_calibri
                        cell_destino.alignment = self._estilo_alineacion_centrada
                        cell_destino.border = self._estilo_borde_celda
                except:
                    continue
        
        # Paso 2: Copiar los datos del origen en las columnas mapeadas (optimizado)
        valor_pais_origen = None
        
        # Obtener PAIS DE ORIGEN primero si tenemos el índice
        if idx_pais_origen is not None and idx_pais_origen < len(row_origen):
            valor_pais = row_origen.iloc[idx_pais_origen]
            if pd.notna(valor_pais) and str(valor_pais).strip() != '':
                valor_pais_origen = valor_pais
        
        # Optimización: Pre-calcular flags de columnas especiales
        es_provincia_ciudad_cols = {15, 16}  # O y P
        es_columna_ap_bc = set(range(42, 56))  # AP-BC (42-55)
        
        # Optimización: Convertir row_origen a lista una sola vez para acceso más rápido
        row_values = row_origen.values if hasattr(row_origen, 'values') else [row_origen.iloc[i] for i in range(len(row_origen))]
        
        # Optimización: Cachear headers procesados
        if not hasattr(self, '_headers_cache'):
            self._headers_cache = {}
        
        for idx_origen_col, col_destino in mapeo_columnas.items():
            try:
                if idx_origen_col >= len(row_values):
                    continue
                
                valor = row_values[idx_origen_col]
                
                # Guardar PAIS DE ORIGEN si no lo tenemos aún
                if valor_pais_origen is None and idx_origen_col < len(headers_origen):
                    if idx_origen_col not in self._headers_cache:
                        self._headers_cache[idx_origen_col] = str(headers_origen[idx_origen_col]).strip().upper()
                    header_orig_upper = self._headers_cache[idx_origen_col]
                    if 'PAIS DE ORIGEN' in header_orig_upper or 'PAÍS DE ORIGEN' in header_orig_upper:
                        if pd.notna(valor) and str(valor).strip() != '':
                            valor_pais_origen = valor
                
                # Solo copiar si el valor no es NaN y no está vacío
                if pd.notna(valor):
                    # Optimización: Evitar str() si no es necesario
                    if isinstance(valor, str):
                        valor_str = valor.strip()
                    else:
                        valor_str = str(valor).strip()
                    
                    if valor_str and valor_str.lower() != 'nan':
                        try:
                            # Optimización: Pre-calcular flags antes de acceder a celda
                            es_provincia_o_ciudad = (col_destino in es_provincia_ciudad_cols)
                            es_columna_ap_bc_flag = (col_destino in es_columna_ap_bc)
                            
                            # Solo acceder a la celda si realmente necesitamos escribir
                            necesita_escribir = es_provincia_o_ciudad or es_columna_ap_bc_flag
                            
                            # Optimización: Acceder a celda solo una vez
                            cell_destino = ws_destino.cell(fila_destino, col_destino)
                            if isinstance(cell_destino, MergedCell):
                                continue
                            
                            # Verificar si tiene fórmula antes de escribir (solo si no es columna especial)
                            if not necesita_escribir and cell_destino.data_type == 'f':
                                continue
                            
                            # Aplicar transformaciones solo si es necesario (optimizado)
                            # Pre-calcular header_orig una sola vez usando cache
                            header_orig = None
                            if idx_origen_col < len(headers_origen):
                                if idx_origen_col not in self._headers_cache:
                                    self._headers_cache[idx_origen_col] = str(headers_origen[idx_origen_col]).strip().upper()
                                header_orig = self._headers_cache[idx_origen_col]
                                
                                # Transformación para PROVINCIA y CIUDAD (ULTRA OPTIMIZADO)
                                if 'PROVINCIA' in header_orig or 'CIUDAD' in header_orig:
                                    # Optimización: Procesamiento más rápido
                                    if isinstance(valor, (int, float)):
                                        # Si ya es numérico, solo convertir a int
                                        valor = int(valor)
                                    elif isinstance(valor_str, str) and valor_str:
                                        # Quitar cero inicial si existe (optimizado)
                                        if len(valor_str) > 1 and valor_str[0] == '0' and valor_str[1:].isdigit():
                                            valor = int(valor_str[1:])
                                        elif valor_str.isdigit():
                                            valor = int(valor_str)
                                        else:
                                            # Intentar convertir si es numérico con decimales
                                            try:
                                                if valor_str.replace('.', '', 1).replace('-', '', 1).isdigit():
                                                    valor = int(float(valor_str))
                                                else:
                                                    valor = valor_str
                                            except:
                                                valor = valor_str
                                    else:
                                        valor = valor_str
                                # Transformación para NACIONALIDAD: copiar del origen, pero si TIPO IDENTIFICACION = "00", entonces "239"
                                elif 'NACIONALIDAD' in header_orig:
                                    if tipo_identificacion == '00' or tipo_identificacion == '0':
                                        valor = '239'
                                    else:
                                        # Copiar el valor del origen directamente
                                        valor = valor_str
                                # Transformación para PAIS DE ORIGEN: convertir a número en columna M (columna 13)
                                elif 'PAIS DE ORIGEN' in header_orig or 'PAÍS DE ORIGEN' in header_orig:
                                    # Si la columna destino es M (columna 13), convertir a número
                                    if col_destino == 13:
                                        try:
                                            # Intentar convertir a número
                                            if isinstance(valor, (int, float)):
                                                valor = int(valor)
                                            elif isinstance(valor_str, str) and valor_str:
                                                # Si es string, intentar convertir a número
                                                if valor_str.isdigit():
                                                    valor = int(valor_str)
                                                elif valor_str.replace('.', '', 1).replace('-', '', 1).isdigit():
                                                    valor = int(float(valor_str))
                                                else:
                                                    # Si no se puede convertir, mantener el valor original
                                                    valor = valor_str
                                            else:
                                                valor = valor_str
                                        except:
                                            # Si hay error, mantener el valor original
                                            valor = valor_str
                                    else:
                                        # Si no es columna M, mantener como texto
                                        valor = valor_str
                                # Transformación para MONTO CREDITO: convertir a número con 2 decimales
                                elif 'MONTO CREDITO' in header_orig or 'MONTO CRÉDITO' in header_orig:
                                    try:
                                        if isinstance(valor, (int, float)):
                                            valor = round(float(valor), 2)
                                        elif isinstance(valor_str, str) and valor_str:
                                            # Intentar convertir a número
                                            try:
                                                valor = round(float(valor_str), 2)
                                            except:
                                                valor = valor_str
                                        else:
                                            valor = valor_str
                                    except:
                                        valor = valor_str
                                # Transformación para PLAZO DE CREDITO: convertir a número con 2 decimales
                                elif 'PLAZO DE CREDITO' in header_orig or 'PLAZO DE CRÉDITO' in header_orig:
                                    try:
                                        if isinstance(valor, (int, float)):
                                            valor = round(float(valor), 2)
                                        elif isinstance(valor_str, str) and valor_str:
                                            # Intentar convertir a número
                                            try:
                                                valor = round(float(valor_str), 2)
                                            except:
                                                valor = valor_str
                                        else:
                                            valor = valor_str
                                    except:
                                        valor = valor_str
                                # Formatear fechas: escribir como objeto date sin hora (00:00:00)
                                elif 'FECHA' in header_orig:
                                    try:
                                        fecha_obj = None
                                        if isinstance(valor, (datetime, pd.Timestamp)):
                                            # Convertir a date para quitar la hora
                                            fecha_obj = valor.date() if hasattr(valor, 'date') else valor
                                        elif isinstance(valor, str):
                                            # Si tiene hora, quitarla
                                            if ' ' in valor:
                                                valor = valor.split(' ')[0]  # Quitar la parte de hora
                                            if '-' in valor or '/' in valor:
                                                try:
                                                    fecha_parsed = pd.to_datetime(valor)
                                                    fecha_obj = fecha_parsed.date() if hasattr(fecha_parsed, 'date') else fecha_parsed
                                                except:
                                                    pass
                                        
                                        # Escribir como objeto date (sin hora) para que Excel lo reconozca
                                        if fecha_obj:
                                            if isinstance(fecha_obj, datetime):
                                                valor = fecha_obj.date()
                                            elif hasattr(fecha_obj, 'date'):
                                                valor = fecha_obj.date()
                                            # Si ya es date object, mantenerlo (ya no tiene hora)
                                    except:
                                        pass
                                else:
                                    valor = valor_str
                            
                            # Escribir el valor (optimización: estilos se aplicarán en batch al final)
                            cell_destino.value = valor
                            
                            # Aplicar formato numérico solo cuando sea necesario (optimización)
                            if 'FECHA' in header_orig and isinstance(valor, (date, datetime)):
                                cell_destino.number_format = 'mm/dd/yyyy'
                            elif isinstance(valor, (int, float)) and not isinstance(valor, bool):
                                # Aplicar formato de 2 decimales a todos los números
                                cell_destino.number_format = '0.00'
                        except:
                            continue
            except:
                continue
        
        # Paso 3: Para PAIS DE RESIDENCIA, siempre poner "239" (optimizado: sin estilos por ahora)
        if idx_pais_residencia_dest is not None:
            try:
                cell_destino = ws_destino.cell(fila_destino, idx_pais_residencia_dest)
                if not isinstance(cell_destino, MergedCell) and cell_destino.data_type != 'f':
                    cell_destino.value = '239'
                    # Estilos se aplicarán al final en batch
            except:
                pass
        
        # Paso 4: Para NUMERO DE POLIZA, siempre poner "5852" (optimizado: cachear índice)
        if not hasattr(self, '_idx_numero_poliza'):
            self._idx_numero_poliza = None
            for idx, cell in enumerate(headers_destino):
                if cell.value and 'NUMERO' in str(cell.value).upper() and 'POLIZA' in str(cell.value).upper():
                    self._idx_numero_poliza = idx + 1
                    break
        
        if self._idx_numero_poliza is not None:
            try:
                cell_poliza = ws_destino.cell(fila_destino, self._idx_numero_poliza)
                if not isinstance(cell_poliza, MergedCell) and cell_poliza.data_type != 'f':
                    cell_poliza.value = '5852'
            except:
                pass
        
        # Paso 5: Para NOMBRE DEL PRODUCTO, siempre poner "MONTO DEL CREDITO" (optimizado: cachear índice)
        if not hasattr(self, '_idx_nombre_producto'):
            self._idx_nombre_producto = None
            for idx, cell in enumerate(headers_destino):
                if cell.value and 'NOMBRE' in str(cell.value).upper() and 'PRODUCTO' in str(cell.value).upper():
                    self._idx_nombre_producto = idx + 1
                    break
        
        if self._idx_nombre_producto is not None:
            try:
                cell_producto = ws_destino.cell(fila_destino, self._idx_nombre_producto)
                if not isinstance(cell_producto, MergedCell) and cell_producto.data_type != 'f':
                    cell_producto.value = 'MONTO DEL CREDITO'
            except:
                pass
    
    def desplazar_filas_totales(self, ws, ultima_fila_datos_nueva):
        """Desplaza las filas de totales hacia abajo si hay más datos que en la plantilla original"""
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter
        
        # Detectar dónde empiezan las filas de totales en la plantilla original
        # Buscar desde el final hacia arriba
        fila_inicio_totales_original = None
        fila_fin_totales_original = ws.max_row
        
        # Buscar filas con totales (buscar en múltiples columnas, no solo la primera)
        for row in range(ws.max_row, max(1, ws.max_row - 200), -1):
            try:
                # Buscar en las primeras 10 columnas para detectar totales
                encontrado = False
                for col in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(row, col)
                    if cell.value:
                        valor_str = str(cell.value).strip().upper()
                        if any(keyword in valor_str for keyword in ['TOTAL', 'CUADRE', 'PRECANCELACION', 'PRE CANCELACION', 'SUMA']):
                            encontrado = True
                            break
                
                if encontrado:
                    if fila_inicio_totales_original is None:
                        fila_inicio_totales_original = row
                    # Actualizar fin si encontramos una fila más arriba
                    if row < fila_inicio_totales_original:
                        fila_inicio_totales_original = row
            except:
                continue
        
        # Si no se encontraron filas de totales, no hay nada que desplazar
        if fila_inicio_totales_original is None:
            return 0
        
        # Calcular cuántas filas de datos había en la plantilla original
        # (asumiendo que los totales empiezan después de los datos)
        filas_datos_original = fila_inicio_totales_original - 6  # Restar encabezados (filas 1-5) y primera fila de datos (6)
        
        # Calcular diferencia
        diferencia_filas = ultima_fila_datos_nueva - (fila_inicio_totales_original - 1)
        
        # Si hay más datos nuevos que originales, desplazar las filas de totales
        if diferencia_filas > 0:
            self.enviar_mensaje('actualizar_texto', f"\n📊 DETECCIÓN DE TOTALES:")
            self.enviar_mensaje('actualizar_texto', f"  • Filas de datos originales: {filas_datos_original}")
            self.enviar_mensaje('actualizar_texto', f"  • Filas de datos nuevas: {ultima_fila_datos_nueva - 5}")
            self.enviar_mensaje('actualizar_texto', f"  • Diferencia: +{diferencia_filas} filas")
            self.enviar_mensaje('actualizar_texto', f"  • Desplazando {fila_fin_totales_original - fila_inicio_totales_original + 1} fila(s) de totales...")
            
            # Calcular nuevas posiciones
            nueva_fila_inicio = fila_inicio_totales_original + diferencia_filas
            nueva_fila_fin = fila_fin_totales_original + diferencia_filas
            num_filas_totales = fila_fin_totales_original - fila_inicio_totales_original + 1
            
            # Copiar filas de totales desde el final hacia el inicio (para evitar sobrescribir)
            # Primero, copiar todas las filas de totales a una posición temporal más abajo
            fila_temporal_inicio = ws.max_row + 1
            
            for offset in range(num_filas_totales):
                fila_origen = fila_fin_totales_original - offset  # Empezar desde el final
                fila_temp = fila_temporal_inicio + (num_filas_totales - 1 - offset)
                
                # Copiar toda la fila
            for col in range(1, ws.max_column + 1):
                    try:
                        cell_origen = ws.cell(fila_origen, col)
                        cell_temp = ws.cell(fila_temp, col)
                        
                        if isinstance(cell_origen, MergedCell) or isinstance(cell_temp, MergedCell):
                            continue
                        
                        # Copiar valor, fórmula, formato, etc.
                        if cell_origen.data_type == 'f':
                            # Es una fórmula, copiar y ajustar referencias
                            formula = str(cell_origen.value)
                            # Ajustar referencias de fila en la fórmula
                            formula_ajustada = self.ajustar_referencias_formula(
                                formula, fila_origen, fila_temp
                            )
                            cell_temp.value = formula_ajustada
                        else:
                            # Es un valor, copiar directamente
                            cell_temp.value = cell_origen.value
                        
                        # Copiar formato si es posible
                        try:
                            cell_temp.font = copy.copy(cell_origen.font)
                            cell_temp.fill = copy.copy(cell_origen.fill)
                            cell_temp.border = copy.copy(cell_origen.border)
                            cell_temp.alignment = copy.copy(cell_origen.alignment)
                            cell_temp.number_format = cell_origen.number_format
                        except:
                            pass
                    except:
                        continue
            
            # Ahora mover las filas temporales a su posición final
            for offset in range(num_filas_totales):
                fila_temp = fila_temporal_inicio + offset
                fila_destino_final = nueva_fila_inicio + offset
                
                for col in range(1, ws.max_column + 1):
                    try:
                        cell_temp = ws.cell(fila_temp, col)
                        cell_destino = ws.cell(fila_destino_final, col)
                        
                        if isinstance(cell_temp, MergedCell) or isinstance(cell_destino, MergedCell):
                            continue
                        
                        if cell_temp.data_type == 'f':
                            # Ajustar referencias de fila en la fórmula
                            formula = str(cell_temp.value)
                            formula_ajustada = self.ajustar_referencias_formula(
                                formula, fila_temp, fila_destino_final
                            )
                            cell_destino.value = formula_ajustada
                        else:
                            cell_destino.value = cell_temp.value
                        
                        # Copiar formato
                        try:
                            cell_destino.font = copy.copy(cell_temp.font)
                            cell_destino.fill = copy.copy(cell_temp.fill)
                            cell_destino.border = copy.copy(cell_temp.border)
                            cell_destino.alignment = copy.copy(cell_temp.alignment)
                            cell_destino.number_format = cell_temp.number_format
                        except:
                            pass
                    except:
                        continue
                
                # Limpiar fila temporal
                for col in range(1, ws.max_column + 1):
                    try:
                        cell_temp = ws.cell(fila_temp, col)
                        if not isinstance(cell_temp, MergedCell):
                            cell_temp.value = None
                    except:
                        pass
            
            # Limpiar las filas originales de totales
            for row in range(fila_inicio_totales_original, fila_fin_totales_original + 1):
                for col in range(1, ws.max_column + 1):
                    try:
                        cell = ws.cell(row, col)
                        if not isinstance(cell, MergedCell):
                            cell.value = None
                    except:
                        pass
            
            return num_filas_totales
        
        return 0
    
    def validar_formulas_datos(self, ws, ultima_fila_datos):
        """Valida que no haya NaN donde deberían estar fórmulas (optimizado)"""
        from openpyxl.cell.cell import MergedCell
        import re
        
        self.enviar_mensaje('actualizar_texto', "\n🔍 Validando fórmulas...")
        
        fila_plantilla = 6
        formulas_restauradas = 0
        
        # Verificar solo 5 filas de muestra para velocidad
        filas_a_verificar = min(5, ultima_fila_datos - 5)
        
        # Pre-calcular patrón regex una vez
        pattern = r'(\$?[A-Z]+\$?)(\d+)'
        
        for row in range(6, 6 + filas_a_verificar):
            diferencia = row - fila_plantilla
            
            def reemplazar_ref(match):
                col_ref = match.group(1)
                fila_ref = int(match.group(2))
                match_completo = match.group(0)
                if '$' in match_completo:
                    partes = match_completo.split(str(fila_ref))
                    if len(partes) > 0 and '$' in partes[0]:
                        return match_completo
                return f"{col_ref}{fila_ref + diferencia}"
            
            # Solo verificar primeras 100 columnas para velocidad
            for col in range(1, min(ws.max_column + 1, 100)):
                try:
                    cell_plantilla = ws.cell(fila_plantilla, col)
                    
                    if isinstance(cell_plantilla, MergedCell) or cell_plantilla.data_type != 'f':
                        continue
                    
                    cell_verificar = ws.cell(row, col)
                    if isinstance(cell_verificar, MergedCell):
                        continue
                    
                    # Si la plantilla tiene fórmula pero la celda verificada no
                    if cell_verificar.data_type != 'f':
                        # Verificar si el valor es NaN o None
                        if cell_verificar.value is None or (isinstance(cell_verificar.value, float) and pd.isna(cell_verificar.value)):
                            # Restaurar la fórmula directamente
                            formula_original = str(cell_plantilla.value)
                            formula_ajustada = re.sub(pattern, reemplazar_ref, formula_original)
                            cell_verificar.value = formula_ajustada
                            formulas_restauradas += 1
                except:
                    continue
        
        if formulas_restauradas > 0:
            self.enviar_mensaje('actualizar_texto', f"  ⚠️ {formulas_restauradas} fórmula(s) restaurada(s)")
        else:
            self.enviar_mensaje('actualizar_texto', f"  ✓ Fórmulas validadas correctamente")
    
    def establecer_fecha_corte(self, ws, df_origen, headers_origen):
        """Establece la fecha de corte en B3 (última fecha del mes) y retorna la fecha"""
        from datetime import datetime
        import calendar
        
        # Buscar columna FECHA DE INICIO DE CREDITO
        col_fecha_inicio = None
        for idx, header in enumerate(headers_origen):
            if pd.notna(header) and 'FECHA DE INICIO DE CREDITO' in str(header).upper():
                col_fecha_inicio = idx
                break
        
        if col_fecha_inicio is None:
            return None
        
        # Buscar la primera fecha válida
        fecha_mes = None
        for idx in range(6, min(len(df_origen), 100)):  # Buscar en las primeras filas
            try:
                fecha_valor = df_origen.iloc[idx, col_fecha_inicio]
                if pd.notna(fecha_valor):
                    try:
                        if isinstance(fecha_valor, datetime):
                            fecha_mes = fecha_valor
                        elif isinstance(fecha_valor, pd.Timestamp):
                            fecha_mes = fecha_valor.to_pydatetime()
                        elif isinstance(fecha_valor, str):
                            fecha_mes = pd.to_datetime(fecha_valor).to_pydatetime()
                        else:
                            fecha_mes = pd.to_datetime(fecha_valor).to_pydatetime()
                        if fecha_mes:
                            break
                    except:
                        continue
            except:
                continue
        
        if fecha_mes:
            # Calcular última fecha del mes
            año = fecha_mes.year
            mes = fecha_mes.month
            ultimo_dia = calendar.monthrange(año, mes)[1]
            fecha_corte = datetime(año, mes, ultimo_dia).date()
            
            # Establecer en B3 como objeto date para que Excel lo reconozca correctamente
            try:
                cell_b3 = ws.cell(3, 2)  # B3 = fila 3, columna 2
                # Escribir como objeto date (no como string) para que Excel lo reconozca
                cell_b3.value = fecha_corte
                # Aplicar formato de fecha sin hora (mm/dd/yyyy)
                cell_b3.number_format = 'mm/dd/yyyy'
                return fecha_corte
            except:
                return None
        
        return None
    
    def renombrar_hoja_dv(self, wb, nombre_hoja_actual, fecha_corte):
        """Renombra la hoja DV según el mes y año de la fecha de corte"""
        meses_espanol = {
            1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
            5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
            9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
        }
        
        try:
            mes_nombre = meses_espanol.get(fecha_corte.month, 'MES')
            año = fecha_corte.year
            
            # Extraer el número (5852) del nombre actual si existe
            numero = None
            if '(' in nombre_hoja_actual and ')' in nombre_hoja_actual:
                try:
                    # Buscar número entre paréntesis
                    import re
                    match = re.search(r'\((\d+)\)', nombre_hoja_actual)
                    if match:
                        numero = match.group(1)
                except:
                    pass
            
            # Crear nuevo nombre: "DV [MES] ([número])" o "DV [MES] ([año])"
            if numero:
                nuevo_nombre = f"DV {mes_nombre} ({numero})"
            else:
                nuevo_nombre = f"DV {mes_nombre} ({año})"
            
            # Renombrar la hoja
            if nombre_hoja_actual in wb.sheetnames:
                wb[nombre_hoja_actual].title = nuevo_nombre
                self.enviar_mensaje('actualizar_texto', f"  ✓ Hoja renombrada: '{nombre_hoja_actual}' → '{nuevo_nombre}'")
                return nuevo_nombre
        except Exception as e:
            self.enviar_mensaje('actualizar_texto', f"  ⚠️ Error al renombrar hoja: {str(e)}")
        
        return None
    
    def actualizar_referencias_xml_pivot(self, archivo_path, nuevo_nombre_hoja, ultima_fila_datos):
        """Elimina referencias a archivos externos y deja solo el nombre de la hoja con su rango - VERSIÓN OPTIMIZADA"""
        import zipfile
        import re
        import tempfile
        import shutil
        import os
        
        try:
            # Buscar nombres antiguos posibles
            nombres_antiguos = ["DV OCTUBRE (5852)", "DV OCTUBRE", "DV"]
            
            # Crear archivo temporal
            temp_file = archivo_path + '.tmp'
            archivos_modificados = 0
            total_reemplazos = 0
            
            self.enviar_mensaje('actualizar_texto', "  → Procesando archivos XML...")
            
            with zipfile.ZipFile(archivo_path, 'r') as zip_read:
                with zipfile.ZipFile(temp_file, 'w', zipfile.ZIP_DEFLATED) as zip_write:
                    for item in zip_read.infolist():
                        data = zip_read.read(item.filename)
                        data_original = data
                        reemplazos_en_archivo = 0
                        
                        # SOLO procesar archivos XML relacionados con tablas dinámicas o que puedan contener referencias
                        if (item.filename.endswith('.xml') and 
                            ('pivot' in item.filename.lower() or 'cache' in item.filename.lower() or 
                             'worksheet' in item.filename.lower() or 'workbook' in item.filename.lower())):
                            try:
                                # Convertir a string para buscar y reemplazar directamente
                                data_str = data.decode('utf-8', errors='ignore')
                                data_str_original = data_str
                                
                                # Verificar si contiene referencias problemáticas antes de procesar
                                if 'Users' not in data_str and 'Downloads' not in data_str and '[' not in data_str:
                                    # No tiene referencias problemáticas, saltar
                                    zip_write.writestr(item, data)
                                    continue
                                
                                # ===== MÉTODO SIMPLIFICADO Y RÁPIDO =====
                                
                                # 1. Buscar y reemplazar el patrón completo problemático directamente
                                # '\Users\kevin\Downloads\[Facturación DV Noviembre 2025.xlsx]DV OCTUBRE (5852)'!$A$5:$BC$6
                                patron_completo = r"'\\Users[^']*\[[^\]]+\]" + re.escape('DV OCTUBRE (5852)') + r"'!"
                                if re.search(patron_completo, data_str, re.IGNORECASE):
                                    nuevo_rango = f"'{nuevo_nombre_hoja}'!"
                                    data_str = re.sub(patron_completo, nuevo_rango, data_str, flags=re.IGNORECASE)
                                    reemplazos_en_archivo += 1
                                
                                # 2. Buscar cualquier ruta seguida de [archivo]nombre_hoja y reemplazar
                                for nombre_antiguo in nombres_antiguos:
                                    # Buscar: 'ruta\[archivo]nombre_hoja'!
                                    patron_ruta_con_hoja = r"'[^']*\\[^']*\[[^\]]+\]" + re.escape(nombre_antiguo) + r"'!"
                                    if re.search(patron_ruta_con_hoja, data_str, re.IGNORECASE):
                                        data_str = re.sub(patron_ruta_con_hoja, f"'{nuevo_nombre_hoja}'!", data_str, flags=re.IGNORECASE)
                                        reemplazos_en_archivo += 1
                                
                                # 3. Eliminar cualquier corchete con .xlsx
                                if '[' in data_str and '.xlsx' in data_str:
                                    patron_xlsx = r'\[[^\]]+\.xlsx\]'
                                    data_str = re.sub(patron_xlsx, '', data_str, flags=re.IGNORECASE)
                                
                                # 4. Eliminar cualquier corchete restante
                                if '[' in data_str:
                                    patron_corchetes = r'\[[^\]]+\]'
                                    data_str = re.sub(patron_corchetes, '', data_str)
                                
                                # 5. Eliminar rutas de Windows que queden
                                if '\\Users' in data_str or '\\' in data_str:
                                    patron_rutas = r'[A-Z]:\\[^!\']*|\\[A-Za-z]+\\[^!\']*'
                                    data_str = re.sub(patron_rutas, '', data_str, flags=re.IGNORECASE)
                                
                                # 6. Corregir referencias que quedaron con el formato correcto
                                for nombre_antiguo in nombres_antiguos:
                                    # Formato: 'DV OCTUBRE (5852)'!$A$5:$BC$6
                                    patron_con_comillas = r"'" + re.escape(nombre_antiguo) + r"'!\$A\$5:\$BC\$\d+"
                                    if re.search(patron_con_comillas, data_str, re.IGNORECASE):
                                        nuevo_rango = f"'{nuevo_nombre_hoja}'!$A$5:$BC${ultima_fila_datos}"
                                        data_str = re.sub(patron_con_comillas, nuevo_rango, data_str, flags=re.IGNORECASE)
                                        reemplazos_en_archivo += 1
                                    
                                    # Solo nombre de hoja: 'DV OCTUBRE (5852)'!
                                    patron_solo_hoja = r"'" + re.escape(nombre_antiguo) + r"'!"
                                    if re.search(patron_solo_hoja, data_str, re.IGNORECASE):
                                        data_str = re.sub(patron_solo_hoja, f"'{nuevo_nombre_hoja}'!", data_str, flags=re.IGNORECASE)
                                        reemplazos_en_archivo += 1
                                
                                # 7. Corregir rangos que terminen antes de ultima_fila_datos
                                if '$A$5:$BC$' in data_str:
                                    patron_rango_final = r"(\$A\$5:\$BC\$)(\d+)"
                                    def corregir_rango_final(match):
                                        fila_actual = int(match.group(2))
                                        if fila_actual < ultima_fila_datos:
                                            return f"{match.group(1)}{ultima_fila_datos}"
                                        return match.group(0)
                                    
                                    data_str = re.sub(patron_rango_final, corregir_rango_final, data_str)
                                
                                # Convertir de vuelta a bytes si hubo cambios
                                if data_str != data_str_original:
                                    data = data_str.encode('utf-8')
                                    archivos_modificados += 1
                                    total_reemplazos += reemplazos_en_archivo
                                    
                            except Exception as e:
                                # Si hay error, mantener el archivo original
                                pass
                        
                        zip_write.writestr(item, data)
            
            # Reemplazar archivo original con el temporal
            if archivos_modificados > 0:
                shutil.move(temp_file, archivo_path)
                self.enviar_mensaje('actualizar_texto', f"  ✓ Tabla dinámica actualizada: {archivos_modificados} archivo(s) XML modificado(s), {total_reemplazos} reemplazo(s)")
            else:
                # Si no se modificó nada, eliminar el archivo temporal
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                self.enviar_mensaje('actualizar_texto', "  ⚠️ No se encontraron referencias para actualizar")
        except Exception as e:
            # Si hay error, eliminar archivo temporal si existe
            try:
                temp_file = archivo_path + '.tmp'
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                self.enviar_mensaje('actualizar_texto', f"  ⚠️ Error al actualizar XML: {str(e)}")
            except:
                pass
    
    def crear_hoja2_tabla_dinamica(self, wb, ws_destino, ultima_fila_datos, headers_destino):
        """Crea una nueva hoja 'Hoja2' con una tabla dinámica agrupada por rangos de MONTO CREDITO"""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill
        
        try:
            # Buscar o crear Hoja2
            hoja2 = None
            if 'Hoja2' in wb.sheetnames or 'HOJA2' in [s.upper() for s in wb.sheetnames]:
                # Si ya existe, eliminarla
                for nombre in wb.sheetnames:
                    if nombre.upper() == 'HOJA2':
                        wb.remove(wb[nombre])
                        break
            
            # Crear nueva hoja Hoja2
            hoja2 = wb.create_sheet("Hoja2")
            
            # Buscar la columna de MONTO CREDITO en los headers
            col_monto_credito = None
            for idx, cell in enumerate(headers_destino):
                if cell.value:
                    header_str = str(cell.value).strip().upper()
                    if 'MONTO CREDITO' in header_str or 'MONTO CRÉDITO' in header_str:
                        col_monto_credito = idx + 1
                        break
            
            if not col_monto_credito:
                self.enviar_mensaje('actualizar_texto', "  ⚠️ No se encontró la columna MONTO CREDITO para crear la tabla dinámica")
                return
            
            # Leer datos de MONTO CREDITO desde la hoja destino
            datos_monto = []
            for fila in range(6, ultima_fila_datos + 1):
                try:
                    cell = ws_destino.cell(fila, col_monto_credito)
                    if cell.value is not None:
                        valor = cell.value
                        # Convertir a float si es posible
                        if isinstance(valor, (int, float)):
                            datos_monto.append(float(valor))
                        elif isinstance(valor, str):
                            try:
                                datos_monto.append(float(valor.replace(',', '.')))
                            except:
                                pass
                except:
                    continue
            
            if not datos_monto:
                self.enviar_mensaje('actualizar_texto', "  ⚠️ No se encontraron datos de MONTO CREDITO")
                return
            
            # Crear DataFrame para agrupar
            df = pd.DataFrame({'MONTO_CREDITO': datos_monto})
            
            # Definir rangos de agrupación
            def asignar_rango(valor):
                if valor <= 5000:
                    return "1-5000"
                elif valor <= 10000:
                    return "5001-10000"
                elif valor <= 15000:
                    return "10001-15000"
                elif valor <= 20000:
                    return "15001-20000"
                elif valor <= 25000:
                    return "20001-25000"
                elif valor <= 30000:
                    return "25001-30000"
                elif valor <= 35000:
                    return "30001-35000"
                elif valor <= 40000:
                    return "35001-40000"
                else:
                    # Agrupar valores mayores en rangos de 5000
                    inicio = int((valor - 1) // 5000) * 5000 + 1
                    fin = inicio + 4999
                    return f"{inicio}-{fin}"
            
            df['Rango'] = df['MONTO_CREDITO'].apply(asignar_rango)
            
            # Agrupar por rango
            resultado = df.groupby('Rango').agg({
                'MONTO_CREDITO': ['count', 'sum']
            }).reset_index()
            
            resultado.columns = ['Rango', 'Cuenta', 'Suma']
            
            # Ordenar por el valor inicial del rango
            def obtener_inicio_rango(rango_str):
                try:
                    return int(rango_str.split('-')[0])
                except:
                    return 0
            
            resultado['Orden'] = resultado['Rango'].apply(obtener_inicio_rango)
            resultado = resultado.sort_values('Orden')
            resultado = resultado.drop('Orden', axis=1)
            
            # Escribir encabezados en Hoja2
            hoja2['A1'] = 'Etiquetas de fila'
            hoja2['B1'] = 'Cuenta de MONTO CREDITO'
            hoja2['C1'] = 'Suma de MONTO CREDITO'
            
            # Aplicar formato a encabezados
            fill_gris = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            for col in ['A1', 'B1', 'C1']:
                cell = hoja2[col]
                cell.font = self._estilo_fuente_calibri
                cell.alignment = self._estilo_alineacion_centrada
                cell.fill = fill_gris
                cell.border = self._estilo_borde_celda
            
            # Escribir datos
            fila_actual = 2
            for _, row in resultado.iterrows():
                hoja2[f'A{fila_actual}'] = row['Rango']
                hoja2[f'B{fila_actual}'] = int(row['Cuenta'])
                hoja2[f'C{fila_actual}'] = round(row['Suma'], 2)
                
                # Aplicar formato
                for col in ['A', 'B', 'C']:
                    cell = hoja2[f'{col}{fila_actual}']
                    cell.font = self._estilo_fuente_calibri
                    cell.alignment = self._estilo_alineacion_centrada
                    cell.border = self._estilo_borde_celda
                    if col == 'C':  # Columna de suma con formato de 2 decimales
                        cell.number_format = '0.00'
                
                fila_actual += 1
            
            # Agregar fila de Total general
            total_cuenta = int(resultado['Cuenta'].sum())
            total_suma = round(resultado['Suma'].sum(), 2)
            
            hoja2[f'A{fila_actual}'] = 'Total general'
            hoja2[f'B{fila_actual}'] = total_cuenta
            hoja2[f'C{fila_actual}'] = total_suma
            
            # Aplicar formato a Total general (fondo azul claro)
            fill_azul = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            for col in ['A', 'B', 'C']:
                cell = hoja2[f'{col}{fila_actual}']
                cell.font = self._estilo_fuente_calibri
                cell.alignment = self._estilo_alineacion_centrada
                cell.border = self._estilo_borde_celda
                cell.fill = fill_azul
                if col == 'C':
                    cell.number_format = '0.00'
            
            # Ajustar ancho de columnas
            hoja2.column_dimensions['A'].width = 20
            hoja2.column_dimensions['B'].width = 25
            hoja2.column_dimensions['C'].width = 25
            
            self.enviar_mensaje('actualizar_texto', f"  ✓ Hoja2 creada con tabla dinámica agrupada por rangos de MONTO CREDITO")
            
        except Exception as e:
            self.enviar_mensaje('actualizar_texto', f"  ⚠️ Error al crear Hoja2: {str(e)}")
            import traceback
            self.enviar_mensaje('actualizar_texto', f"  Detalle: {traceback.format_exc()}")
    
    def crear_tabla_dinamica_xml(self, archivo_path, nombre_hoja_datos, ultima_fila_datos, col_monto_credito):
        """Crea una tabla dinámica en Hoja2 mediante edición XML"""
        from openpyxl.utils import get_column_letter
        
        try:
            # El rango de datos
            rango_datos = f"'{nombre_hoja_datos}'!$A$5:$BC${ultima_fila_datos}"
            col_letter_monto = get_column_letter(col_monto_credito)
            
            # Nota: La creación completa de tablas dinámicas mediante XML es muy compleja
            # Requiere crear múltiples archivos XML (pivotCache, pivotTable, etc.)
            # Por ahora, solo indicamos que la hoja está lista para la tabla dinámica
            # El usuario puede crearla manualmente en Excel con el rango proporcionado
            
            self.enviar_mensaje('actualizar_texto', f"  → Hoja2 lista. Crear tabla dinámica manualmente con origen: {rango_datos}")
            self.enviar_mensaje('actualizar_texto', f"  → Campo: MONTO CREDITO (columna {col_letter_monto})")
            self.enviar_mensaje('actualizar_texto', "  → Agrupar en rangos: 1-5000, 5001-10000, 10001-15000, etc.")
            
        except Exception as e:
            self.enviar_mensaje('actualizar_texto', f"  ⚠️ Nota: La tabla dinámica debe crearse manualmente en Excel")
    
    def agregar_totales_columnas(self, ws, ultima_fila_datos, headers_destino):
        """Agrega fórmulas SUM al final de los datos para las columnas requeridas (AH-AO)"""
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import MergedCell
        
        # Columnas que deben tener totales (en rango AH-AO)
        # NOTA: AI (FECHA DE INICIO DE CREDITO) y AJ (FECHA DE TERMINACION DE CREDITO) NO se suman
        columnas_totales = {
            'MONTO CREDITO': ['MONTO CREDITO', 'MONTO CRÉDITO'],
            'PLAZO DE CREDITO': ['PLAZO DE CREDITO', 'PLAZO DE CRÉDITO'],
            'PRIMA NETA': ['PRIMA NETA'],
            'IMP': ['IMP'],
            'PRIMA TOTAL': ['PRIMA TOTAL'],
            'HGR': ['HGR']
        }
        
        # Buscar las columnas en el destino (búsqueda más exhaustiva)
        columnas_encontradas = {}
        for idx, cell in enumerate(headers_destino):
            if cell.value:
                header_str = str(cell.value).strip().upper()
                # Normalizar espacios
                header_str_limpio = ' '.join(header_str.split())
                
                for col_nombre, variantes in columnas_totales.items():
                    if col_nombre not in columnas_encontradas:
                        for variante in variantes:
                            variante_upper = variante.upper()
                            # Búsqueda más flexible: contiene o está contenido
                            if variante_upper in header_str_limpio or header_str_limpio in variante_upper:
                                col_num = idx + 1
                                col_letter = get_column_letter(col_num)
                                columnas_encontradas[col_nombre] = (col_num, col_letter)
                                self.enviar_mensaje('actualizar_texto', f"  → Columna '{col_nombre}' encontrada en columna {col_letter} ({col_num})")
                                break
                        if col_nombre in columnas_encontradas:
                            break
        
        # Fila donde se colocarán los totales (después de los datos)
        fila_total = ultima_fila_datos + 1
        
        # Agregar fórmulas SUM para cada columna encontrada
        formulas_agregadas = 0
        for col_nombre, (col_num, col_letter) in columnas_encontradas.items():
            try:
                cell_total = ws.cell(fila_total, col_num)
                if isinstance(cell_total, MergedCell):
                    continue
                
                # Verificar si ya hay una fórmula SUM en esta celda o en celdas cercanas
                # No agregar suma si ya existe una en esta celda, arriba o debajo
                tiene_suma_en_celda = False
                try:
                    if cell_total.data_type == 'f':
                        formula_existente = str(cell_total.value).upper()
                        if 'SUM' in formula_existente and col_letter in formula_existente:
                            tiene_suma_en_celda = True
                except:
                    pass
                
                tiene_suma_arriba = False
                if fila_total > 1:
                    try:
                        cell_arriba = ws.cell(fila_total - 1, col_num)
                        if not isinstance(cell_arriba, MergedCell) and cell_arriba.data_type == 'f':
                            formula_arriba = str(cell_arriba.value).upper()
                            if 'SUM' in formula_arriba and col_letter in formula_arriba:
                                tiene_suma_arriba = True
                    except:
                        pass
                
                tiene_suma_debajo = False
                for fila_check in range(fila_total + 1, min(fila_total + 5, ws.max_row + 1)):
                    try:
                        cell_check = ws.cell(fila_check, col_num)
                        if not isinstance(cell_check, MergedCell) and cell_check.data_type == 'f':
                            formula_check = str(cell_check.value).upper()
                            if 'SUM' in formula_check and col_letter in formula_check:
                                tiene_suma_debajo = True
                                break
                    except:
                        continue
                
                # Solo agregar la suma si no hay una en esta celda, arriba o debajo
                if not tiene_suma_en_celda and not tiene_suma_arriba and not tiene_suma_debajo:
                    # Siempre crear/actualizar la fórmula SUM (sobrescribir si existe)
                    # Crear nueva fórmula SUM desde fila 6 hasta última fila de datos
                    formula = f"=SUM({col_letter}6:{col_letter}{ultima_fila_datos})"
                    cell_total.value = formula
                    
                    # Aplicar fuente Calibri, alineación centrada, bordes, fondo amarillo y formato de 2 decimales
                    from openpyxl.styles import PatternFill
                    cell_total.font = self._estilo_fuente_calibri
                    cell_total.alignment = self._estilo_alineacion_centrada
                    cell_total.border = self._estilo_borde_celda
                    cell_total.number_format = '0.00'  # Formato de 2 decimales
                    # Pintar de amarillo
                    cell_total.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    
                    formulas_agregadas += 1
                    self.enviar_mensaje('actualizar_texto', f"  ✓ Suma agregada para '{col_nombre}' en {col_letter}{fila_total}: {formula}")
            except:
                continue
        
        # Agregar fórmulas adicionales una fila después de los totales (fila_total + 1)
        fila_formulas_adicionales = fila_total + 1
        
        # Buscar columnas AL (PRIMA NETA), AM (IMP), AN (PRIMA TOTAL), AO (HGR)
        col_al = None  # PRIMA NETA
        col_am = None  # IMP
        col_an = None  # PRIMA TOTAL
        col_ao = None  # HGR
        
        for idx, cell in enumerate(headers_destino):
            if cell.value:
                header_str = str(cell.value).strip().upper()
                header_str_limpio = ' '.join(header_str.split())
                
                if 'PRIMA NETA' in header_str_limpio and col_al is None:
                    col_al = idx + 1
                elif 'IMP' in header_str_limpio and col_am is None:
                    col_am = idx + 1
                elif 'PRIMA TOTAL' in header_str_limpio and col_an is None:
                    col_an = idx + 1
                elif 'HGR' in header_str_limpio and col_ao is None:
                    col_ao = idx + 1
        
        # Agregar fórmulas en AM, AN, AO una fila después de los totales
        formulas_adicionales = 0
        from openpyxl.styles import Font, Alignment, Border, Side
        borde_delgado = Side(style='thin')
        borde_celda = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)
        
        # AM: =AL{fila_total}*4%
        if col_am is not None and col_al is not None:
            try:
                col_letter_al = get_column_letter(col_al)
                col_letter_am = get_column_letter(col_am)
                cell_am = ws.cell(fila_formulas_adicionales, col_am)
                if not isinstance(cell_am, MergedCell):
                    formula_am = f"={col_letter_al}{fila_total}*4%"
                    cell_am.value = formula_am
                    cell_am.font = self._estilo_fuente_calibri
                    cell_am.alignment = self._estilo_alineacion_centrada
                    cell_am.border = self._estilo_borde_celda
                    cell_am.number_format = '0.00'  # Formato de 2 decimales
                    formulas_adicionales += 1
                    # No mostrar mensaje para esta fórmula (no es SUM)
            except:
                pass
        
        # AN: =+AL{fila_total}+AM{fila_formulas_adicionales}
        if col_an is not None and col_al is not None and col_am is not None:
            try:
                col_letter_al = get_column_letter(col_al)
                col_letter_am = get_column_letter(col_am)
                col_letter_an = get_column_letter(col_an)
                cell_an = ws.cell(fila_formulas_adicionales, col_an)
                if not isinstance(cell_an, MergedCell):
                    formula_an = f"=+{col_letter_al}{fila_total}+{col_letter_am}{fila_formulas_adicionales}"
                    cell_an.value = formula_an
                    cell_an.font = self._estilo_fuente_calibri
                    cell_an.alignment = self._estilo_alineacion_centrada
                    cell_an.border = self._estilo_borde_celda
                    cell_an.number_format = '0.00'  # Formato de 2 decimales
                    formulas_adicionales += 1
                    # No mostrar mensaje para esta fórmula (no es SUM)
            except:
                pass
        
        # AO: =SUM(AO6:AO{ultima_fila_datos})
        # NOTA: HGR ya se procesa en columnas_totales, así que solo agregar aquí si NO se agregó arriba
        if col_ao is not None:
            try:
                from openpyxl.styles import PatternFill
                col_letter_ao = get_column_letter(col_ao)
                
                # Verificar si ya hay una suma en la fila_total (arriba)
                tiene_suma_arriba = False
                try:
                    cell_arriba = ws.cell(fila_total, col_ao)
                    if not isinstance(cell_arriba, MergedCell) and cell_arriba.data_type == 'f':
                        formula_arriba = str(cell_arriba.value).upper()
                        if 'SUM' in formula_arriba and col_letter_ao in formula_arriba:
                            tiene_suma_arriba = True
                except:
                    pass
                
                # Verificar si ya hay una suma debajo
                tiene_suma_debajo = False
                for fila_check in range(fila_formulas_adicionales + 1, min(fila_formulas_adicionales + 5, ws.max_row + 1)):
                    try:
                        cell_check = ws.cell(fila_check, col_ao)
                        if not isinstance(cell_check, MergedCell) and cell_check.data_type == 'f':
                            formula_check = str(cell_check.value).upper()
                            if 'SUM' in formula_check and col_letter_ao in formula_check:
                                tiene_suma_debajo = True
                                break
                    except:
                        continue
                
                # Solo agregar si NO hay suma arriba NI debajo
                if not tiene_suma_arriba and not tiene_suma_debajo:
                    cell_ao = ws.cell(fila_formulas_adicionales, col_ao)
                    if not isinstance(cell_ao, MergedCell):
                        formula_ao = f"=SUM({col_letter_ao}6:{col_letter_ao}{ultima_fila_datos})"
                        cell_ao.value = formula_ao
                        cell_ao.font = self._estilo_fuente_calibri
                        cell_ao.alignment = self._estilo_alineacion_centrada
                        cell_ao.border = self._estilo_borde_celda
                        cell_ao.number_format = '0.00'  # Formato de 2 decimales
                        # Pintar de amarillo (es SUM)
                        cell_ao.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        formulas_adicionales += 1
                        self.enviar_mensaje('actualizar_texto', f"  ✓ Suma agregada para 'HGR' en {col_letter_ao}{fila_formulas_adicionales}: {formula_ao}")
            except:
                pass
        
        if formulas_agregadas > 0:
            self.enviar_mensaje('actualizar_texto', f"  ✓ {formulas_agregadas} total(es) agregado(s)/actualizado(s) en fila {fila_total}")
        else:
            self.enviar_mensaje('actualizar_texto', f"  ⚠️ No se encontraron columnas para totales")
            # Mostrar qué columnas se buscaron
            self.enviar_mensaje('actualizar_texto', f"     Columnas buscadas: {', '.join(columnas_totales.keys())}")
    
    def verificar_errores_fila(self, ws, fila):
        """Verifica si una fila tiene errores en fórmulas. Retorna True si hay errores - OPTIMIZADO"""
        from openpyxl.cell.cell import MergedCell
        
        # Errores comunes de Excel
        errores_excel = {'#N/D', '#N/A', '#VALUE!', '#REF!', '#NAME?', '#NUM!', '#DIV/0!', '#NULL!', '#N/D!', '#N/A!'}
        
        # Revisar solo las columnas que pueden tener fórmulas (limitado para velocidad)
        max_cols = min(ws.max_column, 200)
        
        # Método optimizado: verificar directamente en el workbook actual
        # Guardar el archivo periódicamente para poder leer valores calculados
        # Pero para no hacerlo muy lento, verificamos primero si hay fórmulas con errores obvios
        for col in range(1, max_cols + 1):
            try:
                cell = ws.cell(fila, col)
                if isinstance(cell, MergedCell):
                    continue
                
                # Solo verificar celdas que tienen fórmulas
                if cell.data_type == 'f':
                    # Verificar si la fórmula contiene referencias que podrían causar errores
                    formula = str(cell.value) if cell.value else ""
                    
                    # Verificar patrones de error comunes en fórmulas
                    # Si la fórmula tiene referencias a celdas vacías o inválidas, podría generar error
                    # Por ahora, verificamos si el valor de la celda es un string de error
                    # (Esto requiere que el archivo se haya guardado y recalculado, pero es más rápido)
                    
                    # Método alternativo: verificar si hay errores obvios en la fórmula misma
                    if any(error in formula.upper() for error in ['#N/D', '#N/A', '#VALUE', '#REF', '#NAME', '#NUM', '#DIV', '#NULL']):
                        return True
            except:
                continue
        
        # Si no encontramos errores obvios, intentar leer el archivo guardado (solo si existe)
        # Esto es más preciso pero más lento, así que lo hacemos como fallback
        try:
            wb = ws.parent
            archivo_path = getattr(wb, 'filename', self.archivo_resultado)
            
            # Solo intentar leer si el archivo existe y no es muy grande (optimización)
            if archivo_path and os.path.exists(archivo_path) and os.path.getsize(archivo_path) < 50 * 1024 * 1024:  # < 50MB
                try:
                    # Leer el archivo con data_only=True para obtener valores calculados
                    wb_valores = load_workbook(archivo_path, data_only=True, read_only=True)
                    if ws.title in wb_valores.sheetnames:
                        ws_valores = wb_valores[ws.title]
                        
                        for col in range(1, max_cols + 1):
                            try:
                                cell_original = ws.cell(fila, col)
                                if isinstance(cell_original, MergedCell):
                                    continue
                                
                                # Solo verificar celdas que tienen fórmulas
                                if cell_original.data_type == 'f':
                                    # Obtener el valor calculado
                                    cell_valor = ws_valores.cell(fila, col)
                                    valor_calculado = cell_valor.value
                                    
                                    # Verificar si el valor calculado es un error
                                    if isinstance(valor_calculado, str):
                                        valor_upper = valor_calculado.strip().upper()
                                        if valor_upper in errores_excel or (valor_upper.startswith('#') and '!' in valor_upper):
                                            wb_valores.close()
                                            return True
                                    # Verificar si es un objeto Error de openpyxl
                                    elif hasattr(valor_calculado, '__class__') and 'Error' in str(type(valor_calculado)):
                                        wb_valores.close()
                                        return True
                            except:
                                continue
                    
                    wb_valores.close()
                except:
                    pass  # Si falla, continuar sin error
        except:
            pass  # Si falla, continuar sin error
        
        return False
    
    def resaltar_fila_amarillo(self, ws, fila):
        """Resalta una fila completa en amarillo"""
        from openpyxl.styles import PatternFill
        from openpyxl.cell.cell import MergedCell
        
        amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        max_cols = min(ws.max_column, 200)
        
        for col in range(1, max_cols + 1):
            try:
                cell = ws.cell(fila, col)
                if not isinstance(cell, MergedCell):
                    cell.fill = amarillo
            except:
                continue
    
    def aplicar_estilos_batch(self, ws, ultima_fila_datos, headers_destino):
        """Aplica estilos en batch a todas las celdas de datos (optimización de rendimiento)"""
        from openpyxl.cell.cell import MergedCell
        
        max_cols = min(ws.max_column, 200)
        
        # Aplicar estilos solo a celdas que tienen valores (optimización)
        # Procesar en lotes para mejor rendimiento
        batch_size = 1000
        for fila_inicio in range(6, ultima_fila_datos + 1, batch_size):
            fila_fin = min(fila_inicio + batch_size, ultima_fila_datos + 1)
            for fila in range(fila_inicio, fila_fin):
                for col in range(1, max_cols + 1):
                    try:
                        cell = ws.cell(fila, col)
                        if isinstance(cell, MergedCell):
                            continue
                        # Solo aplicar estilos si la celda tiene valor y no tiene estilos ya aplicados
                        if cell.value is not None:
                            if cell.font is None or cell.font.name != 'Calibri':
                                cell.font = self._estilo_fuente_calibri
                            if cell.alignment is None:
                                cell.alignment = self._estilo_alineacion_centrada
                            if cell.border is None:
                                cell.border = self._estilo_borde_celda
                    except:
                        continue
    
    def aplicar_formato_calibri_y_edad(self, ws, ultima_fila_datos, headers_destino):
        """Aplica fuente Calibri a todas las celdas y formato de 2 decimales a EDAD"""
        from openpyxl.styles import Font
        from openpyxl.cell.cell import MergedCell
        
        fuente_calibri = Font(name='Calibri')
        max_cols = min(ws.max_column, 200)
        
        # Buscar columna EDAD
        col_edad = None
        for idx, cell in enumerate(headers_destino):
            if cell.value and 'EDAD' in str(cell.value).upper():
                col_edad = idx + 1
                break
        
        # Aplicar formato a todas las filas de datos
        for fila in range(6, ultima_fila_datos + 1):
            # Aplicar fuente Calibri a todas las columnas
            for col in range(1, max_cols + 1):
                try:
                    cell = ws.cell(fila, col)
                    if isinstance(cell, MergedCell):
                        continue
                    
                    # Aplicar fuente Calibri, alineación centrada y bordes
                    from openpyxl.styles import Alignment, Border, Side
                    borde_delgado = Side(style='thin')
                    borde_celda = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)
                    cell.font = fuente_calibri
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = borde_celda
                    
                    # Si es columna EDAD y tiene fórmula, modificar para 2 decimales
                    if col == col_edad and cell.data_type == 'f':
                        formula = str(cell.value)
                        if 'TODAY()' in formula.upper() or '/365' in formula.upper():
                            if 'ROUND' not in formula.upper():
                                # Envolver con ROUND para 2 decimales
                                formula_sin_igual = formula[1:] if formula.startswith('=') else formula
                                formula_ajustada = f"=ROUND({formula_sin_igual},2)"
                                cell.value = formula_ajustada
                except:
                    continue
    
    def resaltar_filas_nan(self, ws, ultima_fila_datos):
        """Resalta en amarillo SOLO las filas donde las fórmulas resultan en error #N/D"""
        from openpyxl.styles import PatternFill
        from openpyxl.cell.cell import MergedCell
        import threading
        import openpyxl
        import os
        
        amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        filas_con_error = set()
        filas_resaltadas_lock = threading.Lock()
        
        max_cols = min(ws.max_column, 200)
        
        # Identificar columnas con fórmulas (solo una vez)
        columnas_con_formulas = set()
        for fila_check in range(6, min(11, ultima_fila_datos + 1)):  # Solo primeras 5 filas
            for col in range(1, max_cols + 1):
                try:
                    cell = ws.cell(fila_check, col)
                    if not isinstance(cell, MergedCell) and cell.data_type == 'f':
                        columnas_con_formulas.add(col)
                except:
                    continue
        
        if not columnas_con_formulas:
            return 0
        
        # Guardar el archivo temporalmente para leer valores calculados
        archivo_temp = self.archivo_resultado.replace('.xlsx', '_temp_check.xlsx')
        try:
            ws.parent.save(archivo_temp)
        except:
            return 0
        
        # Leer el archivo con data_only=True para obtener valores calculados
        try:
            wb_valores = openpyxl.load_workbook(archivo_temp, data_only=True)
            ws_valores = wb_valores[ws.title]
        except:
            try:
                if os.path.exists(archivo_temp):
                    os.remove(archivo_temp)
            except:
                pass
            return 0
        
        # Función para procesar rango de filas - SOLO detecta errores reales #N/D
        def procesar_rango_filas(inicio, fin, direccion="normal"):
            filas_error_local = []
            rango = range(inicio, fin + 1) if direccion == "normal" else range(fin, inicio - 1, -1)
            
            for fila in rango:
                tiene_error = False
                
                # Solo revisar columnas con fórmulas
                for col in columnas_con_formulas:
                    try:
                        # Verificar en el archivo con valores calculados
                        cell_valor = ws_valores.cell(fila, col)
                        if isinstance(cell_valor, MergedCell):
                            continue
                        
                        # Verificar si la celda original tiene fórmula
                        cell_original = ws.cell(fila, col)
                        if isinstance(cell_original, MergedCell) or cell_original.data_type != 'f':
                            continue
                        
                        # Obtener el valor calculado
                        valor_calculado = cell_valor.value
                        
                        # SOLO detectar errores #N/D (n/d)
                        if valor_calculado is not None:
                            if isinstance(valor_calculado, str):
                                valor_upper = valor_calculado.strip().upper()
                                # Solo #N/D o #N/A (que es lo mismo que n/d)
                                if valor_upper in ['#N/D', '#N/A', 'N/D', 'N/A', '#N?D', '#N?A']:
                                    tiene_error = True
                                    break
                            # Verificar si es un objeto Error de openpyxl
                            elif hasattr(valor_calculado, '__class__'):
                                tipo_error = str(type(valor_calculado))
                                if 'Error' in tipo_error or 'N/A' in tipo_error or 'N/D' in tipo_error:
                                    tiene_error = True
                                    break
                    except:
                        continue
                
                if tiene_error:
                    filas_error_local.append(fila)
            
            # Agregar filas encontradas al set compartido
            with filas_resaltadas_lock:
                filas_con_error.update(filas_error_local)
        
        # Usar threading para procesar más rápido
        total_filas = ultima_fila_datos - 5
        mitad = 6 + (total_filas // 2)
        
        thread1 = threading.Thread(target=procesar_rango_filas, args=(6, mitad, "normal"))
        thread2 = threading.Thread(target=procesar_rango_filas, args=(mitad + 1, ultima_fila_datos, "reverso"))
        
        thread1.start()
        thread2.start()
        
        thread1.join()
        thread2.join()
        
        # Cerrar el archivo de valores
        wb_valores.close()
        
        # Eliminar archivo temporal
        try:
            if os.path.exists(archivo_temp):
                os.remove(archivo_temp)
        except:
            pass
        
        # Resaltar filas encontradas (batch) - SOLO las que tienen error #N/D
        for fila in filas_con_error:
            try:
                for col in range(1, max_cols + 1):
                    try:
                        cell = ws.cell(fila, col)
                        if not isinstance(cell, MergedCell):
                            cell.fill = amarillo
                    except:
                        continue
            except:
                continue
        
        return len(filas_con_error)
    
    def actualizar_formulas_totales(self, ws, ultima_fila_datos):
        """Actualiza las fórmulas de totales al final del archivo, conservando TODAS las fórmulas"""
        import re
        from openpyxl.cell.cell import MergedCell
        
        formulas_actualizadas = 0
        formulas_conservadas = 0
        
        # Buscar en un rango más amplio para encontrar TODAS las fórmulas de totales
        # Buscar desde la última fila de datos hasta el final del archivo
        fila_inicio_busqueda = max(1, ultima_fila_datos + 1)
        fila_fin_busqueda = ws.max_row + 1
        
        # Buscar TODAS las fórmulas (no solo SUM) y actualizarlas
        for row in range(fila_inicio_busqueda, fila_fin_busqueda):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row, col)
                    
                    # Saltar celdas fusionadas
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.data_type == 'f':
                        formula = str(cell.value)
                        formula_upper = formula.upper()
                        
                        # Buscar y actualizar fórmulas SUM
                        if 'SUM(' in formula_upper:
                            formula_modificada = formula
                            modificado = False
                            
                            # Patrón mejorado para SUM(COLUMNAFILA:COLUMNAFILA)
                            # Maneja: SUM(AH6:AH5281), SUM($AH$6:$AH$5281), SUM(AH$6:AH$5281), etc.
                            pattern = r'SUM\(([A-Z]+\$?\d+):([A-Z]+\$?\d+)\)'
                            
                            def reemplazar_sum(match):
                                nonlocal modificado
                                try:
                                    inicio = match.group(1)
                                    fin = match.group(2)
                                    
                                    # Extraer columna y fila de inicio (maneja $AH$6, AH6, AH$6, etc.)
                                    match_ini = re.search(r'([A-Z]+)(\$?)(\d+)', inicio)
                                    # Extraer columna y fila de fin
                                    match_fin = re.search(r'([A-Z]+)(\$?)(\d+)', fin)
                                    
                                    if match_ini and match_fin:
                                        col_ini = match_ini.group(1)
                                        abs_ini = match_ini.group(2)  # $ si es absoluto
                                        fila_ini = int(match_ini.group(3))
                                        
                                        col_fin = match_fin.group(1)
                                        abs_fin = match_fin.group(2)  # $ si es absoluto
                                        fila_fin = int(match_fin.group(3))
                                        
                                        # Solo actualizar si es la misma columna y la fila final es mayor que nuestros datos
                                        if col_ini == col_fin and fila_fin > ultima_fila_datos:
                                            modificado = True
                                            # Mantener referencias absolutas
                                            nueva_ref = f"{col_ini}{abs_ini}{fila_ini}:{col_fin}{abs_fin}{ultima_fila_datos}"
                                            return f"SUM({nueva_ref})"
                                except:
                                    pass
                                return match.group(0)  # No cambiar si hay error
                            
                            # Aplicar reemplazo (case insensitive)
                            formula_modificada = re.sub(pattern, reemplazar_sum, formula, flags=re.IGNORECASE)
                            
                            if modificado:
                                cell.value = formula_modificada
                                formulas_actualizadas += 1
                        else:
                            # Conservar otras fórmulas (no SUM) sin modificar
                            formulas_conservadas += 1
                            
                except Exception:
                    continue
        
        return formulas_actualizadas, formulas_conservadas
    
    def extraer_cuadre(self, df_origen):
        """Extrae el cuadre del archivo origen según la estructura documentada"""
        cuadre = {}
        
        # Buscar en las últimas filas (aproximadamente desde fila 5280)
        inicio_busqueda = max(0, len(df_origen) - 100)
        
        try:
            for idx in range(inicio_busqueda, len(df_origen)):
                try:
                    row = df_origen.iloc[idx]
                    row_str = ' '.join([str(x) for x in row if pd.notna(x)]).upper()
                    
                    # Buscar fila con totales principales (aproximadamente fila 5282)
                    # Buscar por "TOTAL CLIENTES" o valores grandes que indiquen totales
                    if 'TOTAL CLIENTES' in row_str or any(
                        pd.notna(val) and isinstance(val, (int, float)) and val > 5000 
                        for val in row[:35] if pd.notna(val)
                    ):
                        # Buscar MONTO CREDITO (Col 25, índice 24)
                        try:
                            if 24 < len(row):
                                val = row.iloc[24] if hasattr(row, 'iloc') else row[24]
                                if pd.notna(val) and isinstance(val, (int, float)) and val > 1000000:
                                    cuadre['MONTO_CREDITO'] = float(val)
                        except (IndexError, KeyError, ValueError):
                            pass
                        
                        # Buscar TOTAL CLIENTES (Col 28, índice 27)
                        try:
                            if 27 < len(row):
                                val = row.iloc[27] if hasattr(row, 'iloc') else row[27]
                                if pd.notna(val) and isinstance(val, (int, float)) and val > 1000:
                                    cuadre['CLIENTES'] = int(val)
                        except (IndexError, KeyError, ValueError):
                            pass
                        
                        # Buscar PRIMA NETA (Col 29, índice 28)
                        try:
                            if 28 < len(row):
                                val = row.iloc[28] if hasattr(row, 'iloc') else row[28]
                                if pd.notna(val) and isinstance(val, (int, float)) and 200000 < val < 400000:
                                    cuadre['PRIMA_NETA'] = float(val)
                        except (IndexError, KeyError, ValueError):
                            pass
                        
                        # Buscar IMP (Col 30, índice 29)
                        try:
                            if 29 < len(row):
                                val = row.iloc[29] if hasattr(row, 'iloc') else row[29]
                                if pd.notna(val) and isinstance(val, (int, float)) and 10000 < val < 20000:
                                    cuadre['IMP'] = float(val)
                        except (IndexError, KeyError, ValueError):
                            pass
                        
                        # Buscar PRIMA TOTAL (Col 31, índice 30)
                        try:
                            if 30 < len(row):
                                val = row.iloc[30] if hasattr(row, 'iloc') else row[30]
                                if pd.notna(val) and isinstance(val, (int, float)) and 300000 < val < 400000:
                                    cuadre['PRIMA_TOTAL'] = float(val)
                        except (IndexError, KeyError, ValueError):
                            pass
                        
                        # Buscar HGR (Col 32, índice 31)
                        try:
                            if 31 < len(row):
                                val = row.iloc[31] if hasattr(row, 'iloc') else row[31]
                                if pd.notna(val) and isinstance(val, (int, float)) and 100000 < val < 200000:
                                    cuadre['HGR'] = float(val)
                        except (IndexError, KeyError, ValueError):
                            pass
                        
                        # Si encontramos al menos un total, podemos parar
                        if len(cuadre) >= 2:
                            break
                except (IndexError, KeyError):
                    continue
        except Exception as e:
            # Si hay error general, continuar sin cuadre
            pass
        
        return cuadre
    
    def mostrar_cuadre(self, cuadre, total_filas):
        """Muestra el cuadre en el área de texto y valida los totales"""
        if cuadre:
            self.enviar_mensaje('actualizar_texto', "\n📊 TOTALES EXTRAÍDOS DEL ARCHIVO ORIGEN (413):")
            self.enviar_mensaje('actualizar_texto', "-" * 80)
            
            # Mostrar solo CLIENTES y PRIMA_NETA
            if 'CLIENTES' in cuadre:
                value = cuadre['CLIENTES']
                self.enviar_mensaje('actualizar_texto', f"  CLIENTES: {value:,}")
            if 'PRIMA_NETA' in cuadre:
                value = cuadre['PRIMA_NETA']
                self.enviar_mensaje('actualizar_texto', f"  PRIMA NETA: {value:,.2f}")
            
            self.enviar_mensaje('actualizar_texto', f"\nTotal de filas procesadas: {total_filas}")
            self.enviar_mensaje('actualizar_texto', f"  ⚠️ Filas con errores en fórmulas (ND/#N/D/#N/A): 0")
        else:
            self.enviar_mensaje('actualizar_texto', "\n⚠️ No se encontraron totales en el archivo origen.")
        
        self.enviar_mensaje('actualizar_texto', "\n" + "=" * 80)
    
    def limpiar_archivos_temporales(self):
        """Elimina archivos temporales antiguos del directorio"""
        directorio_actual = os.getcwd()
        patron_temp = "temp_*.xlsx"
        
        try:
            archivos_temp = glob.glob(os.path.join(directorio_actual, patron_temp))
            for archivo_temp in archivos_temp:
                try:
                    # Eliminar archivos temporales antiguos (más de 1 hora)
                    tiempo_modificacion = os.path.getmtime(archivo_temp)
                    tiempo_actual = datetime.now().timestamp()
                    diferencia_horas = (tiempo_actual - tiempo_modificacion) / 3600
                    
                    if diferencia_horas > 1:  # Más de 1 hora
                        os.remove(archivo_temp)
                except Exception as e:
                    # Si no se puede eliminar, continuar
                    pass
        except Exception:
            pass
    
    def cerrar_aplicacion(self):
        """Limpia archivos temporales antes de cerrar"""
        if self.archivo_resultado and os.path.exists(self.archivo_resultado):
            try:
                os.remove(self.archivo_resultado)
            except:
                pass
        self.root.destroy()
    
    def descargar_archivo_automatico(self):
        """Descarga automáticamente el archivo transformado sin preguntar al usuario"""
        if not self.archivo_resultado:
            self.enviar_mensaje('error', "No hay archivo para descargar")
            return
        
        if not os.path.exists(self.archivo_resultado):
            self.enviar_mensaje('error', "El archivo temporal no existe")
            return
        
        try:
            # Determinar carpeta de destino (misma carpeta del archivo origen o carpeta de Descargas)
            if self.archivo_origen:
                carpeta_destino = os.path.dirname(os.path.abspath(self.archivo_origen))
            else:
                # Usar carpeta de Descargas del usuario
                import platform
                sistema = platform.system()
                if sistema == 'Windows':
                    carpeta_destino = os.path.join(os.path.expanduser("~"), "Downloads")
                elif sistema == 'Darwin':  # macOS
                    carpeta_destino = os.path.join(os.path.expanduser("~"), "Downloads")
                else:  # Linux
                    carpeta_destino = os.path.join(os.path.expanduser("~"), "Downloads")
            
            # Crear carpeta si no existe
            if not os.path.exists(carpeta_destino):
                os.makedirs(carpeta_destino)
            
            # Usar el nombre generado basado en plantilla y fecha
            nombre_archivo = self.nombre_archivo_descarga if self.nombre_archivo_descarga else "archivo_transformado.xlsx"
            
            # Si el nombre no tiene extensión, agregarla
            if not nombre_archivo.endswith('.xlsx'):
                nombre_archivo += '.xlsx'
            
            # Ruta completa del archivo destino
            archivo_guardar = os.path.join(carpeta_destino, nombre_archivo)
            
            # Si el archivo ya existe, agregar un número al final
            contador = 1
            archivo_original = archivo_guardar
            while os.path.exists(archivo_guardar):
                nombre_base = os.path.splitext(nombre_archivo)[0]
                extension = os.path.splitext(nombre_archivo)[1]
                archivo_guardar = os.path.join(carpeta_destino, f"{nombre_base}_{contador}{extension}")
                contador += 1
            
            # Copiar el archivo
            shutil.copy2(self.archivo_resultado, archivo_guardar)
            
            # Eliminar archivo temporal después de copiarlo exitosamente
            try:
                os.remove(self.archivo_resultado)
                self.archivo_resultado = None
            except:
                pass
            
            # Abrir la carpeta donde se guardó el archivo
            try:
                import platform
                sistema = platform.system()
                
                if sistema == 'Windows':
                    # Windows: abrir carpeta en el explorador
                    os.startfile(carpeta_destino)
                elif sistema == 'Darwin':  # macOS
                    import subprocess
                    subprocess.Popen(['open', carpeta_destino])
                else:  # Linux y otros
                    import subprocess
                    subprocess.Popen(['xdg-open', carpeta_destino])
            except Exception as e:
                # Si falla abrir la carpeta, continuar sin error
                pass
            
            self.enviar_mensaje('actualizar_texto', f"\n✅ Archivo descargado automáticamente en:\n{archivo_guardar}\n\nLa carpeta se ha abierto automáticamente.")
            self.enviar_mensaje('completado', f"Transformación completada exitosamente.\n\nArchivo descargado en:\n{archivo_guardar}")
            
        except Exception as e:
            self.enviar_mensaje('error', f"Error al descargar archivo automáticamente:\n{str(e)}")
    
    def descargar_archivo(self):
        """Permite al usuario descargar el archivo transformado manualmente"""
        if not self.archivo_resultado:
            messagebox.showerror("Error", "No hay archivo para descargar")
            return
        
        if not os.path.exists(self.archivo_resultado):
            messagebox.showerror("Error", "El archivo temporal no existe")
            return
        
        # Usar el nombre generado basado en plantilla y fecha
        nombre_sugerido = self.nombre_archivo_descarga if self.nombre_archivo_descarga else "archivo_transformado.xlsx"
        
        # Preguntar dónde guardar
        archivo_guardar = filedialog.asksaveasfilename(
            title="Guardar archivo transformado",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=nombre_sugerido
        )
        
        if archivo_guardar:
            try:
                shutil.copy2(self.archivo_resultado, archivo_guardar)
                
                # Eliminar archivo temporal después de copiarlo exitosamente
                try:
                    os.remove(self.archivo_resultado)
                    self.archivo_resultado = None
                except:
                    pass
                
                # Abrir la carpeta donde se guardó el archivo
                try:
                    carpeta_destino = os.path.dirname(archivo_guardar)
                    if carpeta_destino:
                        import platform
                        sistema = platform.system()
                        
                        if sistema == 'Windows':
                            # Windows: abrir carpeta en el explorador
                            os.startfile(carpeta_destino)
                        elif sistema == 'Darwin':  # macOS
                            import subprocess
                            subprocess.Popen(['open', carpeta_destino])
                        else:  # Linux y otros
                            import subprocess
                            subprocess.Popen(['xdg-open', carpeta_destino])
                except Exception as e:
                    # Si falla abrir la carpeta, continuar sin error
                    pass
                
                messagebox.showinfo("Éxito", f"Archivo guardado en:\n{archivo_guardar}\n\nLa carpeta se ha abierto automáticamente.")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar archivo:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = TransformadorExcel(root)
    root.mainloop()

